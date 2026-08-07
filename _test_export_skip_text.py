# -*- coding: utf-8 -*-
"""导出：脚注文本不导出；质检拒标但有数值的二维表仍导出。"""
import os
import tempfile

from codes.pdf_extractor.table_export_filter import is_exportable_table
from codes.pdf_extractor import ExcelExporter


def test_text_footnote_not_exportable():
    t = {
        "type": "text",
        "parse_status": "success",
        "is_real_table": False,
        "quality_decision": "rejected",
        "table_category": "非表格",
        "data": "1.上表中各项数据均为最近一个季度内92个自然日的简单算术平均值。",
        "page": 39,
    }
    assert not is_exportable_table(t)


def test_rejected_but_numeric_grid_still_exportable():
    """质检标 rejected/非表格，但多行列含数值 → 仍应导出。"""
    t = {
        "type": "table",
        "parse_status": "success",
        "is_real_table": False,
        "quality_decision": "rejected",
        "table_category": "非表格",
        "data": [
            ["项目", "金额", "比例"],
            ["现金", "100", "10%"],
            ["存款", "200", "20%"],
        ],
        "page": 10,
    }
    assert is_exportable_table(t)


def test_real_table_exportable():
    t = {
        "type": "table",
        "parse_status": "success",
        "is_real_table": True,
        "quality_decision": "accepted",
        "table_category": "财务数据表",
        "data": [["项目", "金额"], ["现金", "100"]],
        "page": 1,
    }
    assert is_exportable_table(t)


def test_single_col_char_soup_not_exportable():
    t = {
        "type": "table",
        "parse_status": "success",
        "data": [["1"], ["."], ["上"], ["表"], ["中"]],
        "page": 39,
    }
    assert not is_exportable_table(t)


def test_export_tables_skips_text_keeps_rejected_numeric():
    footnote = "1.上表中各项数据均为最近一个季度内92个自然日。"
    tables = [
        {
            "type": "text",
            "parse_status": "success",
            "is_real_table": False,
            "quality_decision": "rejected",
            "extractor": "table_engine",
            "page": 39,
            "data": footnote,
            "title": "脚注",
        },
        {
            "type": "table",
            "parse_status": "success",
            "is_real_table": False,
            "quality_decision": "rejected",
            "table_category": "非表格",
            "extractor": "table_engine",
            "page": 40,
            "data": [["项目", "金额"], ["现金", "100"]],
            "title": "被拒真表",
        },
    ]
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "out.xlsx")
        assert ExcelExporter.export_tables(tables, path)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        flat = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=40, max_col=6, values_only=True):
                for v in row:
                    if v is not None and str(v).strip():
                        flat.append(str(v).strip())
        assert any("100" in v for v in flat), flat
        assert not any(v == "上" for v in flat), flat


if __name__ == "__main__":
    test_text_footnote_not_exportable()
    test_rejected_but_numeric_grid_still_exportable()
    test_real_table_exportable()
    test_single_col_char_soup_not_exportable()
    test_export_tables_skips_text_keeps_rejected_numeric()
    print("OK")

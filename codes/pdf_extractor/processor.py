# -*- coding: utf-8 -*-
"""
处理模块 - PDF处理、LLM识别、Excel导出、工作线程
"""

import os
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

from PyQt5.QtCore import QThread, pyqtSignal

from .utils import (
    load_config, TEMP_DIR
)
from .pdf_context import PDFContext

# 提前导入 openpyxl，避免运行时首次导入时 GC 触发 C 扩展 refcount bug
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# V3 架构修复：统一去重引擎
from codes.table_validator.dedup_engine import DeduplicationEngine, DedupPolicy
from codes.table_validator.page_layout_model import PageLayoutModel
from codes.table_validator.table_block_decider import TableBlockDecider, decide_table_blocks


# ============================================================
# 混合分割辅助函数
# ============================================================


def _extract_paragraphs_for_hybrid(
    liteparse_data: dict,
    hybrid_tables: list,
) -> list:
    """为混合表格提取未被覆盖的 liteparse text_items 段落。

    混合表格使用 liteparse 行列重建，用表格/region 的 Y+X 范围判断覆盖。
    表格内容不得重复落入段落（与 hybrid 产出的 table 条目互斥）。

    Args:
        liteparse_data: ParseResult.to_dict()
        hybrid_tables: hybrid_segment_tables 输出的表格列表

    Returns:
        段落条目列表（含 _source_item_indices 用于精确去重）
    """
    pages = liteparse_data.get("pages", [])
    if not pages:
        return []

    from codes.table_validator.liteparse_table_segmenter import _build_items
    from codes.table_validator.cell_differ import _normalize_for_search
    import re

    # 按页组织覆盖矩形 (y0, y1, x0, x1)
    _layout_cache: dict = {}
    _Y_MARGIN = 25.0
    for t in hybrid_tables:
        p = t.get("page", 0)
        if p not in _layout_cache:
            items = t.get("text_items", [])
            if items:
                _layout_cache[p] = PageLayoutModel.from_text_items(items, page_num=p)
    if _layout_cache:
        _Y_MARGIN = sum(m.table_y_margin for m in _layout_cache.values()) / len(_layout_cache)

    coverage_by_page: dict = {}

    def _add_coverage(page_num: int, y0: float, y1: float, x0: float, x1: float) -> None:
        if y1 <= y0:
            return
        coverage_by_page.setdefault(page_num, []).append((
            y0 - _Y_MARGIN,
            y1 + _Y_MARGIN,
            max(0.0, x0 - 8.0),
            x1 + 8.0,
        ))

    # ① liteparse table_regions：覆盖整段检测区域（子表 y0 偏小时仍算已覆盖）
    for lp_page in pages:
        page_num = lp_page.get("page_number", 0)
        for region in lp_page.get("table_regions", []):
            if region.get("confidence", 0) < 0.3:
                continue
            _add_coverage(
                page_num,
                float(region.get("y0", 0)),
                float(region.get("y1", 0)),
                float(region.get("x0", 0)),
                float(region.get("x1", 9999)),
            )

    # ② hybrid 条目 bbox（text 类型跳过）
    for t in hybrid_tables:
        if t.get("type") in ("text", "paragraph", "annotation"):
            continue
        p = t.get("page", 0)
        y0 = float(t.get("y0", 0) or 0)
        y1 = float(t.get("y1", 0) or 0)
        if y0 <= 0 and y1 <= 0:
            continue
        x0 = float(t.get("x0", 0) or 0)
        x1 = float(t.get("x1", 0) or 0)
        if x1 <= x0:
            x0, x1 = 0.0, 9999.0
        _add_coverage(p, y0, y1, x0, x1)

    # ③ 各页表格单元格文本指纹（内容去重安全网）
    table_tokens_by_page: dict = {}
    for t in hybrid_tables:
        if t.get("type") in ("text", "paragraph", "annotation"):
            continue
        p = t.get("page", 0)
        tokens = table_tokens_by_page.setdefault(p, set())
        for row in t.get("data", []):
            if not isinstance(row, list):
                continue
            for cell in row:
                cs = str(cell).strip()
                if len(cs) >= 2:
                    tokens.add(_normalize_for_search(cs))
                for part in re.findall(r"[\u4e00-\u9fff]{2,}|\d[\d,\.]+", cs):
                    if len(part) >= 2:
                        tokens.add(_normalize_for_search(part))

    def _is_covered(itx0, ity0, itx1, ity1, p):
        """text_item 中心点落在任一覆盖矩形内 → 已归属表格/region。"""
        if p not in coverage_by_page:
            return False
        cx = (itx0 + itx1) / 2
        cy = (ity0 + ity1) / 2
        for cy0, cy1, cx0, cx1 in coverage_by_page[p]:
            if cx0 <= cx <= cx1 and cy0 <= cy <= cy1:
                return True
        return False

    def _text_duplicates_table_cells(text: str, page_num: int) -> bool:
        """段落正文与当页表格单元格高度重合 → 视为表格重复，不输出为文本。"""
        tokens = table_tokens_by_page.get(page_num)
        if not tokens:
            return False
        raw_parts = re.findall(r"[\u4e00-\u9fff]{2,}|\d[\d,\.]{2,}", text)
        if not raw_parts:
            return False
        norm = [_normalize_for_search(p) for p in raw_parts if len(p) >= 2]
        if len(norm) < 2:
            return False
        hit = sum(1 for p in norm if p in tokens)
        if hit >= 3 and hit / len(norm) >= 0.45:
            return True
        # 典型表头/数据行特征
        if hit >= 2 and any(k in text for k in ("期数", "占比", "百万元", "12月31日")):
            if sum(1 for c in text if c.isdigit()) >= 4:
                return True
        return False

    paragraphs = []
    pure_text_pages = {
        int(r.get("page", 0) or 0)
        for r in hybrid_tables
        if r.get("_is_pure_text_page")
    }
    for lp_page in pages:
        page_num = lp_page.get("page_number", 0)
        if page_num in pure_text_pages:
            continue
        regions = lp_page.get("table_regions") or []
        if not any(float(r.get("confidence", 0) or 0) >= 0.3 for r in regions):
            continue
        text_items_raw = lp_page.get("text_items", [])
        if not text_items_raw:
            continue

        indexed_items = _build_items(text_items_raw, page_num)

        orphan_items = []
        for it in indexed_items:
            text = it.get("text", "").strip()
            if not text or len(text) < 2:
                continue
            x0 = it.get("x0", 0)
            y0 = it.get("y0", 0)
            x1 = it.get("x1", 0)
            y1 = it.get("y1", 0)
            if _is_covered(x0, y0, x1, y1, page_num):
                continue
            orphan_items.append(it)

        if not orphan_items:
            continue

        orphan_items.sort(key=lambda it: it["y_mid"])
        blocks = _cluster_orphans_to_text_blocks(orphan_items)

        for block in blocks:
            text = block["text"].strip()
            if len(text) < 5:
                continue
            if _text_duplicates_table_cells(text, page_num):
                continue
            paragraphs.append({
                "page": page_num,
                "type": "paragraph",
                "data": text,
                "text": text,
                "y0": block.get("y0", 0),
                "y1": block.get("y1", 0),
                "extractor": "liteparse_hybrid",
                "confidence": 0.5,
                "rows": block.get("line_count", 1),
                "cols": 1,
                "bbox": [
                    round(block["x0"], 2),
                    round(block["y0"], 2),
                    round(block["x1"], 2),
                    round(block["y1"], 2),
                ],
                "_source_item_indices": block.get("_source_item_indices", []),
            })

    return paragraphs


def _mark_page_types(results: list) -> None:
    """为每个结果项添加 page_type 标记：纯表格 or 半表格。

    按页分析 results 中的 type 字段，判定每页的文档类型：
    - "pure_table"：此页仅有表格，无段落/注解
    - "mixed"：此页同时包含表格和段落/注解

    标记直接写入每个结果 dict 的 page_type 字段。
    """
    if not results:
        return

    # 按页分组统计
    from collections import defaultdict
    page_stats: dict = defaultdict(lambda: {"has_table": False, "has_paragraph": False})

    for r in results:
        p = r.get("page", 0)
        t = r.get("type", "")
        stats = page_stats[p]
        if t == "table":
            stats["has_table"] = True
        elif t in ("paragraph", "annotation", "text"):
            stats["has_paragraph"] = True

    # 计算每页类型
    for p, stats in page_stats.items():
        if stats["has_table"] and not stats["has_paragraph"]:
            page_type = "pure_table"
        elif stats["has_paragraph"] and not stats["has_table"]:
            page_type = "pure_text"
        else:
            page_type = "mixed"

        # 写回该页所有结果项
        for r in results:
            if r.get("page") == p:
                r["page_type"] = page_type


def _cluster_orphans_to_text_blocks(orphan_items):
    """将孤儿 text_items 按 Y 间距聚类为文本块。"""
    if not orphan_items:
        return []

    import re
    # 使用与 liteparse_table_segmenter 相同的聚类逻辑
    Y_GAP_THRESHOLD = 15.0

    blocks = []
    current = [orphan_items[0]]
    prev = orphan_items[0]

    for it in orphan_items[1:]:
        gap = it["y_mid"] - prev["y_mid"]
        if gap < Y_GAP_THRESHOLD:
            current.append(it)
        else:
            blocks.append(_build_text_block(current))
            current = [it]
        prev = it

    if current:
        blocks.append(_build_text_block(current))

    return blocks


def _build_text_block(items):
    """构建文本块 dict（含 _source_item_indices 用于精确去重）。"""
    if not items:
        return {"text": "", "x0": 0, "y0": 0, "x1": 0, "y1": 0, "line_count": 0}

    items_sorted = sorted(items, key=lambda it: (it["y_mid"], it["x0"]))
    x0 = min(it["x0"] for it in items)
    y0 = min(it["y0"] for it in items)
    x1 = max(it["x1"] for it in items)
    y1 = max(it["y1"] for it in items)

    # 按行去重
    lines = {}
    for it in items:
        y_key = round(it["y_mid"], 1)
        if y_key not in lines:
            lines[y_key] = []
        lines[y_key].append(it)

    # 每行按 X 排序拼接
    line_texts = []
    for y_key in sorted(lines.keys()):
        line_items = sorted(lines[y_key], key=lambda it: it["x0"])
        line_text = " ".join(it["text"] for it in line_items)
        line_texts.append(line_text)

    # 用换行连接各行（保持原文格式）
    text_block = "\n".join(line_texts)

    return {
        "text": text_block,
        "x0": x0, "y0": y0,
        "x1": x1, "y1": y1,
        "line_count": len(lines),
        "_source_item_indices": [
            it["item_index"] for it in items_sorted
            if it.get("item_index")
        ],
    }


# ============================================================
# 跨表去重（自动流水线用）
# ============================================================


def _sort_y_for_page_order(item: dict) -> float:
    """提取条目的页内 Y 坐标用于排序，保证表格和段落的正确数据顺序。

    表格使用 y0 字段，段落/注解使用 bbox[1]（top 坐标）。
    failed 类型的条目（无坐标）排在最后（返回大值）。
    """
    if item.get("type") in ("paragraph", "annotation"):
        bbox = item.get("bbox")
        if isinstance(bbox, (list, tuple)) and len(bbox) >= 2:
            return float(bbox[1])
        return float(item.get("y0", 0))
    # 表格：优先 y0，降级到 bbox
    y = item.get("y0")
    if y is not None and y > 0:
        return float(y)
    bbox = item.get("bbox")
    if isinstance(bbox, (list, tuple)) and len(bbox) >= 2:
        return float(bbox[1])
    return 0.0


def _verify_results_ordering(results: list) -> None:
    """验证 results 列表按页码+页内 Y 坐标严格升序排列。

    发现乱序时打印警告但不断言失败（不阻断主流程）。
    """
    if len(results) < 2:
        return

    violations = []
    for i in range(len(results) - 1):
        curr, nxt = results[i], results[i + 1]
        curr_page = curr.get("page", 0)
        nxt_page = nxt.get("page", 0)

        if curr_page > nxt_page:
            violations.append(
                f"  页码倒退: #{i}(P{curr_page}) → #{i+1}(P{nxt_page})"
            )
        elif curr_page == nxt_page:
            curr_y = _sort_y_for_page_order(curr)
            nxt_y = _sort_y_for_page_order(nxt)
            if curr_y > nxt_y + 0.1:  # 允许微小浮点误差
                violations.append(
                    f"  P{curr_page} 同页 Y 倒退: #{i}(y={curr_y:.1f} type={curr.get('type','?')})"
                    f" → #{i+1}(y={nxt_y:.1f} type={nxt.get('type','?')})"
                )

    if violations:
        print(f"  [顺序警告] 发现 {len(violations)} 处顺序异常:")
        for v in violations:
            print(v)
    else:
        passed = sum(1 for r in results if r.get("type") not in ("failed",))
        print(f"  [顺序验证] {len(results)} 个条目顺序正确（{passed} 个有效条目）")


def _deduplicate_text_against_tables(results: list) -> list:
    """表格→文本去重：已被表格内容覆盖的段落/注解不应重复出现在结果中。

    三层策略（按优先级）：
    - Tier 1（精确）：段落有 _source_item_indices 且表格有 text_items →
      直接用 item_index 集合差集判定。零误杀。
    - Tier 2（回退1）：段落无 _source_item_indices 但表格有 text_items →
      收集所有已占用 item_index，用 bbox 辅助判定。
    - Tier 3（回退2）：表格无 text_items（hybrid）→ bbox 空间重叠 + token
      内容重叠（保留原逻辑，阈值保守）。
    """
    tables = [r for r in results if r.get("type") not in ("paragraph", "annotation", "failed")]
    text_entries = [r for r in results if r.get("type") in ("paragraph", "annotation")]

    if not text_entries:
        return results

    # ===== Tier 1 & 2 准备：按页收集表格已占用的 item_index =====
    table_assigned_indices = {}  # {page: set of item_index}
    table_y_ranges = {}          # {page: [(y0, y1, x0, x1), ...]} (Tier 3 fallback)
    table_tokens = {}            # {page: set of normalized tokens} (Tier 3 fallback)

    tables_with_text_items = 0
    tables_without_text_items = 0

    for t in tables:
        pg = t.get("page", 0)
        data = t.get("data", [])

        # 尝试收集 item_index（Tier 1/2）
        text_items = t.get("text_items", [])
        if text_items:
            tables_with_text_items += 1
            indices = set()
            for it in text_items:
                idx = it.get("item_index", 0)
                if idx:
                    indices.add(idx)
            table_assigned_indices.setdefault(pg, set()).update(indices)
        else:
            tables_without_text_items += 1

        # Tier 3 fallback: 构建 token 和 bbox 索引
        if data:
            tokens = set()
            for row in data:
                if isinstance(row, list):
                    for cell in row:
                        if cell:
                            tokens.update(_tokenize_cell(str(cell)))
                elif isinstance(row, str):
                    tokens.update(_tokenize_cell(str(row)))
            table_tokens.setdefault(pg, set()).update(tokens)

            y0 = t.get("y0")
            y1 = t.get("y1")
            bbox = t.get("bbox")
            if (y0 is None or y0 <= 0) and bbox:
                y0 = bbox[1] if len(bbox) >= 2 else 0
                y1 = bbox[3] if len(bbox) >= 4 else 0
            if y0 and y1 and y0 > 0 and y1 > 0:
                x0 = t.get("x0", 0) or (bbox[0] if bbox else 0)
                x1_pad = t.get("x1", 9999) or 9999
                table_y_ranges.setdefault(pg, []).append((y0, y1, x0, x1_pad))

    # ===== 逐条检查 =====
    keep_results = []
    removed_count = 0
    index_removed = 0
    fallback_removed = 0

    for r in results:
        typ = r.get("type")
        if typ not in ("paragraph", "annotation", "text"):
            keep_results.append(r)
            continue

        pg = r.get("page", 0)
        should_remove = False

        # ── Tier 1: 精确 item_index 去重 ──
        source_indices = r.get("_source_item_indices")
        if source_indices and pg in table_assigned_indices:
            assigned = table_assigned_indices[pg]
            if any(idx in assigned for idx in source_indices):
                should_remove = True
                index_removed += 1

        # ── Tier 2 & 3: 回退到 bbox / token 去重 ──
        if not should_remove:
            text_data = r.get("data", "")
            if isinstance(text_data, list):
                text_data = " ".join(str(c) for c in text_data if c)
            text_data = str(text_data).strip()
            if not text_data:
                keep_results.append(r)
                continue

            # --- 空间重叠（paragraph Y >70% 被 table 覆盖）---
            spatial_overlap = False
            para_bbox = r.get("bbox")
            if para_bbox and len(para_bbox) >= 4:
                para_y0 = para_bbox[1]
                para_y1 = para_bbox[3]
                para_x0 = para_bbox[0]
                para_x1 = para_bbox[2]
            else:
                para_y0 = r.get("y0", 0)
                para_y1 = r.get("y1", 0)
                para_x0 = r.get("x0", 0)
                para_x1 = r.get("x1", 9999)

            para_h = para_y1 - para_y0
            if para_h > 0 and pg in table_y_ranges:
                for ty0, ty1, tx0, tx1 in table_y_ranges[pg]:
                    y_overlap = min(para_y1, ty1) - max(para_y0, ty0)
                    if y_overlap > para_h * 0.7:
                        x_overlap = min(para_x1, tx1) - max(para_x0, tx0)
                        if x_overlap > (para_x1 - para_x0) * 0.5 if para_x1 > para_x0 else True:
                            spatial_overlap = True
                            break

            # --- 内容重叠（paragraph token >50% 在 table 中）---
            content_overlap = False
            para_tokens = _tokenize_cell(text_data)
            if para_tokens and pg in table_tokens and table_tokens[pg]:
                t_tokens = table_tokens[pg]
                common = len(para_tokens & t_tokens)
                ratio = common / len(para_tokens) if para_tokens else 0
                if ratio >= 0.5:
                    content_overlap = True

            if spatial_overlap or content_overlap:
                should_remove = True
                fallback_removed += 1

        if should_remove:
            removed_count += 1
            continue

        keep_results.append(r)

    if removed_count > 0:
        detail = ""
        if index_removed > 0:
            detail += f"{index_removed} 个精确(index)"
        if fallback_removed > 0:
            if detail:
                detail += " + "
            detail += f"{fallback_removed} 个回退(bbox)"
        print(f"  [表格→文本去重] 移除了 {removed_count} 个段落/注解条目（{detail}）")

    return keep_results


def _tokenize_cell(text: str) -> frozenset:
    """将单元格文本规范化为可用于去重比较的 token 集合。

    策略：
    1. 去除空白 → 小写
    2. 移除数字中的千分位逗号（1,200 → 1200），避免逗号切割破坏数值完整性
    3. 移除中英文标点（保留字母数字和中文）
    4. findall 提取 ≥2 字中文片段 + ≥2 位数字/字母，单次遍历

    这样表格 cell "1,200" 和段落 "营业收入1,200万元"
    都能产生 token "1200"，实现跨提取源的数值匹配。
    """
    import re
    cleaned = re.sub(r'\s+', '', str(text)).lower()
    # 移除数字中的千分位逗号：1,200 → 1200
    cleaned = re.sub(r'(\d),(\d)', r'\1\2', cleaned)
    # 移除中英文标点（保留字母数字+中文）
    cleaned = re.sub(r'[，。；：、！？—…·,.;:!?()\[\]{}（）、\s/\\"\'%%＃＋－]+',
                     '', cleaned, flags=re.UNICODE)

    # findall：每次从上次结束位置继续，非重叠，逐词交替提取
    tokens = set(re.findall(
        r'[\u4e00-\u9fff]{2,}|\d{2,}|[a-z]{2,}',
        cleaned,
        flags=re.UNICODE,
    ))

    return frozenset(tokens)


def _is_table_data_subset(data_a: list, data_b: list, _row_fp_fn, _num_ratio_fn) -> bool:
    """检测表A的全部非空行是否完全是表B的前缀子集。

    用于整表去重：当表A的所有行恰好等于表B的前N行，且表B有更多内容时，
    表A是冗余碎片，应整体移除。

    判定条件（全部满足）：
    1. 表A有 >= 2 个非空行
    2. 表B非空行严格多于表A
    3. 表A的每一行都与表B同位置的行的指纹精确匹配
    """
    non_empty_a = [r for r in data_a if any(str(c).strip() for c in r)]
    non_empty_b = [r for r in data_b if any(str(c).strip() for c in r)]

    if len(non_empty_a) < 2:
        return False
    if len(non_empty_b) <= len(non_empty_a):
        return False

    # 表A所有行必须精确匹配表B的前面行（同位同指纹）
    for i, row_a in enumerate(non_empty_a):
        if i >= len(non_empty_b):
            return False
        fp_a = _row_fp_fn(row_a)
        fp_b = _row_fp_fn(non_empty_b[i])
        if not fp_a or not fp_b or fp_a != fp_b:
            return False

    # 额外校验：表A不能有表B没有的独特行（集合级子集检查兜底）
    fps_a = {_row_fp_fn(r) for r in non_empty_a if _row_fp_fn(r)}
    fps_b = {_row_fp_fn(r) for r in non_empty_b if _row_fp_fn(r)}
    if not fps_a.issubset(fps_b):
        return False

    return True


def _dedup_adjacent_tables_in_pipeline(results: list) -> list:
    """对同页相邻表进行跨表去重：前表尾/头部 ↔ 后表头部 方向性去重。

    检测同页内相邻表对 (Table_i, Table_{i+1})：
    - A0: 整表子集检测 — 表A全部行=表B前缀 → 移除表A
    - A1: 表i头部 ↔ 表i+1头部（表头行重叠）
    - A2: 表i尾部 ↔ 表i+1头部（数据行重叠）
    若表i+1结构完整 → 从表i删除重叠行。

    Args:
        results: 标准格式表格列表，每项含 data (List[List[str]]) 和 page 字段

    Returns:
        去重后的 results（原地修改 data 字段）
    """
    if not results or len(results) < 2:
        return results

    from codes.table_validator.rule_based_repair import (
        _has_complete_table_structure,
        _normalize_cell_content,
        _is_effectively_empty,
        _is_numeric_cell,
    )

    MAX_CHECK = 8

    def _row_fp(row):
        parts = []
        for c in row:
            s = _normalize_cell_content(str(c))
            if s:
                parts.append(s)
        return " | ".join(parts)

    def _row_cell_set(row):
        return {
            _normalize_cell_content(str(c))
            for c in row
            if _normalize_cell_content(str(c))
        }

    def _num_ratio(row):
        non_empty = [c for c in row if not _is_effectively_empty(str(c))]
        if not non_empty:
            return 0.0
        return sum(1 for c in non_empty if _is_numeric_cell(str(c))) / len(non_empty)

    def _jaccard(s1, s2):
        if not s1 and not s2:
            return 1.0
        if not s1 or not s2:
            return 0.0
        return len(s1 & s2) / len(s1 | s2)

    def _detect_header_indices(data):
        indices = []
        for i in range(min(MAX_CHECK, len(data))):
            row = data[i]
            nr = _num_ratio(row)
            if nr < 0.3 and any(not _is_effectively_empty(str(c)) for c in row):
                indices.append(i)
            else:
                break
        return indices

    def _row_matches(row_a, row_b):
        fp_a = _row_fp(row_a)
        fp_b = _row_fp(row_b)
        if fp_a and fp_b and fp_a == fp_b:
            return True
        set_a = _row_cell_set(row_a)
        set_b = _row_cell_set(row_b)
        if set_a and set_b and _jaccard(set_a, set_b) >= 0.7:
            sz = min(len(set_a), len(set_b)) / max(len(set_a), len(set_b))
            if sz >= 0.5:
                return True
        return False

    # 按页分组（仅表格，跳过段落）
    page_groups = {}
    for i, tbl in enumerate(results):
        if tbl.get("type") == "paragraph":
            continue
        p = tbl.get("page", 0)
        page_groups.setdefault(p, []).append(i)

    total_removed = 0
    entries_to_remove = set()  # 整表子集 → 完全移除的条目索引

    for page, indices in page_groups.items():
        if len(indices) < 2:
            continue

        for j in range(len(indices) - 1):
            idx_a, idx_b = indices[j], indices[j + 1]
            tbl_a = results[idx_a]
            tbl_b = results[idx_b]
            data_a = tbl_a.get("data", [])
            data_b = tbl_b.get("data", [])

            if not data_a or not data_b:
                continue
            if not _has_complete_table_structure(data_b):
                continue

            rows_to_remove = set()

            # A0: 整表子集检测 — 表A全部非空行=表B前缀 → 整表移除
            # 场景：表A仅有表头行，表B有相同表头+数据行，表A是冗余碎片
            if _is_table_data_subset(data_a, data_b, _row_fp, _num_ratio):
                saved_rows = []
                for ri, row in enumerate(data_a):
                    if any(str(c).strip() for c in row):
                        saved_rows.append({
                            "index": ri,
                            "cells": [str(c) for c in row],
                            "numeric_ratio": round(_num_ratio(row), 2),
                            "reason": "entire_table_subset_of_next",
                        })
                tbl_a.setdefault("_dedup_moved_rows", [])
                tbl_a["_dedup_moved_rows"].extend(saved_rows)
                tbl_a["_dedup_entire_table_moved"] = True
                tbl_a["_dedup_moved_to_id"] = tbl_b.get("table_id", idx_b)
                tbl_a["data"] = []
                tbl_a["rows"] = 0
                entries_to_remove.add(idx_a)
                total_removed += len(saved_rows)
                print(f"  [整表去重] P{page} 表#{j + 1}→表#{j + 2}: "
                      f"整表({len(saved_rows)}行)是下一表前缀子集，移除")
                continue

            # A1: 前表头部 ↔ 后表头部
            # 仅当表A是"表头碎片"（几乎无数据行）时才删除其表头
            # 防止从有数据内容的表中错误删除正常表头（如连续财务表共享列名）
            # V5 修复：有2+行表头 + ≥1行数据的表是完整迷你表，不是碎片
            # 例如：1行数据 + 2行表头的 mini 表不应被去重吞掉表头
            data_rows_a = sum(1 for r in data_a if _num_ratio(r) >= 0.3)
            header_rows_a = len(_detect_header_indices(data_a))
            is_header_fragment = (
                data_rows_a == 0 or
                (len(data_a) <= 5 and data_rows_a <= 1 and header_rows_a <= 1)
            )

            if is_header_fragment:
                head_a = _detect_header_indices(data_a)
                head_b = _detect_header_indices(data_b)
                head_b_rows = [data_b[i] for i in head_b]
                for hi in head_a:
                    for hb_row in head_b_rows:
                        if _row_matches(data_a[hi], hb_row):
                            rows_to_remove.add(hi)
                            break

            # A2: 前表尾部 ↔ 后表头部（排除表头行）
            # V6 修复：小型表（行数 < MAX_CHECK）时 tail_start=0 会把表头行
            # 误当尾部数据行来匹配，导致连续同类表格的表头被反复删除。
            # 表头行的去重只应由 A1 处理（且仅限碎片表），A2 应仅处理数据行重叠。
            head_a_indices = set(_detect_header_indices(data_a))
            tail_start = max(0, len(data_a) - MAX_CHECK)
            tail = data_a[tail_start:]
            head = data_b[:MAX_CHECK]
            for offset, row in enumerate(tail):
                actual_idx = tail_start + offset
                if actual_idx in head_a_indices:
                    continue
                if not _row_fp(row):
                    continue
                for h_row in head:
                    if _row_matches(row, h_row):
                        rows_to_remove.add(actual_idx)
                        break

            # A2+: 处理孤立章节标题行
            # 章节标题行（如"48 其他资产减值损失"）非空单元格少，
            # 无法通过常规 Jaccard 匹配到下一表的数据行。
            # 检测条件：尾部最后2行内、低数值占比、1-2个非空单元格、
            # 且后表有自己的表头结构 → 标题应归属下一个表
            head_b_for_title = _detect_header_indices(data_b)
            if head_b_for_title:
                for offset in range(1, min(3, len(data_a) + 1)):
                    tail_idx = len(data_a) - offset
                    if tail_idx in rows_to_remove:
                        continue
                    row = data_a[tail_idx]
                    nr = _num_ratio(row)
                    non_empty_count = sum(1 for c in row if not _is_effectively_empty(str(c)))
                    if nr < 0.3 and 1 <= non_empty_count <= 2:
                        rows_to_remove.add(tail_idx)

            # 执行重新分割（V2优化：保留完整行数据到元数据，而非简单删除）
            if rows_to_remove:
                # 先保存被移除的完整行数据，以便恢复
                saved_rows = []
                for row_idx in sorted(rows_to_remove):
                    saved_rows.append({
                        "index": row_idx,
                        "cells": [str(c) for c in data_a[row_idx]],
                        "numeric_ratio": round(_num_ratio(data_a[row_idx]), 2),
                        "reason": "moved_to_table_below",
                    })
                # 记录到表A的元数据
                tbl_a.setdefault("_dedup_moved_rows", [])
                tbl_a["_dedup_moved_rows"].extend(saved_rows)

                # 从尾部移除（这些行归属于下一张表，保留完整副本在元数据中）
                sorted_indices = sorted(rows_to_remove, reverse=True)
                for row_idx in sorted_indices:
                    del data_a[row_idx]
                removed = len(sorted_indices)
                total_removed += removed
                print(f"  [跨表去重] P{page} 表#{j + 1}→表#{j + 2}: "
                      f"从表#{j + 1}尾部重新分割 {removed} 行到下一表（后表结构完整）")

    if total_removed > 0:
        print(f"  [跨表去重] 总计删除 {total_removed} 行重叠数据")
    if entries_to_remove:
        results = [r for i, r in enumerate(results) if i not in entries_to_remove]
        print(f"  [整表去重] 移除 {len(entries_to_remove)} 张冗余子集表")
    return results


# ============================================================
# PDF处理器
# ============================================================
class PDFProcessor:
    """PDF处理器"""

    # === v2 表格提取算法参数配置 ===
    V2_CONFIG = {
        # 行分组
        "y_threshold_factor": 0.4,       # 动态阈值：中位gap × 因子
        "y_threshold_min": 2.0,          # 最小值
        "y_threshold_max": 15.0,          # 最大值

        # 列检测
        "align_tolerance": 4.0,          # 对齐聚簇容差(pt)
        "gap_factor": 0.3,               # gap阈值：中位gap + stdev × 因子
        "gap_min": 10.0,                 # gap最小值
        "line_merge_tolerance": 2.0,     # 竖线去重容差(pt)
        "column_line_min_count": 2,      # 指令1最少竖线条数

        # 表格区域
        "table_min_width_ratio": 0.3,    # 表格最小宽度/页宽
        "table_min_height": 20.0,        # 表格最小高度
        "density_grid": 10,              # 文本密度网格数
        "density_threshold": 0.8,        # 密度阈值(×平均值倍数)

        # 单元格分配
        "row_margin_factor": 0.2,        # 行分配允许越界比例

        # 置信度
        "confidence_col_weight": 0.35,   # 列数一致性权重
        "confidence_empty_weight": 0.25, # 空值率权重
        "confidence_num_weight": 0.25,   # 数值占比权重
        "confidence_line_bonus": 0.15,   # 表格线加分

        # 过滤（严格模式：V2宁缺毋滥，漏掉的表格由docx通道补充）
        "financial_keywords": [
            "万元", "元", "百万", "十亿", "%", "比率",
            "资产", "负债", "收入", "利润", "现金", "股东",
            "资本", "充足率", "率", "额", "数"
        ],
        "min_text_length": 50,           # 最小文本长度

        # pdfplumber降级
        "pdfplumber_min_words": 20,      # 单页最低word数
        "pdfplumber_min_row_words": 3,   # 每行最低word数
    }

    def __init__(self):
        self.config = load_config()

    def is_image_pdf(self, pdf_path=None, context=None):
        """检测是否为图片型PDF（扫描件）
        使用 get_text('dict') 检测实际文本块，比字符数判断更可靠；
        采样前5页+中部若干页，避免封面/签章页导致全局误判
        
        Args:
            pdf_path: PDF 文件路径（向后兼容，context 为 None 时使用）
            context: PDFContext 共享上下文（优先使用）
        """
        import fitz  # PyMuPDF

        if context:
            doc = context.doc
            close_doc = False
        else:
            doc = fitz.open(pdf_path)
            close_doc = True

        total = len(doc)
        # 采样页：前5页 + 中部区域（避免仅靠封面判断）
        sample_pages = list(range(min(5, total)))
        if total > 10:
            mid = total // 2
            for p in range(mid - 2, min(mid + 3, total)):
                if p not in sample_pages:
                    sample_pages.append(p)

        image_pages = 0
        text_pages = 0
        details = []

        for page_num in sample_pages:
            page = doc[page_num]
            text_dict = page.get_text("dict")
            blocks = text_dict.get("blocks", [])

            # 统计有实际文本内容的 span 数量（比字符数更可靠）
            text_spans = 0
            total_chars = 0
            for block in blocks:
                if block.get("type") == 0:  # 文本块
                    for line in block.get("lines", []):
                        for span in line.get("spans", []):
                            t = span.get("text", "").strip()
                            if t:
                                total_chars += len(t)
                                text_spans += 1

            images = page.get_images()

            # 真正扫描件：没有任何文本 span 但有图片
            # 有文本 span(>=1)就认为是文本页，不再用字符数阈值（CJK 文本容易误判）
            if text_spans == 0 and len(images) > 0:
                image_pages += 1
                details.append(f"p{page_num+1}=图片")
            elif text_spans > 0:
                text_pages += 1
                details.append(f"p{page_num+1}=文本({text_spans}span/{total_chars}字)")
            else:
                details.append(f"p{page_num+1}=空白")

        if close_doc:
            doc.close()

        result = image_pages > text_pages
        print(f"  [PDF检测] 采样{len(sample_pages)}页: {', '.join(details)}")
        print(f"  [PDF检测] 文本页={text_pages}, 图片页={image_pages} → {'图片型PDF' if result else '文本型PDF'}")
        return result

    def extract_text_tables(self, pdf_path=None, max_pages=None, context=None, progress_callback=None, progress_base=20, skip_drawings=False):
        """提取文本型PDF中的表格，保留位置信息
        
        Args:
            pdf_path: PDF 文件路径（向后兼容）
            max_pages: 最大处理页数
            context:  PDFContext 共享上下文（优先使用）
            progress_callback: callback(value, message) 逐页进度
            progress_base: 进度条起始值（默认20）
            skip_drawings: 跳过 drawings（避免 PyMuPDF 崩溃）
        """
        import fitz

        version = self.config.get("extraction_version", "v2")
        if version == "v2":
            return self._extract_text_tables_v2(pdf_path, max_pages, context, progress_callback, progress_base, skip_drawings)

        # ========== v1 逻辑（原有代码，完全不动）==========
        import re
        import pdfplumber

        if context:
            doc = context.doc
            close_doc = False
        else:
            doc = fitz.open(pdf_path)
            close_doc = True

        total_pages = len(doc)

        if max_pages:
            total_pages = min(max_pages, total_pages)

        results = []

        for page_num in range(total_pages):
            page = doc[page_num]
            page_rect = page.rect

            if progress_callback:
                pct = progress_base + int((page_num + 1) / total_pages * 10)
                progress_callback(pct, f"V1提取表格: 第{page_num + 1}/{total_pages}页...")

            # 方法1: 使用PyMuPDF直接获取页面的完整文本和位置信息（确保不丢失边缘数据）
            try:
                text_dict = page.get_text("dict")
                blocks = text_dict.get("blocks", [])

                page_x0 = page_rect.x0
                page_x1 = page_rect.x1
                page_y0 = page_rect.y0
                page_y1 = page_rect.y1

                words = []
                for block in blocks:
                    if block.get("type") == 0:  # 文本块
                        for line in block.get("lines", []):
                            for span in line.get("spans", []):
                                text = span.get("text", "").strip()
                                if text:
                                    bbox = span.get("bbox", [0, 0, 0, 0])
                                    words.append({
                                        "text": text,
                                        "x0": bbox[0],
                                        "y0": bbox[1],
                                        "x1": bbox[2],
                                        "y1": bbox[3],
                                    })

                if words:
                    full_text = " ".join([w["text"] for w in words])
                    financial_keywords = ["万元", "元", "百万", "十亿", "%", "比率", "资产", "负债", "收入", "利润",
                                         "现金", "股东", "资本", "充足率", "率", "额", "数"]
                    has_financial = any(kw in full_text for kw in financial_keywords)

                    if has_financial and len(full_text) > 50:
                        table_data = self._reconstruct_table_from_blocks_improved(words, page_rect)
                        if table_data and len(table_data) > 1:
                            table_data = self._normalize_table_columns(table_data)
                            results.append({
                                "page": page_num + 1,
                                "type": "table",
                                "data": table_data,
                                "text": full_text,
                                "extractor": "pymupdf_position"
                            })
            except Exception as e:
                print(f"  PyMuPDF位置提取第{page_num + 1}页失败: {e}")

            # 方法2: 使用pdfplumber的表格检测获取行边界
            pdfplumber_page = None
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    if page_num < len(pdf.pages):
                        pdfplumber_page = pdf.pages[page_num]
                        # 新版本pdfplumber直接调用find_tables()，不需要settings参数
                        found_tables = pdfplumber_page.find_tables()
                        all_words = pdfplumber_page.extract_words()

                        if all_words and found_tables:
                            full_text = " ".join([w.get("text", "") for w in all_words])
                            financial_keywords = ["万元", "元", "百万", "十亿", "%", "比率", "资产", "负债", "收入",
                                                 "利润", "现金", "股东", "资本", "充足率", "率", "额", "数"]
                            has_financial = any(kw in full_text for kw in financial_keywords)

                            if has_financial and len(full_text) > 50:
                                table_data = self._reconstruct_table_with_pdfplumber_rows(
                                    pdfplumber_page, found_tables, all_words
                                )
                                if table_data and len(table_data) > 1:
                                    table_data = self._normalize_table_columns(table_data)
                                    results.append({
                                        "page": page_num + 1,
                                        "type": "table",
                                        "data": table_data,
                                        "text": full_text,
                                        "extractor": "pdfplumber_hybrid"
                                    })
            except Exception as e:
                print(f"  pdfplumber提取第{page_num + 1}页失败: {e}")

            # 方法3: 使用PyMuPDF BLOCK模式提取带位置的文本
            if not any(r.get("page") == page_num + 1 and r.get("type") == "table" for r in results):
                text_dict = page.get_text("dict")
                blocks = text_dict.get("blocks", [])

                text_blocks = []
                for block in blocks:
                    if block.get("type") == 0:
                        for line in block.get("lines", []):
                            for span in line.get("spans", []):
                                text = span.get("text", "").strip()
                                if text:
                                    text_blocks.append({
                                        "text": text,
                                        "x0": span.get("bbox", [0, 0, 0, 0])[0],
                                        "y0": span.get("bbox", [0, 0, 0, 0])[1],
                                        "x1": span.get("bbox", [0, 0, 0, 0])[2],
                                        "y1": span.get("bbox", [0, 0, 0, 0])[3],
                                    })

                if text_blocks:
                    full_text = " ".join([b["text"] for b in text_blocks])
                    financial_keywords = ["万元", "元", "百万", "十亿", "%", "比率", "资产", "负债", "收入", "利润",
                                         "现金", "股东", "资本", "充足率"]
                    has_financial = any(kw in full_text for kw in financial_keywords)

                    if has_financial and len(full_text) > 50:
                        table_data = self._reconstruct_table_from_blocks_improved(text_blocks, page_rect)
                        if table_data and len(table_data) > 1:
                            table_data = self._normalize_table_columns(table_data)
                            results.append({
                                "page": page_num + 1,
                                "type": "table",
                                "data": table_data,
                                "text": full_text,
                                "extractor": "position_based"
                            })
                        else:
                            table_data = self._reconstruct_table_from_blocks(text_blocks, page_rect.width)
                            if table_data and len(table_data) > 1:
                                table_data = self._normalize_table_columns(table_data)
                                results.append({
                                    "page": page_num + 1,
                                    "type": "table",
                                    "data": table_data,
                                    "text": full_text,
                                    "extractor": "position_based_fallback"
                                })

        if close_doc:
            doc.close()

        # V1 不再合并同一页的多个表格，每个表格独立保留
        # results = self._merge_tables_on_same_page(results)
        return results

    def _merge_tables_on_same_page(self, results):
        """合并同一页的多个表格"""
        if not results:
            return results

        page_groups = {}
        for table in results:
            page = table.get("page", 0)
            if page not in page_groups:
                page_groups[page] = []
            page_groups[page].append(table)

        merged_results = []
        for page in sorted(page_groups.keys()):
            tables = page_groups[page]

            if len(tables) == 1:
                merged_results.append(tables[0])
            else:
                merged_data = []
                merged_extractors = []

                # 先计算所有表格的最大列数
                all_max_cols = 0
                for table in tables:
                    data = table.get("data", [])
                    if data:
                        all_max_cols = max(all_max_cols, max(len(row) for row in data))

                # 再合并表格
                for i, table in enumerate(tables):
                    data = table.get("data", [])
                    if not data:
                        continue

                    if merged_data:
                        separator_row = ["--- 表格" + str(i) + " ---"] + [""] * (all_max_cols - 1)
                        merged_data.append(separator_row)

                    for row in data:
                        padded_row = list(row) + [None] * (all_max_cols - len(row))
                        merged_data.append(padded_row)

                    merged_extractors.append(table.get("extractor", "unknown"))

                merged_results.append({
                    "page": page,
                    "type": "table",
                    "data": merged_data,
                    "text": "",
                    "extractor": "+".join(merged_extractors)
                })

        return merged_results

    def _normalize_table_columns(self, table_data):
        """规范化表格 — 委托给独立函数。"""
        return _normalize_table_columns(table_data)

    def _reconstruct_table_from_blocks(self, text_blocks, page_width):
        """根据文本块位置信息重建表格结构"""
        if not text_blocks:
            return None

        rows = []
        current_row = []
        current_y = None
        y_threshold = 5

        sorted_blocks = sorted(text_blocks, key=lambda b: (round(b["y0"] / y_threshold), b["x0"]))

        for block in sorted_blocks:
            y = round(block["y0"] / y_threshold)
            if current_y is None or abs(y - current_y) <= 1:
                current_row.append(block)
                current_y = y
            else:
                if current_row:
                    rows.append(current_row)
                current_row = [block]
                current_y = y

        if current_row:
            rows.append(current_row)

        table_data = []
        for row in rows:
            sorted_row = sorted(row, key=lambda b: b["x0"])
            row_data = [span["text"] for span in sorted_row]
            row_text = "".join(row_data)
            if len(row_text.strip()) > 0:
                table_data.append(row_data)

        return table_data

    def _reconstruct_table_from_blocks_improved(self, text_blocks, page_rect):
        """改进的表格重建方法"""
        if not text_blocks:
            return None

        if hasattr(page_rect, 'width'):
            page_width = page_rect.width
            page_x0 = page_rect.x0 if hasattr(page_rect, 'x0') else 0
        else:
            page_width = page_rect[2] if len(page_rect) > 2 else page_rect[0]
            page_x0 = page_rect[0] if len(page_rect) > 0 else 0

        all_x0 = [b["x0"] for b in text_blocks]
        all_x1 = [b["x1"] for b in text_blocks]

        if not all_x0:
            return None

        min_x = min(all_x0)
        max_x = max(all_x1)

        x_points = sorted(set(all_x0 + all_x1))
        if len(x_points) < 2:
            return None

        # 计算自适应列边界阈值（基于x坐标分布）
        gaps = []
        for i in range(len(x_points) - 1):
            gap = x_points[i + 1] - x_points[i]
            gaps.append((x_points[i], x_points[i + 1], gap))

        # 自适应阈值：取gap的中位数*1.5，更能适应不同PDF
        if gaps:
            all_gaps = [g[2] for g in gaps if g[2] > 0]
            if all_gaps:
                import statistics
                median_gap = statistics.median(all_gaps)
                gap_threshold = max(median_gap * 1.5, 10)  # 最小10pt
            else:
                gap_threshold = 15
        else:
            gap_threshold = 15

        column_boundaries = []

        for x_start, x_end, gap in gaps:
            if gap > gap_threshold:
                column_boundaries.append((x_start + x_end) / 2)

        if not column_boundaries:
            column_boundaries = [min_x, max_x]
        else:
            column_boundaries = sorted(set(column_boundaries))
            # 确保左右边界包含所有内容
            if column_boundaries[0] > min_x:
                column_boundaries.insert(0, (min_x + column_boundaries[0]) / 2)
            if column_boundaries[-1] < max_x:
                column_boundaries.append((column_boundaries[-1] + max_x) / 2)

        y_threshold = 5
        sorted_blocks = sorted(text_blocks, key=lambda b: b["y0"])

        rows = []
        current_row = []
        current_y = None

        for block in sorted_blocks:
            y = round(block["y0"] / y_threshold)
            if current_y is None or abs(y - current_y) <= 1:
                current_row.append(block)
                current_y = y
            else:
                if current_row:
                    rows.append(current_row)
                current_row = [block]
                current_y = y

        if current_row:
            rows.append(current_row)

        table_data = []
        for row_blocks in rows:
            sorted_row = sorted(row_blocks, key=lambda b: b["x0"])
            row_data = [""] * (len(column_boundaries) - 1)

            for block in sorted_row:
                col_idx = self._find_column_index(block["x0"], block["x1"], column_boundaries)
                if 0 <= col_idx < len(row_data):
                    if row_data[col_idx]:
                        row_data[col_idx] += " " + block["text"]
                    else:
                        row_data[col_idx] = block["text"]

            if any(cell.strip() for cell in row_data):
                table_data.append(row_data)

        return table_data if table_data else None

    def _reconstruct_table_with_pdfplumber_rows(self, pdfplumber_page, pdfplumber_tables, words):
        """使用pdfplumber的表格检测获取精确的行列边界"""
        if not pdfplumber_tables or not words:
            return None

        table_data = []

        for table in pdfplumber_tables:
            table_bbox = table.bbox
            if not table_bbox:
                continue

            table_top = table_bbox[1]
            table_bottom = table_bbox[3]
            table_left = table_bbox[0]
            table_right = table_bbox[2]

            table_rows = table.rows
            if not table_rows:
                continue

            # 获取列数：Row对象是可迭代的，但不支持len()，转换为列表
            first_row = list(table_rows[0]) if table_rows else []
            num_cols = len(first_row)

            for row_cells in table_rows:
                row_data = []

                for cell in row_cells:
                    cell_bbox = cell.bbox
                    if not cell_bbox:
                        row_data.append("")
                        continue

                    cell_left = cell_bbox[0]
                    cell_right = cell_bbox[2]
                    cell_top = cell_bbox[1]
                    cell_bottom = cell_bbox[3]

                    cell_texts = []
                    for w in words:
                        word_x0 = w.get("x0", 0)
                        word_x1 = w.get("x1", 0)
                        word_top = w.get("top", 0)
                        word_bottom = w.get("bottom", 0)
                        word_mid_y = (word_top + word_bottom) / 2

                        if cell_top <= word_mid_y <= cell_bottom:
                            if word_x0 < cell_right and word_x1 > cell_left:
                                cell_texts.append(w)

                    if cell_texts:
                        cell_texts.sort(key=lambda w: w.get("x0", 0))
                        cell_text = " ".join([w.get("text", "") for w in cell_texts])
                    else:
                        cell_text = ""

                    row_data.append(cell_text)

                if any(cell.strip() for cell in row_data):
                    table_data.append(row_data)

        return table_data if table_data else None

    def _detect_column_boundaries_by_spacing(self, text_blocks, page_width):
        """根据文本间距检测列边界"""
        if not text_blocks:
            return [0, page_width]

        x_coords = [b["x0"] for b in text_blocks] + [b["x1"] for b in text_blocks]

        bucket_size = 20
        max_x = max(x_coords) if x_coords else page_width
        buckets = {}

        for x in x_coords:
            bucket = int(x / bucket_size)
            buckets[bucket] = buckets.get(bucket, 0) + 1

        if len(buckets) < 2:
            return [0, page_width]

        avg_density = sum(buckets.values()) / len(buckets)

        gaps = []
        sorted_buckets = sorted(buckets.keys())

        for i in range(len(sorted_buckets) - 1):
            bucket1, bucket2 = sorted_buckets[i], sorted_buckets[i + 1]
            mid_buckets = range(bucket1 + 1, bucket2)
            gap_density = sum(buckets.get(b, 0) for b in mid_buckets)

            if gap_density < avg_density * 0.3:
                gaps.append((bucket1 * bucket_size + bucket_size / 2, bucket2 * bucket_size))

        boundaries = [0]
        for start, end in sorted(gaps, key=lambda x: x[0]):
            boundaries.append((start + end) / 2)
        boundaries.append(page_width)

        if len(boundaries) < 3:
            boundaries = [0, page_width * 0.3, page_width * 0.6, page_width]

        return sorted(set(boundaries))

    def _find_column_index(self, x0, x1, column_boundaries):
        """找到文本块属于哪一列"""
        center_x = (x0 + x1) / 2

        for i in range(len(column_boundaries) - 1):
            if column_boundaries[i] <= center_x < column_boundaries[i + 1]:
                return i

        if center_x < column_boundaries[0]:
            return 0
        elif center_x >= column_boundaries[-1]:
            return len(column_boundaries) - 2
        else:
            min_dist = float('inf')
            closest_col = 0
            for i in range(len(column_boundaries) - 1):
                mid = (column_boundaries[i] + column_boundaries[i + 1]) / 2
                dist = abs(center_x - mid)
                if dist < min_dist:
                    min_dist = dist
                    closest_col = i
            return closest_col

    def _parse_text_to_table(self, text):
        """将文本解析为表格格式"""
        import re

        lines = text.split('\n')
        table_data = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            parts = re.split(r'\s{2,}|\t', line)
            parts = [p.strip() for p in parts if p.strip()]

            if parts and (any(c.isdigit() for p in parts for c in p) or
                         any(kw in line for kw in ["资产", "负债", "收入", "利润", "合计", "小计"])):
                table_data.append(parts)

        return table_data if table_data else [[text]]

    def pdf_to_images(self, pdf_path=None, output_dir=None, context=None):
        """将PDF转换为图片
        
        Args:
            pdf_path: PDF 文件路径（向后兼容）
            output_dir:  输出目录
            context:     PDFContext 共享上下文（优先使用）
        """
        import fitz

        if output_dir is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(TEMP_DIR) / f"pdf_images_{timestamp}"
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        if context:
            return context.generate_all_llm_images(output_dir)

        # 向后兼容：无 context 时自己打开 PDF
        doc = fitz.open(pdf_path)
        image_paths = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            mat = fitz.Matrix(2, 2)
            pix = page.get_pixmap(matrix=mat)

            output_path = output_dir / f"page_{page_num + 1}.png"
            pix.save(str(output_path))
            image_paths.append(str(output_path))

        doc.close()
        return image_paths




# ============================================================
# v2 表格提取算法
# ============================================================

    # ---- v2 入口 ----

    def _extract_text_tables_v2(self, pdf_path=None, max_pages=None, context=None,
                                 progress_callback=None, progress_base=20, skip_drawings=False):
        """v2 表格提取入口（委托给 V2Pipeline 模块化管线）
        
        Args:
            pdf_path: PDF 文件路径
            max_pages: 最大处理页数
            context:  PDFContext 共享上下文（优先使用）
            progress_callback: callback(value, message) 逐页进度
            progress_base: 进度条起始值
            skip_drawings: 跳过 get_drawings()
        """
        from codes.v2_steps.pipeline import V2Pipeline

        pipeline = V2Pipeline()
        return pipeline.run(
            pdf_path=pdf_path,
            max_pages=max_pages,
            context=context,
            progress_callback=progress_callback,
            progress_base=progress_base,
            skip_drawings=skip_drawings,
        )

    def _extract_text_tables_v2_legacy(self, pdf_path=None, max_pages=None, context=None,
                                        progress_callback=None, progress_base=20, skip_drawings=False):
        """[回退] v2 表格提取入口（原始单体实现，当 Pipeline 出问题时使用）"""
        import fitz
        import statistics

        if context:
            doc = context.doc
            close_doc = False
        else:
            doc = fitz.open(pdf_path)
            close_doc = True

        total_pages = len(doc)
        if max_pages:
            total_pages = min(max_pages, total_pages)

        results = []
        cfg = self.V2_CONFIG

        for page_num in range(total_pages):
            page = doc[page_num]
            page_rect = page.rect

            if progress_callback:
                pct = progress_base + int((page_num + 1) / total_pages * 10)
                progress_callback(pct, f"V2扫描: 第{page_num + 1}/{total_pages}页")

            words_raw = page.get_text("words")
            words = []
            for w in words_raw:
                words.append({
                    "x0": w[0], "y0": w[1],
                    "x1": w[2], "y1": w[3],
                    "text": w[4],
                    "baseline": w[3],
                })

            drawings = []
            if not skip_drawings:
                try:
                    drawings_raw = page.get_drawings()
                    for d in drawings_raw:
                        rect = d["rect"]
                        w = rect.width
                        h = rect.height
                        direction = None
                        if w > h * 5:
                            direction = "h"
                        elif h > w * 5:
                            direction = "v"
                        drawings.append({
                            "type": "line" if (w < h * 0.3 or h < w * 0.3) else "rect",
                            "direction": direction,
                            "x0": rect.x0, "y0": rect.y0,
                            "x1": rect.x1, "y1": rect.y1,
                            "color": d.get("color"),
                            "width": d.get("width", 1),
                            "fill": d.get("fill"),
                        })
                except Exception:
                    print(f"  [V2] 第{page_num+1}页: get_drawings() 失败，使用纯文本检测")

            if not words:
                print(f"  [V2] 第{page_num+1}页: get_text('words')返回空，尝试dict回退...")
                words = PDFProcessor._extract_words_from_dict(page)
                if words:
                    print(f"  [V2] 第{page_num+1}页: dict回退成功，提取到{len(words)}个文本片段")
                else:
                    print(f"  [V2] 第{page_num+1}页: dict回退也失败，跳过该页")
                    continue

            full_text = " ".join(w["text"] for w in words)
            if not any(kw in full_text for kw in cfg["financial_keywords"]):
                print(f"  [V2] 第{page_num+1}页: 未匹配金融关键词，跳过 (文本长度={len(full_text)}, 预览={full_text[:60]!r})")
                continue
            if len(full_text) < cfg["min_text_length"]:
                print(f"  [V2] 第{page_num+1}页: 文本长度{len(full_text)}不满足最低{cfg['min_text_length']}要求，跳过")
                continue

            table_regions = self._detect_table_region(drawings, page_rect.width, page_rect.height)
            if not table_regions:
                table_regions = self._detect_table_region_by_text(words, page_rect.width, page_rect.height)
            if not table_regions:
                print(f"  [V2] 第{page_num+1}页: 未检测到表格区域，跳过")
                continue

            from codes.content_segmenter.segmenter import ContentSegmenter
            from codes.content_segmenter.segment_logger import SegmentLogger

            segmenter = ContentSegmenter()
            pdf_stem = Path(pdf_path).stem if pdf_path else "unknown"

            for region in table_regions:
                rx0, ry0, rx1, ry1 = region
                region_words = [w for w in words
                                if rx0 <= w["x0"] <= rx1 and ry0 <= w["y0"] <= ry1]
                if len(region_words) < 3:
                    continue

                def _word_getter(w):
                    return (w["x0"], w["x1"], w["y0"], w["y1"], w["text"])

                seg_result = segmenter.segment_region(
                    text_items=region_words,
                    page_width=page_rect.width,
                    page_height=page_rect.height,
                    page_number=page_num + 1,
                    region_bbox=region,
                    item_getter=_word_getter,
                )

                try:
                    SegmentLogger.log_page_diff(
                        pdf_name=pdf_stem,
                        page_number=page_num + 1,
                        before_regions=[{
                            "bbox": [round(rx0, 2), round(ry0, 2), round(rx1, 2), round(ry1, 2)],
                            "item_count": len(region_words),
                            "text_preview": " ".join(w["text"] for w in region_words[:20]),
                        }],
                        after_segment=seg_result,
                        page_size={"width": round(page_rect.width, 2), "height": round(page_rect.height, 2)},
                    )
                except Exception:
                    pass

                paragraph_found_count = 0
                for sr in seg_result.regions:
                    srx0, sry0, srx1, sry1 = sr.x0, sr.y0, sr.x1, sr.y1

                    if sr.is_paragraph:
                        paragraph_found_count += 1
                        para_words = [w for w in words
                                      if srx0 <= w["x0"] <= srx1 and sry0 <= w["y0"] <= sry1]
                        para_words.sort(key=lambda w: (w["y0"], w["x0"]))
                        lines = []
                        cur_line = []
                        cur_y = None
                        for w in para_words:
                            if cur_y is None or abs(w["y0"] - cur_y) <= 5.0:
                                cur_line.append(w["text"])
                                if cur_y is None:
                                    cur_y = w["y0"]
                            else:
                                if cur_line:
                                    lines.append(" ".join(cur_line))
                                cur_line = [w["text"]]
                                cur_y = w["y0"]
                        if cur_line:
                            lines.append(" ".join(cur_line))
                        para_text = "\n".join(lines).strip()
                        if para_text and len(para_text) >= 3:
                            results.append({
                                "page": page_num + 1,
                                "type": "paragraph",
                                "data": para_text,
                                "text": para_text,
                                "extractor": "v2_segmenter",
                                "confidence": sr.confidence,
                                "rows": len(lines),
                                "cols": 1,
                                "bbox": [round(srx0, 2), round(sry0, 2), round(srx1, 2), round(sry1, 2)],
                            })
                        continue

                    sub_region_words = [w for w in words
                                        if rx0 <= w["x0"] <= rx1 and sry0 <= w["y0"] <= sry1]
                    if len(sub_region_words) < 3:
                        continue

                    context_text = PDFProcessor._extract_context_text_from_words(
                        words, srx0, sry0, srx1, sry1, margin=100.0)

                    row_bounds = self._detect_horizontal_lines(page, sub_region_words, drawings)
                    if len(row_bounds) < 2:
                        continue

                    col_bounds = self._detect_vertical_lines(page, sub_region_words, drawings)
                    if len(col_bounds) < 3:
                        continue

                    table_data = self._assign_words_to_grid(sub_region_words, row_bounds, col_bounds)
                    if not table_data or len(table_data) < 2:
                        continue

                    merge_info = {}
                    merge_stats = {}
                    table_data_before_merge = [list(row) for row in table_data]
                    if drawings:
                        table_data, merge_info, merge_stats = self._detect_and_apply_merge_cells(
                            table_data, drawings, row_bounds, col_bounds)
                        if merge_stats.get("total_spans", 0) > 0:
                            print(f"  [V2 Merge] 第{page_num+1}页: 检测到 {merge_stats['total_spans']} 个合并单元格 "
                                  f"(线条={merge_stats.get('line_spans',0)}, 文本={merge_stats.get('text_spans',0)}, "
                                  f"合并={merge_stats.get('cells_merged',0)}个cell)")

                    table_data = self._normalize_table_columns(table_data)

                    has_border = bool([d for d in drawings if d["direction"] in ("h", "v")])
                    confidence = self._compute_table_confidence(table_data, has_border, words)

                    results.append({
                        "page": page_num + 1,
                        "type": "table",
                        "data": table_data,
                        "text": full_text,
                        "extractor": "v2_position_based",
                        "confidence": confidence,
                        "rows": len(table_data),
                        "cols": len(col_bounds) - 1,
                        "has_border": has_border,
                        "context_text": context_text,
                        "merge_info": merge_info,
                        "merge_stats": merge_stats,
                        "table_data_before_merge": table_data_before_merge,
                    })

                if paragraph_found_count > 0:
                    seg_info = seg_result.to_dict()
                    print(f"  [V2] 第{page_num+1}页: 内容分割 → {seg_info['table_regions']}个表格 + {seg_info['paragraph_regions']}个段落")
                    print(f"  [V2] 第{page_num+1}页: 段切详情 → {seg_result.region_count}个子区域, {round(seg_result.segment_time_ms,1)}ms")

        if close_doc:
            doc.close()
        _mark_page_types(results)
        # V3: 使用统一去重引擎替代分散的 _deduplicate_text_against_tables
        dedup_engine = DeduplicationEngine()
        results = dedup_engine.dedup_text_against_tables(results)
        if dedup_engine._debug_log:
            print(f"  [V2去重] {len(dedup_engine._debug_log)} 项文本去重")
        return results

    @staticmethod
    def _extract_context_text_from_words(words, rx0, ry0, rx1, ry1, margin=100.0):
        """从 PDF words 中提取表格区域上方的上下文文本。

        Args:
            words: PDF 页面的所有 words 列表 (每个 word 含 x0, y0, x1, y1, text)
            rx0, ry0, rx1, ry1: 表格区域边界
            margin: 表格上方搜索范围（pt），默认 100pt（约 2-3 行文本）

        Returns:
            str: 拼接后的上下文文本，按 y 坐标从上到下排列
        """
        # 筛选表格上方 margin 范围内且 x 方向与表格有重叠的 words
        context_top = max(0, ry0 - margin)
        context_words = [
            w for w in words
            if context_top <= w["y0"] < ry0
            and w["x1"] > rx0 * 0.8  # 允许 x 方向有 20% 外扩容差
            and w["x0"] < rx1 * 1.2
            and w["text"].strip()
        ]

        if not context_words:
            return ""

        # 按 y 坐标升序分组为行
        context_words.sort(key=lambda w: (w["y0"], w["x0"]))
        lines = []
        current_line = []
        current_y = None
        y_tolerance = 5.0  # 同行的 y 容差

        for w in context_words:
            if current_y is None or abs(w["y0"] - current_y) <= y_tolerance:
                current_line.append(w["text"])
                if current_y is None:
                    current_y = w["y0"]
            else:
                lines.append(' '.join(current_line))
                current_line = [w["text"]]
                current_y = w["y0"]

        if current_line:
            lines.append(' '.join(current_line))

        return '\n'.join(lines).strip()

    @staticmethod
    def _extract_words_from_dict(page):
        """当 get_text('words') 返回空时的回退方案
        从 get_text('dict') 的 blocks/lines/spans 中提取文本和坐标
        """
        text_dict = page.get_text("dict")
        blocks = text_dict.get("blocks", [])
        words = []
        for block in blocks:
            if block.get("type") == 0:  # 文本块
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        text = span.get("text", "").strip()
                        if text:
                            bbox = span.get("bbox", [0, 0, 0, 0])
                            words.append({
                                "x0": bbox[0],
                                "y0": bbox[1],
                                "x1": bbox[2],
                                "y1": bbox[3],
                                "text": text,
                            "baseline": bbox[3],
                        })
        return words

    # ---- docx 表格提取（pdf2docx 通道） ----

    def _extract_tables_via_docx(self, pdf_path=None, context=None, progress_callback=None):
        """通过 pdf2docx 将 PDF 转为 Word，从 Word 表格结构中提取数据。

        全内存操作（BytesIO），不落盘。
        输出格式与 V1/V2 统一。

        Args:
            pdf_path: PDF 文件路径（向后兼容）
            context:  PDFContext 共享上下文（优先使用）
            progress_callback: callback(value, message) 推送进度
        Returns:
            [{page, type, data, extractor, confidence, ...}]
        """
        from io import BytesIO

        if context:
            _pdf_path = context.pdf_path
        else:
            _pdf_path = pdf_path

        print(f"  [docx] 开始 pdf2docx 全内存转换...")
        t0 = time.time()
        total_hint = context.page_count if context else "?"
        if progress_callback:
            progress_callback(22, f"docx: PDF转Word中({total_hint}页,约2-5分钟)...")

        # 步骤1：pdf2docx → 内存 BytesIO
        # 注意：cv.convert() 是阻塞调用，内部无进度回调，此阶段进度条会停留约2-5分钟
        try:
            from pdf2docx import Converter
        except ImportError:
            print(f"  [docx] 错误：未安装 pdf2docx 库，请执行 pip install pdf2docx")
            return []

        import tempfile
        import multiprocessing
        cpu_count = min(multiprocessing.cpu_count(), 6)  # 最多6核，避免资源争抢

        # PyInstaller 冻结环境下，pdf2docx 内部多进程会导致子进程闪退
        _frozen = getattr(sys, 'frozen', False)
        _use_mp = False if _frozen else True
        _mp_cpu = 1 if _frozen else cpu_count
        if _frozen:
            print("  [docx] 冻结环境，禁用 pdf2docx 内部多进程")

        # pdf2docx 多进程模式需要落盘（子进程间通过 JSON 文件交换数据）
        with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as tmp:
            tmp_path = tmp.name

        cv = Converter(_pdf_path)
        cv.convert(
            tmp_path,
            start=0,
            end=None,
            layout=False,           # 流式模式：表格识别更准
            table_deduction=False,  # 保守策略
            multi_processing=_use_mp,
            cpu_count=_mp_cpu,
        )
        cv.close()

        buf = BytesIO()
        with open(tmp_path, 'rb') as f:
            buf.write(f.read())
        os.unlink(tmp_path)
        buf.seek(0)

        elapsed = time.time() - t0
        print(f"  [docx] pdf2docx 转换完成，耗时 {elapsed:.1f}s")
        if progress_callback:
            progress_callback(30, f"docx: 转换完成({elapsed:.0f}s),解析表格...")

        # 步骤2：python-docx 解析
        try:
            from docx import Document
        except ImportError:
            print(f"  [docx] 错误：未安装 python-docx 库，请执行 pip install python-docx")
            return []

        doc = Document(buf)
        buf.close()

        # 步骤3：遍历文档 body 子元素，通过分页符推算每个表格的真实页码
        # 同时收集每个表格上方的段落文本作为上下文
        W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
        body = doc.element.body

        # 先扫描所有 body 子元素
        # page_table_map:  {表格XML元素: PDF页码}
        # table_context_map: {表格XML元素: 表格上方段落文本}
        current_page = 1
        page_table_map = {}
        table_context_map = {}
        pending_paragraphs = []  # 收集表格前的段落文本

        def _extract_paragraph_text(child_elem):
            """从 w:p 元素中提取纯文本"""
            texts = []
            for t_elem in child_elem.iter(f'{W}t'):
                if t_elem.text:
                    texts.append(t_elem.text)
            return ''.join(texts).strip()

        for child in body:
            tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag

            if tag == 'p':
                para_text = _extract_paragraph_text(child)
                # 检查段落中是否包含分页符
                has_page_break = False

                # 方式1：w:br type='page'（显式分页符）
                for br in child.iter(f'{W}br'):
                    br_type = br.get(f'{W}type')
                    if br_type == 'page':
                        current_page += 1
                        has_page_break = True
                        break

                # 方式2：w:lastRenderedPageBreak（Word 渲染分页标记）
                if not has_page_break:
                    for lrpb in child.iter(f'{W}lastRenderedPageBreak'):
                        current_page += 1
                        has_page_break = True
                        break

                # 方式3：段落属性中的分页设置（w:pageBreakBefore）
                if not has_page_break:
                    pPr = child.find(f'{W}pPr')
                    if pPr is not None:
                        # 段前分页
                        pbf = pPr.find(f'{W}pageBreakBefore')
                        if pbf is not None:
                            val = pbf.get(f'{W}val')
                            if val not in ('0', 'false'):
                                current_page += 1
                                has_page_break = True

                # 方式4：sectPr 内嵌在段落中（节分隔符通常也意味着换页）
                if not has_page_break:
                    pPr = child.find(f'{W}pPr')
                    if pPr is not None:
                        sectPr = pPr.find(f'{W}sectPr')
                        if sectPr is not None:
                            sect_type = sectPr.find(f'{W}type')
                            # 仅在 type='nextPage' 或无 type（默认 nextPage）时递增
                            if sect_type is None or sect_type.get(f'{W}val') == 'nextPage':
                                current_page += 1
                                has_page_break = True

                # 方式5：检测页眉页脚中的纯数字页码（如 13、14、15）
                if context and para_text and not has_page_break:
                    stripped = para_text.strip()
                    if stripped.isdigit():
                        pn = int(stripped)
                        total = context.page_count
                        if 1 <= pn <= total and pn > current_page:
                            current_page = pn
                            has_page_break = True

                if has_page_break:
                    pending_paragraphs.clear()

                if para_text:
                    pending_paragraphs.append(para_text)
                # 限制缓存的段落数量，避免积累过多
                if len(pending_paragraphs) > 5:
                    pending_paragraphs = pending_paragraphs[-5:]

            elif tag == 'tbl':
                page_table_map[child] = current_page
                # 取表格前最近 1-3 段作为上下文文本
                ctx_text = '\n'.join(pending_paragraphs[-3:]) if pending_paragraphs else ""
                table_context_map[child] = ctx_text
                pending_paragraphs.clear()  # 表格之后的段落属于下一个表格

            elif tag == 'sectPr':
                # body 级别的 sectPr（文档最后的节属性），也可能意味着换页
                # 但文档末尾的 sectPr 不需要处理，跳过
                pass

        if not page_table_map:
            print(f"  [docx] Word 中未检测到任何表格")
            return []

        # 通过 python-docx 的 Table 对象处理表格数据
        tables = doc.tables
        results = []
        for tbl_idx, table in enumerate(tables):
            try:
                tbl_elem = table._tbl
                page_num = page_table_map.get(tbl_elem, tbl_idx + 1)

                # 提取表格级列宽
                tblGrid = tbl_elem.find(f'{W}tblGrid')
                col_widths = []
                if tblGrid is not None:
                    for gridCol in tblGrid.findall(f'{W}gridCol'):
                        w = float(gridCol.get(f'{W}w', 0))
                        col_widths.append(w)

                # 逐行解析
                rows_data = []
                merge_tracker = {}  # {(row, col): True} 被垂直合并占用的单元格

                for r, tr in enumerate(table.rows):
                    row_cells = []
                    col_idx = 0

                    for cell in tr.cells:
                        # 跳过被垂直合并占用的位置
                        while merge_tracker.get((r, col_idx)):
                            row_cells.append("")
                            col_idx += 1

                        tc = cell._tc
                        tcPr = tc.find(f'{W}tcPr')
                        col_span = 1
                        row_start = True

                        if tcPr is not None:
                            # gridSpan：跨列
                            gridSpan = tcPr.find(f'{W}gridSpan')
                            if gridSpan is not None:
                                col_span = int(gridSpan.get(f'{W}val', 1))

                            # vMerge：垂直合并
                            vMerge = tcPr.find(f'{W}vMerge')
                            if vMerge is not None:
                                val = vMerge.get(f'{W}val')
                                if val != 'restart':
                                    # 被合并的后续行
                                    row_start = False

                        if row_start:
                            text = cell.text.strip()
                            for span in range(col_span):
                                if span == 0:
                                    row_cells.append(text)
                                else:
                                    row_cells.append("")

                            if vMerge is not None:
                                # 标记下方被合并的行
                                for rr in range(r + 1, len(table.rows)):
                                    merge_tracker[(rr, col_idx)] = True
                                    for s in range(1, col_span):
                                        merge_tracker[(rr, col_idx + s)] = True
                        else:
                            # 垂直合并的延续行：占位
                            for span in range(col_span):
                                row_cells.append("")

                        col_idx += col_span

                    if row_cells:
                        rows_data.append(row_cells)

                if rows_data:
                    context_text = table_context_map.get(tbl_elem, "")
                    results.append({
                        "page": page_num,
                        "type": "table",
                        "data": rows_data,
                        "text": "",
                        "extractor": "docx_based",
                        "confidence": 0.85,
                        "rows": len(rows_data),
                        "cols": max(len(r) for r in rows_data) if rows_data else 0,
                        "has_border": True,
                        "context_text": context_text,
                    })
                    if context_text:
                        ctx_preview = context_text[:50].replace('\n', ' ')
                        print(f"  [docx] 表格{tbl_idx+1}(PDF第{page_num}页): {len(rows_data)}行{results[-1]['cols']}列表格, 上下文: {ctx_preview}...")
                    else:
                        print(f"  [docx] 表格{tbl_idx+1}(PDF第{page_num}页): {len(rows_data)}行{results[-1]['cols']}列表格")

            except Exception as e:
                print(f"  [docx] 表格{tbl_idx+1}解析失败: {e}")
                continue

        # 步骤4：V2-Lite 坐标匹配校验页码
        if context and results:
            if progress_callback:
                progress_callback(33, "docx: V2-Lite 坐标匹配校验页码...")
            # V2-Lite 扫描＋Jaccard匹配＋bbox冲突检测＋排序 全部在内部完成
            results = self._verify_docx_page_numbers(results, context)
            # results 已按页码排序，无需再次排序

        if progress_callback:
            progress_callback(36, f"docx提取完成: {len(results)}个表格")
        print(f"  [docx] 共提取 {len(results)} 个表格")
        return results

    # ---- 逐页 pdf2docx 转换（页码 100% 准确） ----
    # 工作函数已移至 _worker.py（独立模块，无 Qt 依赖，PyInstaller 子进程安全）

    def _extract_tables_via_docx_per_page(self, pdf_path=None, context=None,
                                           progress_callback=None):
        """逐页 pdf2docx 转换：每页独立转 DOCX，表格页码 100% 准确。

        相比一次性全文档转换，逐页方案从根本上消除了"猜页码→DP修正"
        的信息损失链条，表格从哪页 DOCX 解析出来就属于哪页。

        性能优化：
        - 多进程并行转换（每进程独立 Python 解释器，完全隔离 PyMuPDF C 扩展）

        TODO: 跨页表格合并（后续单独实现）

        Args:
            pdf_path: PDF 文件路径（向后兼容）
            context:  PDFContext 共享上下文（优先使用）
            progress_callback: callback(value, message) 推送进度
        Returns:
            [{page, type, data, extractor, confidence, context_text, ...}]
        """
        import multiprocessing
        from codes.pdf_extractor._log import write_log, log_exception

        # PyInstaller 兼容：使用 spawn 上下文的 multiprocessing.Pool，
        # ProcessPoolExecutor 在冻结环境中经常失败
        _frozen = getattr(sys, 'frozen', False)
        _pool_cls = multiprocessing.Pool
        if _frozen:
            _ctx = multiprocessing.get_context('spawn')
            _pool_cls = _ctx.Pool
            write_log(f"[docx-per-page] 冻结环境，使用 spawn Pool")
        else:
            write_log(f"[docx-per-page] 普通环境，使用默认 Pool")

        if context:
            _pdf_path = context.pdf_path
        else:
            _pdf_path = pdf_path

        write_log(f"[docx-per-page] 开始: pdf={_pdf_path}")

        total_pages = context.page_count if context else 0
        if total_pages == 0:
            import fitz
            doc = fitz.open(_pdf_path)
            total_pages = len(doc)
            doc.close()

        cpu_count = min(multiprocessing.cpu_count(), 6)

        # 将页分配到各进程（每个进程处理一批页，复用 Converter）
        pages_per_batch = (total_pages + cpu_count - 1) // cpu_count
        page_batches = []
        for t in range(cpu_count):
            start = t * pages_per_batch + 1
            end = min(start + pages_per_batch - 1, total_pages)
            if start <= end:
                page_batches.append(list(range(start, end + 1)))

        write_log(f"[docx-per-page] {total_pages} 页, {cpu_count} 进程, "
                  f"{len(page_batches)} 批次")

        print(f"  [docx-per-page] 逐页转换: {total_pages} 页, "
              f"{cpu_count} 进程并行 × ~{pages_per_batch} 页/进程")
        if progress_callback:
            progress_callback(22,
                f"docx: 逐页转换 {total_pages} 页 ({cpu_count}进程)...")

        t0 = time.time()

        # 延迟导入工作函数（确保 __module__ 为 _worker，子进程安全）
        from codes.pdf_extractor._worker import convert_batch

        all_tables = []
        total_batches = len(page_batches)
        completed_batches = 0

        write_log(f"[docx-per-page] 创建 multiprocessing.Pool (processes={cpu_count})")
        pool = None
        try:
            pool = _pool_cls(processes=cpu_count)
            async_results = []
            for idx, batch in enumerate(page_batches):
                r = pool.apply_async(convert_batch, (_pdf_path, batch))
                async_results.append((idx, r))

            write_log(f"[docx-per-page] {len(async_results)} 个任务已提交")

            for idx, r in async_results:
                try:
                    batch_tables = r.get(timeout=600)  # 10分钟超时
                    all_tables.extend(batch_tables)
                    write_log(f"[docx-per-page] 批次 "
                              f"{completed_batches+1}/{total_batches} 完成: "
                              f"{len(batch_tables)} 个表格")
                except Exception as e:
                    log_exception(f"[docx-per-page] 批次{idx}异常: {e}")
                    print(f"  [docx-per-page] 进程批处理异常: {e}")

                completed_batches += 1
                if progress_callback:
                    pct = 22 + int(completed_batches / total_batches * 8)
                    progress_callback(pct,
                        f"docx: 批次 {completed_batches}/{total_batches}")
        except Exception as e:
            log_exception(f"[docx-per-page] Pool 致命异常: {e}")
            write_log(f"[docx-per-page] 回退到单进程模式")
            print(f"  [docx-per-page] 多进程异常，回退单进程: {e}")
            # 回退：单进程顺序处理
            for batch in page_batches:
                try:
                    batch_tables = convert_batch(_pdf_path, batch)
                    all_tables.extend(batch_tables)
                except Exception as e2:
                    log_exception(f"[docx-per-page] 单进程批次异常: {e2}")
        finally:
            if pool is not None:
                pool.close()
                pool.join()

        # 按页码排序
        all_tables.sort(key=lambda t: t.get("page", 0))

        elapsed = time.time() - t0
        print(f"  [docx-per-page] 转换完成: {len(all_tables)} 个表格, "
              f"耗时 {elapsed:.1f}s")

        # 注意：不做上下文补充，避免 PyMuPDF C 扩展 refcount bug
        # 子进程已完全隔离，主进程 all_tables 是纯 Python 数据（list[dict]），
        # 不包含任何 PyMuPDF C 对象，后续 deepcopy 安全，无需 gc.collect()

        # 打印摘要
        pages_with_tables = len(set(t["page"] for t in all_tables))
        print(f"  [docx-per-page] {len(all_tables)} 表 / "
              f"{pages_with_tables} 页 / {total_pages} 页 → "
              f"覆盖率 {pages_with_tables/total_pages*100:.0f}%")

        if progress_callback:
            progress_callback(30, f"docx: 提取完成 ({len(all_tables)}表格)")

        return all_tables

    def _enrich_per_page_context(self, all_tables, context):
        """为逐页转换的表补充上下文文本（从 PDF 页文本提取，仅用于展示）。"""
        try:
            pages_with_tables = set(t["page"] for t in all_tables)

            for pn in sorted(pages_with_tables):
                try:
                    page = context.get_page(pn - 1)
                    page_text = page.get_text("text")
                    if not page_text:
                        continue
                    # 取页面前 200 字符作为上下文（仅展示用途）
                    ctx = page_text[:200].strip()

                    if ctx:
                        for t in all_tables:
                            if t["page"] == pn and not t.get("context_text"):
                                t["context_text"] = ctx
                except Exception:
                    pass
        except Exception as e:
            print(f"  [docx-per-page] 上下文补充异常: {e}")

    # ---- V2-Lite 页码分配（委托给独立模块 page_assigner） ----

    def _verify_docx_page_numbers(self, results, context):
        """基于 V2-Lite 物理坐标匹配的页码验证。

        - 逐页转换结果（extractor='docx_per_page'）：页码 100% 准确，跳过 DP
        - 全量转换结果（extractor='docx_based'）：委托 page_assigner 修正

        详见 codes/pdf_extractor/page_assigner.py
        """
        if not results:
            return results

        # 逐页转换 → 页码已准确，跳过 DP 验证
        if results[0].get("extractor") == "docx_per_page":
            print(f"  [docx] 逐页转换结果，页码已 100% 准确，跳过 DP 修正")
            return results

        from codes.pdf_extractor.page_assigner import assign_docx_pages
        return assign_docx_pages(results, context)

    # ---- 表格去重与标题提取 ----

    @staticmethod
    def _table_fingerprint(table_data, sample_cells=6):
        """生成表格的轻量指纹：前 sample_cells 个非空单元格文本（向后兼容接口）

        注意：增强版指纹请使用 _table_fingerprint_v2。
        """
        cells = []
        for row in table_data[:2]:  # 只看前两行
            for cell in row:
                if cell and str(cell).strip():
                    s = str(cell).strip()
                    if len(s) >= 2:
                        cells.append(s)
                        if len(cells) >= sample_cells:
                            break
            if len(cells) >= sample_cells:
                break
        return frozenset(cells) if cells else None

    @staticmethod
    def _table_fingerprint_v2(table_data, max_cells=20):
        """增强版指纹：5行 + 20格 + 数值优先 + 行结构哈希

        相比旧版 _table_fingerprint（只看前2行6格），增强版：
        1. 扩展到5行，max_cells=20，大幅提升区分度
        2. 数值型单元格优先（数值几乎不可能跨表重复）
        3. 附加行结构签名（每行的列数序列），区分不同结构的表

        Returns:
            (frozenset, tuple) 或 None: (指纹集合, 行结构签名)
        """
        cells = []
        numeric_cells = []
        text_cells = []

        for row in table_data[:5]:
            for cell in row:
                if not cell:
                    continue
                s = str(cell).strip()
                if len(s) < 2:
                    continue
                cleaned = s.replace(",", "").replace("%", "").replace("(", "-").replace(")", "")
                try:
                    float(cleaned)
                    numeric_cells.append(s.lower())
                except ValueError:
                    text_cells.append(s.lower())

        cells = numeric_cells[:15] + text_cells[:5]

        # 行结构签名：每行的列数序列
        row_structure = tuple(len(row) for row in table_data[:5])

        return (frozenset(cells[:max_cells]), row_structure) if cells else None

    @staticmethod
    def _deduplicate_v2_docx(v2_tables, docx_tables):
        """去重合并：docx 为主力通道，V2 补漏 docx 未覆盖的表格。

        原则：docx(pdf2word) 从 PDF 内容流重建表格结构，准确度高。
        V2 基于视觉坐标推测，存在合并单元格错位风险。

        规则：
        1. docx 所有表格无条件保留(主力通道)，保持原始阅读顺序不动
        2. V2 段落(type="paragraph")不参与指纹匹配，直接按页码插入
        3. 增强指纹匹配：行结构必须一致 + 公共词 >=3 + 公共词占比 >=40%
        4. 搜索范围扩展到 ±1 页容忍（同一张表可能被 docx 和 V2 分配到相邻页）
        5. V2 独有的表格(docx 漏掉的无框表/小表)按页码插入到 docx 序列的正确位置
        """
        # 分离 V2 中的表格和段落
        v2_paragraphs = [vt for vt in v2_tables if vt.get("type") == "paragraph"]
        v2_real_tables = [vt for vt in v2_tables if vt.get("type") != "paragraph"]

        if not docx_tables:
            merged = list(v2_real_tables)
            # 按页码插入段落
            for para in v2_paragraphs:
                _insert_by_page(merged, para, prefer_before=True)
            return merged

        merged = list(docx_tables)  # docx 保持原顺序，绝不动
        matched_v2_ids = set()

        # 对每个 docx 表格，在 ±1 页范围内找 V2 匹配项（仅匹配 V2 表格，跳过段落）
        for di, dt in enumerate(docx_tables):
            dt_page = dt.get("page", 0)
            dt_data = dt.get("data", [])
            dt_fp = PDFProcessor._table_fingerprint_v2(dt_data)

            if not dt_fp:
                continue

            dt_fp_set, dt_structure = dt_fp

            for vi, vt in enumerate(v2_real_tables):
                if vi in matched_v2_ids:
                    continue
                vt_page = vt.get("page", 0)
                # ±1 页容忍
                if abs(vt_page - dt_page) > 1:
                    continue
                vt_data = vt.get("data", [])
                vt_fp = PDFProcessor._table_fingerprint_v2(vt_data)
                if not vt_fp:
                    continue

                vt_fp_set, vt_structure = vt_fp

                # 行结构不同 → 不可能是同一张表
                if dt_structure != vt_structure:
                    continue

                common = dt_fp_set & vt_fp_set
                min_size = min(len(dt_fp_set), len(vt_fp_set))
                if min_size == 0:
                    continue
                # 增强匹配：>=3 个公共词 且 公共词占较小集合的 >=40%
                if len(common) >= 3 and len(common) >= min_size * 0.4:
                    matched_v2_ids.add(vi)
                    page_note = f"P{dt_page}" if vt_page == dt_page else f"P{dt_page}/P{vt_page}(±1)"
                    print(f"  [去重] {page_note}: docx表{di+1} ← V2表{vi+1}(匹配{len(common)}个公共词, 行结构={dt_structure})")

        # 收集未匹配的 V2 表格，按页码排序后插入到 docx 序列
        unmatched_v2 = []
        for vi, vt in enumerate(v2_real_tables):
            if vi not in matched_v2_ids:
                unmatched_v2.append(vt)

        if unmatched_v2:
            unmatched_v2.sort(key=lambda x: x.get("page", 0))
            for vt in reversed(unmatched_v2):
                vt_page = vt.get("page", 0)
                insert_pos = len(merged)
                for i in range(len(merged)):
                    if merged[i].get("page", 0) > vt_page:
                        insert_pos = i
                        break
                merged.insert(insert_pos, vt)

        # V2 段落按页码插入到合并序列中
        for para in v2_paragraphs:
            PDFProcessor._insert_by_page(merged, para, prefer_before=False)

        v2_table_supplement = len(v2_real_tables) - len(matched_v2_ids)
        v2_para_count = len(v2_paragraphs)
        parts = [f"docx主力={len(docx_tables)}个"]
        if v2_table_supplement:
            parts.append(f"V2补漏表={v2_table_supplement}个")
        if v2_para_count:
            parts.append(f"V2段落={v2_para_count}个")
        print(f"  [去重] 汇总: {' + '.join(parts)} = 共{len(merged)}个条目")
        return merged

    @staticmethod
    def _insert_by_page(merged: list, item: dict, prefer_before: bool = False) -> None:
        """将 item 按页码插入到 merged 列表的正确位置。

        Args:
            merged: 已排序的条目列表
            item: 要插入的条目（含 "page" 和可选 "bbox" 字段）
            prefer_before: True=插入到同页已有条目前面, False=插入到后面
        """
        item_page = item.get("page", 0)
        item_y0 = (item.get("bbox") or [0, 0])[1] if item.get("bbox") else 0

        insert_pos = len(merged)
        for i, existing in enumerate(merged):
            existing_page = existing.get("page", 0)
            if existing_page > item_page:
                insert_pos = i
                break
            elif existing_page == item_page:
                if prefer_before:
                    insert_pos = i
                    break
                else:
                    # 在同页中，按 bbox Y 坐标排序
                    existing_y0 = (existing.get("bbox") or [0, 0])[1] if existing.get("bbox") else 0
                    if item_y0 < existing_y0:
                        insert_pos = i
                        break
                    insert_pos = i + 1

        merged.insert(insert_pos, item)

    @staticmethod
    def _filter_table_quality(tables):
        """过滤低质量表格。

        规则：
        1. 只有 1 行数据的跳过，除非是该页第一个表且含 >= 2 个数值
        2. 没有数值类型数据的跳过（纯文本块，不是表格）
        3. 图表误判过滤：坐标轴刻度、孤立单字图例、饼图标签
        4. 段落项（type="paragraph"）始终保留，不参与表格质量过滤
        """
        import re

        def count_numbers(data):
            """数有多少个含数字的单元格"""
            cnt = 0
            for row in data:
                for cell in row:
                    if cell and re.search(r'\d', str(cell)):
                        cnt += 1
            return cnt

        def has_any_number(data):
            return count_numbers(data) > 0

        def _is_chart_like_data(data):
            """快速检测表格数据是否为图表标签（P0 兜底过滤）。"""
            if not data or len(data) < 2:
                return False

            # 收集所有非空文本
            all_texts = []
            for row in data:
                all_texts.extend(str(c).strip() for c in row if str(c).strip())

            if not all_texts:
                return True  # 全空

            # 特征1：大量单中文字符（图例碎片）
            single_cn_count = sum(
                1 for t in all_texts
                if re.match(r'^[\u4e00-\u9fff]$', t)
            )
            if single_cn_count >= 3 and single_cn_count >= len(all_texts) * 0.5:
                return True

            # 特征2：坐标轴刻度模式（每行只有1~2个纯数字）
            # 检查每个数据行中的纯数字列数
            narrow_numeric_rows = 0
            for row in data:
                numeric_cols = []
                for cell in row:
                    c = str(cell).strip()
                    if not c:
                        continue
                    clean = c.replace(",", "").replace(" ", "")
                    try:
                        float(clean)
                        numeric_cols.append(clean)
                    except ValueError:
                        pass
                if 1 <= len(numeric_cols) <= 2:
                    narrow_numeric_rows += 1

            # 80%+ 的行只有1~2个数值 → 图表刻度
            if narrow_numeric_rows >= len(data) * 0.8 and len(data) >= 3:
                # 进一步验证：数值是否呈规律性变化
                values = []
                for row in data:
                    for cell in row:
                        c = str(cell).strip().replace(",", "")
                        try:
                            values.append(float(c))
                        except ValueError:
                            pass
                if len(values) >= 3:
                    # 检查是否为等差数列
                    diffs = [values[i+1] - values[i] for i in range(len(values)-1)]
                    abs_diffs = [abs(d) for d in diffs]
                    if abs_diffs:
                        mean = sum(abs_diffs) / len(abs_diffs)
                        if mean > 0:
                            std = (sum((d - mean) ** 2 for d in abs_diffs) / len(abs_diffs)) ** 0.5
                            if std / mean < 0.2:
                                return True

            return False

        filtered = []
        removed = 0
        chart_removed = 0
        kept_exceptions = 0
        last_page = None

        for t in tables:
            if t.get("type") == "paragraph":
                # 段落项始终保留，不参与表格质量过滤
                filtered.append(t)
                continue
        
            data = t.get("data", [])
            page = t.get("page", 0)
            is_first_on_page = (page != last_page)
            last_page = page

            # 规则1：至少 2 行（例外：页首表 + 单行 + 2个以上数值）
            if len(data) < 2:
                if is_first_on_page and count_numbers(data) >= 2:
                    # 例外：页首关键指标表（如每股收益、ROE等单行汇总）
                    kept_exceptions += 1
                else:
                    removed += 1
                    continue

            # 规则2：至少有一个数字
            if not has_any_number(data):
                removed += 1
                continue

            # 规则3：图表误判过滤（P0 兜底）
            if _is_chart_like_data(data):
                chart_removed += 1
                continue

            filtered.append(t)

        if removed or chart_removed or kept_exceptions:
            parts = []
            if removed:
                parts.append(f"移除{removed}个")
            if chart_removed:
                parts.append(f"图表过滤{chart_removed}个")
            if kept_exceptions:
                parts.append(f"保留{kept_exceptions}个页首单行表(含数值)")
            print(f"  [质量过滤] {'; '.join(parts)}")
        return filtered

    class TableAutoCorrector:
        """基于列数据特征的无框表格自动纠错器。

        核心策略：
        1. 分析每列的主导数据类型（数字/文本）
        2. 检测表头区域中被垂直拆分的单元格（如"加权平均"+"净资产收益率"）
        3. 通过子表头检测避免误合并父子层级（如"每股收益"+"基本"）
        """

        @staticmethod
        def correct(table_data):
            if not table_data or len(table_data) < 2:
                return table_data

            data = [list(row) for row in table_data]
            max_cols = max((len(r) for r in data), default=0)
            for r in data:
                while len(r) < max_cols:
                    r.append("")

            col_types = PDFProcessor.TableAutoCorrector._analyze_col_types(data)
            data = PDFProcessor.TableAutoCorrector._merge_vertical_headers(data, col_types)

            # 清理空行并重新规范化
            data = [r for r in data if any(str(c).strip() for c in r)]
            if data:
                max_cols = max(len(r) for r in data)
                for r in data:
                    while len(r) < max_cols:
                        r.append("")
            return data

        @staticmethod
        def _analyze_col_types(data):
            import re
            max_cols = max(len(r) for r in data)
            # 数据区从第2行开始（跳过可能的表头），至少跳过1行
            data_start = min(2, max(1, len(data) // 2))

            types = []
            for c in range(max_cols):
                vals = []
                for r in range(data_start, len(data)):
                    if c < len(data[r]):
                        v = str(data[r][c]).strip()
                        if v:
                            vals.append(v)

                if not vals:
                    types.append("empty")
                    continue

                numeric = 0
                for v in vals:
                    v_clean = v.replace(",", "").replace("(", "-").replace(")", "").replace("%", "").replace("\u2030", "")
                    try:
                        float(v_clean)
                        numeric += 1
                    except ValueError:
                        pass

                if numeric > len(vals) * 0.6:
                    types.append("numeric")
                elif numeric == 0:
                    types.append("text")
                else:
                    types.append("mixed")
            return types

        @staticmethod
        def _merge_vertical_headers(data, col_types, max_header_rows=2):
            if len(data) < 3:
                return data

            corrected = []
            i = 0
            while i < len(data):
                row = list(data[i])

                if i + 1 < len(data) and i < max_header_rows:
                    next_row = list(data[i + 1])
                    merged_any = False

                    for c in range(min(len(row), len(next_row), len(col_types))):
                        if not row[c] or not next_row[c]:
                            continue

                        a = str(row[c]).strip()
                        b = str(next_row[c]).strip()

                        # 跳过子表头（如"基本"不应与"每股收益"合并）
                        if PDFProcessor.TableAutoCorrector._is_likely_child_header(data, i + 1, c):
                            continue

                        # 合并条件：短文本、无数字、该列数据区以数字为主
                        if (len(a) <= 8 and len(b) <= 8 and
                                len(a) + len(b) <= 15 and
                                col_types[c] == "numeric" and
                                not PDFProcessor.TableAutoCorrector._has_digit(a) and
                                not PDFProcessor.TableAutoCorrector._has_digit(b)):
                            row[c] = a + b
                            next_row[c] = ""
                            merged_any = True

                    if merged_any:
                        corrected.append(row)
                        if any(str(x).strip() for x in next_row):
                            corrected.append(next_row)
                        i += 2
                        continue

                corrected.append(row)
                i += 1
            return corrected

        @staticmethod
        def _is_likely_child_header(data, row_idx, col_idx):
            if row_idx >= len(data) or col_idx >= len(data[row_idx]):
                return False

            val = str(data[row_idx][col_idx]).strip()
            if not val or len(val) > 5:
                return False

            row = data[row_idx]
            short_cols = []
            for c, cell in enumerate(row):
                if cell and str(cell).strip():
                    s = str(cell).strip()
                    if len(s) <= 5 and not PDFProcessor.TableAutoCorrector._has_digit(s):
                        short_cols.append(c)

            if len(short_cols) < 2:
                return False

            if row_idx > 0:
                prev_row = data[row_idx - 1]
                parent_vals = []
                for c in short_cols:
                    if c < len(prev_row) and prev_row[c] and str(prev_row[c]).strip():
                        parent_vals.append(str(prev_row[c]).strip())
                parent_vals = [v for v in parent_vals if v]
                if len(parent_vals) <= 1:
                    return True
                if len(set(parent_vals)) == 1:
                    return True
            return False

        @staticmethod
        def _has_digit(s):
            import re
            return bool(re.search(r'\d', str(s)))

    @staticmethod
    def _extract_table_title(table_data):
        """从表格数据中提取标题文字（用于 Sheet 命名）。

        规则：取第一个长度>=4的非空单元格作为标题，
        如果没有，取表格第一行的前 3 个非空单元拼接。
        """
        if not table_data:
            return "表格"
        for row in table_data:
            for cell in row:
                if cell and str(cell).strip():
                    s = str(cell).strip()
                    if len(s) >= 4:
                        # 截取前12个字符
                        return s[:12].replace("/", "-").replace("\\", "-").replace("*", "")
        # fallback：拼接前几个非空
        parts = []
        for row in table_data:
            for cell in row:
                if cell and str(cell).strip():
                    s = str(cell).strip()
                    if s not in parts:
                        parts.append(s[:4])
                    if len(parts) >= 3:
                        break
            if len(parts) >= 3:
                break
        return "-".join(parts)[:20] if parts else "表格"

    # ---- 表格区域检测 ----

    def _detect_table_region(self, drawings, page_width, page_height):
        """从 drawing 中检测表格外框区域"""
        cfg = self.V2_CONFIG

        rectangles = [
            d for d in drawings
            if d["type"] == "rect"
            and d["x1"] - d["x0"] > page_width * cfg["table_min_width_ratio"]
            and d["y1"] - d["y0"] > cfg["table_min_height"]
        ]

        h_lines = [
            d for d in drawings
            if d["type"] == "line" and d["direction"] == "h"
            and d["x1"] - d["x0"] > page_width * cfg["table_min_width_ratio"]
        ]
        v_lines = [
            d for d in drawings
            if d["type"] == "line" and d["direction"] == "v"
            and d["y1"] - d["y0"] > cfg["table_min_height"]
        ]

        regions = []

        for rect in rectangles:
            regions.append((rect["x0"], rect["y0"], rect["x1"], rect["y1"]))

        if len(h_lines) >= 2 and len(v_lines) >= 2:
            x0 = min(l["x0"] for l in v_lines)
            x1 = max(l["x1"] for l in v_lines)
            y0 = min(l["y0"] for l in h_lines)
            y1 = max(l["y1"] for l in h_lines)
            if x1 - x0 > page_width * cfg["table_min_width_ratio"] and y1 - y0 > cfg["table_min_height"]:
                if not any(self._has_overlap((x0, y0, x1, y1), [r]) for r in regions):
                    regions.append((x0, y0, x1, y1))

        return regions

    def _has_overlap(self, rect, regions):
        """检测两个区域是否重叠"""
        rx0, ry0, rx1, ry1 = rect
        for gx0, gy0, gx1, gy1 in regions:
            if not (rx1 <= gx0 or rx0 >= gx1 or ry1 <= gy0 or ry0 >= gy1):
                return True
        return False

    def _detect_table_region_by_text(self, words, page_width, page_height):
        """无框表格区域检测（文本密度法）- 恢复原逻辑"""
        cfg = self.V2_CONFIG
        if not words or len(words) < 20:
            return []

        grid_rows = cfg["density_grid"]
        grid_cols = cfg["density_grid"]
        cell_h = page_height / grid_rows
        cell_w = page_width / grid_cols

        density = [[0] * grid_cols for _ in range(grid_rows)]
        for w in words:
            col = int((w["x0"] + w["x1"]) / 2 / cell_w)
            row = int((w["y0"] + w["y1"]) / 2 / cell_h)
            if 0 <= row < grid_rows and 0 <= col < grid_cols:
                density[row][col] += 1

        row_density = [sum(density[r]) for r in range(grid_rows)]
        avg = sum(row_density) / max(len(row_density), 1)
        avg = max(avg, 3)

        table_row_indices = [
            r for r in range(grid_rows)
            if row_density[r] > avg * cfg["density_threshold"]
        ]

        if not table_row_indices:
            return []

        # 合并连续行
        table_row_ranges = self._merge_consecutive(table_row_indices)

        # 恢复原逻辑：只检测上下边界（行），左右边界交给列检测处理
        regions = []
        for start, end in table_row_ranges:
            y0 = start * cell_h
            y1 = (end + 1) * cell_h
            # 左右边界使用整个页面宽度，让列检测算法决定真正的边界
            regions.append((0, y0, page_width, y1))

        return regions

    def _merge_consecutive(self, indices):
        """合并连续整数索引为 [(start, end), ...]"""
        if not indices:
            return []
        indices = sorted(set(indices))
        ranges = []
        start = indices[0]
        end = indices[0]
        for i in indices[1:]:
            if i == end + 1:
                end = i
            else:
                ranges.append((start, end))
                start = i
                end = i
        ranges.append((start, end))
        return ranges

    # ---- 行边界检测 ----

    def _detect_horizontal_lines(self, page, words, page_drawings):
        """检测行边界，返回 [(y_top, y_bottom), ...]"""
        cfg = self.V2_CONFIG

        h_lines = sorted(set(
            d["y0"] for d in page_drawings
            if d["type"] == "line" and d["direction"] == "h"
            and d["x1"] - d["x0"] > page.rect.width * cfg["table_min_width_ratio"]
        ))

        if len(h_lines) >= 2:
            row_bounds = []
            for i in range(len(h_lines) - 1):
                row_bounds.append((h_lines[i], h_lines[i + 1]))
            return row_bounds

        # 无水平线 → 动态阈值分组
        y_threshold = self._compute_dynamic_y_threshold(words)
        rows = self._group_words_into_rows(words, y_threshold)

        row_bounds = []
        for row_words in rows:
            if row_words:
                y_top = min(w["y0"] for w in row_words)
                y_bot = max(w["y1"] for w in row_words)
                row_bounds.append((y_top, y_bot))

        return row_bounds

    def _compute_dynamic_y_threshold(self, words):
        """动态计算行分组阈值"""
        import statistics
        cfg = self.V2_CONFIG

        if not words or len(words) < 3:
            return 5.0

        y_positions = sorted(set(w["y0"] for w in words if w["text"].strip()))
        if len(y_positions) < 5:
            return 5.0

        gaps = []
        for i in range(len(y_positions) - 1):
            gap = y_positions[i + 1] - y_positions[i]
            if 0.5 < gap < 50:
                gaps.append(gap)

        if len(gaps) < 3:
            return 5.0

        median_gap = statistics.median(gaps)
        threshold = median_gap * cfg["y_threshold_factor"]
        return max(cfg["y_threshold_min"], min(cfg["y_threshold_max"], threshold))

    def _group_words_into_rows(self, words, y_threshold):
        """按 y 坐标对 words 进行行分组"""
        if not words:
            return []

        sorted_words = sorted(words, key=lambda w: w["y0"])
        rows = []
        current_row = [sorted_words[0]]
        current_y = sorted_words[0]["y0"]

        for w in sorted_words[1:]:
            if abs(w["y0"] - current_y) <= y_threshold:
                current_row.append(w)
                current_y = (current_y + w["y0"]) / 2
            else:
                rows.append(sorted(current_row, key=lambda ww: ww["x0"]))
                current_row = [w]
                current_y = w["y0"]

        if current_row:
            rows.append(sorted(current_row, key=lambda ww: ww["x0"]))

        return rows

    # ---- 列边界检测 ----

    def _detect_vertical_lines(self, page, words, page_drawings):
        """
        检测列边界（v2.1规格：三指令融合 + 线条锚点增强）
        
        增强点：
        - 指令1放宽阈值：>=2条竖线即尝试按线切分
        - 新增线条去重（合并1pt内的相近线）
        - 指令3增加线条锚点：部分竖线作为gap检测的强制分割点
        
        返回值：[x0, x1, x2, ...] 列分割线位置
        """
        import statistics
        cfg = self.V2_CONFIG
        page_width = page.rect.width

        # ----- 提取并去重垂直线 -----
        raw_v_lines = sorted(set(
            d["x0"] for d in page_drawings
            if d["type"] == "line" and d["direction"] == "v"
        ))
        v_lines = self._merge_nearby_lines(raw_v_lines, cfg["line_merge_tolerance"])

        # 区分为内线和外线
        inner_lines = [x for x in v_lines
                       if page_width * 0.05 < x < page_width * 0.95]

        # ----- 指令1：垂直线直接切分（增强版） -----
        min_line_count = cfg.get("column_line_min_count", 2)
        if len(v_lines) >= min_line_count and len(inner_lines) >= 1:
            # 至少有1条内部竖线 → 可以用线条定义列结构
            # 组合：左边距 + 所有线条 + 右边距（去重排序）
            boundaries = sorted(set([0] + v_lines + [page_width]))
            if len(boundaries) >= 3:  # 确保至少有2列
                return boundaries

        # 收集可用于后续融合的锚点线（内部竖线是强信号）
        anchor_lines = inner_lines[:]

        # ----- 指令2：文本对齐聚簇 -----
        x0_list = [w["x0"] for w in words if w["text"].strip()]
        x1_list = [w["x1"] for w in words if w["text"].strip()]

        if x0_list:
            # x0对齐点检测
            left_aligns = self._cluster_1d(x0_list, cfg["align_tolerance"])
            right_aligns = self._cluster_1d(x1_list, cfg["align_tolerance"])

            # 合并左右对齐点
            all_aligns = sorted(set(left_aligns + right_aligns))

            # 如果有锚点线，与对齐点融合
            if anchor_lines:
                all_aligns = self._fuse_line_anchors_with_aligns(
                    all_aligns, anchor_lines, cfg["align_tolerance"])

            # 如果对齐点足够，直接返回
            if len(all_aligns) >= 3:
                return all_aligns

        # ----- 指令3：gap检测（兜底） -----
        all_x = sorted(set(x0_list + x1_list))

        if len(all_x) < 3:
            return [0, page_width]

        # 计算gap
        gaps = []
        gap_positions = []
        for i in range(len(all_x) - 1):
            gap = all_x[i + 1] - all_x[i]
            if gap > 0:
                gaps.append(gap)
                gap_positions.append((all_x[i], all_x[i + 1]))

        if not gaps:
            return [0, page_width]

        # 用中位数 + 标准差作为阈值
        median_gap = statistics.median(gaps)
        stdev_gap = statistics.stdev(gaps) if len(gaps) >= 2 else median_gap * 0.5
        gap_threshold = max(median_gap + stdev_gap * cfg["gap_factor"], cfg["gap_min"])

        # 找到gap大于阈值的位置
        boundaries = [0]
        for (left, right), gap in zip(gap_positions, gaps):
            if gap > gap_threshold:
                boundaries.append((left + right) / 2)

        # 线条锚点注入：确保锚点线位置被纳入列边界
        if anchor_lines:
            boundaries = self._fuse_line_anchors_with_aligns(
                boundaries, anchor_lines, cfg["gap_min"])
        else:
            boundaries.append(page_width)

        return sorted(set(boundaries))

    # ---- 线条辅助方法 ----

    @staticmethod
    def _merge_nearby_lines(lines, tolerance):
        """合并容差范围内的邻近竖线（PDF常见双线/冗余线问题）
        
        例如：[100, 101, 200, 300] tolerance=2.0 → [100.5, 200, 300]
        """
        if not lines:
            return []
        sorted_lines = sorted(lines)
        merged = [sorted_lines[0]]
        for x in sorted_lines[1:]:
            if x - merged[-1] <= tolerance:
                # 合并为均值
                merged[-1] = round((merged[-1] + x) / 2, 1)
            else:
                merged.append(x)
        return merged

    @staticmethod
    def _fuse_line_anchors_with_aligns(aligns, anchors, tolerance):
        """将线条锚点融合到对齐点/边界列表中
        
        策略：
        1. 如果锚点与已有对齐点距离 ≤ tolerance*2 → 用锚点替换（锚点更精确）
        2. 如果锚点远离已有对齐点 → 作为新边界插入
        3. 确保首尾节点存在
        """
        if not anchors:
            return sorted(set(aligns))

        result = list(aligns) if isinstance(aligns, list) else list(aligns)
        result_set = set(result)

        for anchor in anchors:
            # 找最近的已有边界
            if result:
                nearest = min(result, key=lambda x: abs(x - anchor))
                if abs(nearest - anchor) <= tolerance * 2:
                    # 替换为更精确的锚点坐标
                    if nearest not in result_set:
                        continue
                    idx = result.index(nearest)
                    result[idx] = anchor
                    result_set.discard(nearest)
                    result_set.add(anchor)
                else:
                    # 作为新分割点插入
                    if anchor not in result_set:
                        result.append(anchor)
                        result_set.add(anchor)

        return sorted(set(result))

    def _cluster_1d(self, values, tolerance=4):
        """一维坐标聚簇，找出文本对齐位置（v2规格：最小簇大小=3）"""
        if not values:
            return []

        sorted_vals = sorted(values)
        clusters = []
        current_cluster = [sorted_vals[0]]

        for v in sorted_vals[1:]:
            if v - current_cluster[-1] <= tolerance:
                current_cluster.append(v)
            else:
                if len(current_cluster) >= 3:
                    clusters.append(sum(current_cluster) / len(current_cluster))
                current_cluster = [v]

        if len(current_cluster) >= 3:
            clusters.append(sum(current_cluster) / len(current_cluster))

        return clusters

    # ---- 网格填充 ----

    def _assign_words_to_grid(self, words, row_bounds, col_bounds):
        """将 words 分配到行列网格中（v2规格：重叠面积法）"""
        cfg = self.V2_CONFIG
        n_rows = len(row_bounds)
        n_cols = len(col_bounds) - 1

        if n_rows == 0 or n_cols == 0:
            return []

        grid = [[[] for _ in range(n_cols)] for _ in range(n_rows)]

        for w in words:
            wx0, wy0, wx1, wy1 = w["x0"], w["y0"], w["x1"], w["y1"]
            text = w["text"]

            if not text.strip():
                continue

            # 行分配
            row_idx = None
            center_y = (wy0 + wy1) / 2
            margin = (row_bounds[0][1] - row_bounds[0][0]) * cfg["row_margin_factor"] if row_bounds else 0
            for r, (y_top, y_bot) in enumerate(row_bounds):
                if (y_top - margin) <= center_y <= (y_bot + margin):
                    row_idx = r
                    break

            # 列分配：简单重叠法
            col_idx = None
            max_overlap = 0

            for c in range(n_cols):
                col_left = col_bounds[c]
                col_right = col_bounds[c + 1]
                
                overlap = max(0.0, min(wx1, col_right) - max(wx0, col_left))
                if overlap > max_overlap:
                    max_overlap = overlap
                    col_idx = c
            
            # 兜底：如果没有任何重叠，使用最近列中心
            if col_idx is None:
                center_x = (wx0 + wx1) / 2
                min_dist = float('inf')
                for c in range(n_cols):
                    col_center = (col_bounds[c] + col_bounds[c + 1]) / 2
                    dist = abs(center_x - col_center)
                    if dist < min_dist:
                        min_dist = dist
                        col_idx = c

            if row_idx is not None and col_idx is not None:
                grid[row_idx][col_idx].append(text)

        # 合并单元格文本
        result = []
        for r in range(n_rows):
            row_data = []
            for c in range(n_cols):
                cell_texts = grid[r][c]
                if cell_texts:
                    row_data.append(" ".join(cell_texts))
                else:
                    row_data.append("")
            result.append(row_data)

        return result

    # ---- 置信度评分 ----

    def _compute_table_confidence(self, table_data, has_border, page_words):
        """计算表格提取结果的置信度"""
        import statistics
        cfg = self.V2_CONFIG

        if not table_data or len(table_data) < 2:
            return 0.0

        scores = []

        # 因子1: 列数一致性
        col_counts = [len(row) for row in table_data if row]
        if col_counts and len(col_counts) >= 2:
            mean_cols = statistics.mean(col_counts)
            cv = statistics.stdev(col_counts) / mean_cols if mean_cols > 0 else 1.0
            col_consistency = max(0.0, 1.0 - cv * 2)
            scores.append((col_consistency, cfg["confidence_col_weight"]))
        else:
            scores.append((0.5, cfg["confidence_col_weight"]))

        # 因子2: 空值率
        total_cells = sum(len(row) for row in table_data)
        empty_cells = sum(1 for row in table_data for cell in row if not str(cell).strip())
        empty_ratio = empty_cells / max(total_cells, 1)
        if empty_ratio < 0.05:
            empty_score = 0.7
        elif empty_ratio > 0.5:
            empty_score = 0.3
        else:
            empty_score = 1.0 - empty_ratio
        scores.append((empty_score, cfg["confidence_empty_weight"]))

        # 因子3: 数值占比
        def is_numeric(text):
            text = str(text).strip().replace(",", "").replace("(", "-").replace(")", "")
            if not text:
                return False
            try:
                float(text)
                return True
            except:
                if text.endswith("%"):
                    try:
                        float(text[:-1])
                        return True
                    except:
                        return False
                return False

        numeric_count = sum(1 for row in table_data for cell in row
                            if is_numeric(str(cell).strip()))
        numeric_ratio = numeric_count / max(total_cells, 1)
        numeric_score = min(numeric_ratio * 2, 1.0) if numeric_ratio < 0.5 else 1.0
        scores.append((numeric_score, cfg["confidence_num_weight"]))

        # 加权综合
        weighted_sum = sum(s * w for s, w in scores)
        weighted_total = sum(w for _, w in scores)
        confidence = weighted_sum / weighted_total
        if has_border:
            confidence += cfg["confidence_line_bonus"]

        return min(1.0, max(0.0, confidence))

    # ---- 合并单元格视觉恢复（Step 2） ----

    def _detect_merge_cells_from_lines(self, drawings, row_bounds, col_bounds):
        """从表格线检测合并单元格（v2 Step 2：视觉线索）
        
        原理：
        - 若某列边界在特定行范围内缺少竖线 → 该行此处为 colspan 合并
        - 若某行边界在特定列范围内缺少横线 → 该列此处为 rowspan 合并
        
        Args:
            drawings: get_drawings() 提取的线条列表
            row_bounds: [(y_top, y_bottom), ...] 行边界
            col_bounds: [x0, x1, ...] 列分割线
        
        Returns:
            [(row, col, rowspan, colspan, confidence), ...]
        """
        if not drawings or len(row_bounds) < 2 or len(col_bounds) < 3:
            return []

        n_rows = len(row_bounds)
        n_cols = len(col_bounds) - 1
        line_merge_tol = self.V2_CONFIG.get("line_merge_tolerance", 2.0)

        # 提取线条
        v_lines = [(d["x0"], d["y0"], d["y1"]) for d in drawings
                    if d["type"] == "line" and d["direction"] == "v"]
        h_lines = [(d["y0"], d["x0"], d["x1"]) for d in drawings
                    if d["type"] == "line" and d["direction"] == "h"]

        merge_spans = []

        # ---- 检测横向合并（colspan） ----
        # 把竖线按 x 坐标合并去重
        v_x_groups = {}
        for x, y0, y1 in v_lines:
            matched = None
            for gx in v_x_groups:
                if abs(x - gx) <= line_merge_tol * 2:
                    matched = gx
                    break
            if matched is not None:
                v_x_groups[matched].append((y0, y1))
            else:
                v_x_groups[x] = [(y0, y1)]

        # 只有存在内部竖线时才信任线条检测（无线表格不适用）
        has_inner_v_lines = any(
            min(cx for cx in v_x_groups) < col_bounds[-1] * 0.9
            for _ in v_x_groups
        ) if v_x_groups else False

        for r, (ry0, ry1) in enumerate(row_bounds):
            row_center_y = (ry0 + ry1) / 2
            row_height = ry1 - ry0

            # 统计当前行中各列边界位置的竖线覆盖情况
            missing_boundaries = []
            for c in range(1, n_cols):
                boundary_x = col_bounds[c]
                has_line = False
                for gx, segs in v_x_groups.items():
                    if abs(gx - boundary_x) <= line_merge_tol * 3:
                        for sy0, sy1 in segs:
                            overlap = max(0.0, min(sy1, ry1) - max(sy0, ry0))
                            if overlap > row_height * 0.6:
                                has_line = True
                                break
                        if has_line:
                            break
                if not has_line:
                    missing_boundaries.append(c)

            if not missing_boundaries:
                continue

            # 分组连续缺失边界 → 每组对应一个可能的 colspan 合并
            groups = []
            group_start = missing_boundaries[0]
            group_prev = missing_boundaries[0]
            for b in missing_boundaries[1:]:
                if b == group_prev + 1:
                    group_prev = b
                else:
                    groups.append((group_start, group_prev))
                    group_start = b
                    group_prev = b
            groups.append((group_start, group_prev))

            for gs, ge in groups:
                # 起始列：缺失边界左侧的第一列
                start_col = gs - 1
                if start_col < 0:
                    start_col = 0
                # 结束边界索引：缺失边界的最后一个 + 1
                end_boundary = ge + 1
                span_cols = end_boundary - start_col

                # 排除整行合并和单列
                if 2 <= span_cols < n_cols:
                    merge_spans.append((r, start_col, 1, span_cols, 0.85))

        # ---- 检测纵向合并（rowspan） ----
        h_y_groups = {}
        for y, x0, x1 in h_lines:
            matched = None
            for gy in h_y_groups:
                if abs(y - gy) <= line_merge_tol * 2:
                    matched = gy
                    break
            if matched is not None:
                h_y_groups[matched].append((x0, x1))
            else:
                h_y_groups[y] = [(x0, x1)]

        # 只有存在内部横线时才信任线条检测
        has_inner_h_lines = any(
            min(gy for gy in h_y_groups) < row_bounds[-1][1] * 0.9
            for _ in h_y_groups
        ) if h_y_groups else False

        for c in range(n_cols):
            cx0, cx1 = col_bounds[c], col_bounds[c + 1]
            col_width = cx1 - cx0

            # 统计当前列中各行边界位置的横线覆盖情况
            missing_boundaries = []
            for r in range(1, n_rows):
                boundary_y = row_bounds[r][0]
                has_line = False
                for gy, segs in h_y_groups.items():
                    if abs(gy - boundary_y) <= line_merge_tol * 3:
                        for sx0, sx1 in segs:
                            overlap = max(0.0, min(sx1, cx1) - max(sx0, cx0))
                            if overlap > col_width * 0.6:
                                has_line = True
                                break
                        if has_line:
                            break
                if not has_line:
                    missing_boundaries.append(r)

            if not missing_boundaries:
                continue

            # 排除整列无横线（说明列宽范围内根本没有横线覆盖）
            if len(missing_boundaries) >= n_rows - 1:
                continue

            # 分组连续缺失
            groups = []
            gs = missing_boundaries[0]
            gp = missing_boundaries[0]
            for b in missing_boundaries[1:]:
                if b == gp + 1:
                    gp = b
                else:
                    groups.append((gs, gp))
                    gs = b
                    gp = b
            groups.append((gs, gp))

            for gs, ge in groups:
                start_r = gs - 1  # 缺失边界的上方行
                if start_r < 0:
                    start_r = 0
                span_rows = (ge + 1) - start_r
                if 2 <= span_rows < n_rows:
                    merge_spans.append((start_r, c, span_rows, 1, 0.8))

        # 去重 & 合并重叠 span
        return self._merge_overlapping_spans(merge_spans, n_rows, n_cols)

    def _detect_merge_cells_from_text(self, table_data):
        """从文本模式检测合并单元格（v2 Step 2：文本重复模式）
        
        检测模式：
        1. 相邻行同一列内容完全相同 → rowspan（纵向合并）
        2. 同行连续空单元格 → 可能 colspan（需配合线条检测确认）
        
        Args:
            table_data: 当前网格填充结果 List[List[str]]
        
        Returns:
            [(row, col, rowspan, colspan, confidence), ...]
        """
        if not table_data or len(table_data) < 2:
            return []

        n_rows = len(table_data)
        n_cols = max(len(row) for row in table_data) if table_data else 0
        if n_cols < 2:
            return []

        merge_spans = []

        # ---- 模式1：相邻行同一列内容完全相同 → rowspan ----
        for c in range(n_cols):
            r = 0
            while r < n_rows - 1:
                cur_val = self._safe_cell(table_data, r, c)
                if not cur_val or len(cur_val) < 2:
                    r += 1
                    continue

                # 向下查找相同内容的连续行
                span_rows = 1
                for nr in range(r + 1, n_rows):
                    next_val = self._safe_cell(table_data, nr, c)
                    if next_val == cur_val:
                        span_rows += 1
                    else:
                        break

                if span_rows >= 2:
                    merge_spans.append((r, c, span_rows, 1, 0.7))
                    r += span_rows
                else:
                    r += 1

        # ---- 模式2：表头行连续空单元格 → 可能 colspan ----
        # 只检查前几行（表头区域）
        header_rows = min(n_rows, max(2, n_rows // 3))
        for r in range(header_rows):
            c = 0
            while c < n_cols:
                cur_val = self._safe_cell(table_data, r, c)
                if cur_val and len(cur_val) >= 2:
                    # 向右查找连续空单元格
                    empty_count = 0
                    for nc in range(c + 1, n_cols):
                        next_val = self._safe_cell(table_data, r, nc)
                        if not next_val or len(next_val.strip()) == 0:
                            empty_count += 1
                        else:
                            break

                    if empty_count >= 1:
                        merge_spans.append((r, c, 1, empty_count + 1, 0.55))
                        c += empty_count + 1
                    else:
                        c += 1
                else:
                    c += 1

        # 去重
        return self._merge_overlapping_spans(merge_spans, n_rows, n_cols)

    @staticmethod
    def _safe_cell(table_data, row, col):
        """安全获取单元格值"""
        if 0 <= row < len(table_data) and 0 <= col < len(table_data[row]):
            return str(table_data[row][col]).strip()
        return ""

    @staticmethod
    def _merge_overlapping_spans(spans, n_rows, n_cols):
        """合并重叠的合并单元格 span，去重并处理冲突
        
        策略：
        1. 线条检测优先（高置信度）
        2. 合并相邻的有重叠的 span
        3. 限制 span 不超过表格边界
        """
        if not spans:
            return []

        # 按置信度降序排列（高优先级先处理）
        sorted_spans = sorted(spans, key=lambda s: s[4], reverse=True)

        # 占用网格
        occupied = [[False] * n_cols for _ in range(n_rows)]
        result = []

        for row, col, rowspan, colspan, conf in sorted_spans:
            # 限制范围
            rowspan = min(rowspan, n_rows - row)
            colspan = min(colspan, n_cols - col)

            if rowspan < 1 or colspan < 1:
                continue
            if rowspan == 1 and colspan == 1:
                continue  # 不算合并

            # 检查是否与已有 span 冲突
            conflict = False
            for dr in range(rowspan):
                for dc in range(colspan):
                    if occupied[row + dr][col + dc]:
                        conflict = True
                        break
                if conflict:
                    break

            if not conflict:
                result.append((row, col, rowspan, colspan, round(conf, 2)))
                for dr in range(rowspan):
                    for dc in range(colspan):
                        occupied[row + dr][col + dc] = True

        return sorted(result, key=lambda s: (s[0], s[1]))

    def _apply_merge_cells(self, table_data, merge_spans):
        """将检测到的合并 span 应用到表格数据
        
        对于每个合并 span：
        - 保留 (row, col) 位置的值
        - 将被合并的单元格设为 None 或合并标记
        
        Args:
            table_data: 2D 表格数据
            merge_spans: [(row, col, rowspan, colspan, confidence), ...]
        
        Returns:
            (modified_table, merge_info_dict)
            merge_info_dict: {(row, col): {"rowspan": n, "colspan": m, "confidence": c}}
        """
        if not merge_spans:
            return table_data, {}

        n_rows = len(table_data)
        n_cols = max(len(row) for row in table_data) if table_data else 0

        # 先补齐所有行到相同列数
        normalized = []
        for row in table_data:
            r = list(row)
            while len(r) < n_cols:
                r.append("")
            normalized.append(r)

        merge_info = {}
        for row, col, rowspan, colspan, conf in merge_spans:
            if rowspan <= 1 and colspan <= 1:
                continue
            merge_info[(row, col)] = {
                "rowspan": rowspan,
                "colspan": colspan,
                "confidence": conf,
            }
            # 清空被合并的单元格（保留原始位置的文本）
            for dr in range(rowspan):
                for dc in range(colspan):
                    if dr == 0 and dc == 0:
                        continue
                    if row + dr < n_rows and col + dc < n_cols:
                        normalized[row + dr][col + dc] = ""

        return normalized, merge_info

    # ---- 合并单元格检测主入口 ----

    def _detect_and_apply_merge_cells(self, table_data, drawings, row_bounds, col_bounds):
        """合并单元格检测与恢复主入口（v2 Step 2）
        
        三阶段策略：
        1. 线条视觉检测（最高置信度）
        2. 文本模式检测（辅助）
        3. 应用合并
        
        Returns:
            (table_data_with_merge, merge_info, stats)
            stats: {"line_spans": N, "text_spans": N, "total_spans": N, "cells_merged": N}
        """
        # 阶段1：线条视觉检测
        line_spans = self._detect_merge_cells_from_lines(drawings, row_bounds, col_bounds)

        # 阶段2：文本模式检测
        text_spans = self._detect_merge_cells_from_text(table_data)

        # 合并两阶段结果（线条优先）
        all_spans = line_spans + text_spans
        n_rows = len(table_data)
        n_cols = max(len(row) for row in table_data) if table_data else 0
        merged_spans = self._merge_overlapping_spans(all_spans, max(n_rows, 1), max(n_cols, 1))

        # 阶段3：应用合并
        modified_table, merge_info = self._apply_merge_cells(table_data, merged_spans)

        stats = {
            "line_spans": len(line_spans),
            "text_spans": len(text_spans),
            "total_spans": len(merged_spans),
            "cells_merged": sum(rs * cs - 1 for _, _, rs, cs, _ in merged_spans),
        }

        return modified_table, merge_info, stats


# ============================================================
# LLM视觉识别模块
# ============================================================
class VisionLLM:
    """视觉大模型接口"""

    def __init__(self, api_key=None, endpoint=None, model=None):
        self.config = load_config()
        self.api_key = api_key or self.config.get("doubao_api_key", "")
        self.endpoint = endpoint or self.config.get("doubao_endpoint", "ark.cn-beijing.volces.com")
        self.model = model or self.config.get("doubao_model", "doubao-pro-32k")

    def test_connection(self):
        """测试API连接"""
        import requests
        api_url = f"https://{self.endpoint}/api/v3/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        data = {
            "model": self.model,
            "messages": [{"role": "user", "content": "Hi"}],
            "max_tokens": 10
        }
        try:
            resp = requests.post(api_url, headers=headers, json=data, timeout=15)
            resp.raise_for_status()
            return True, "API连接成功！"
        except requests.exceptions.Timeout:
            return False, "连接超时，请检查网络或API地址"
        except requests.exceptions.RequestException as e:
            return False, f"连接失败: {str(e)}"

    def recognize_table(self, image_path):
        """识别图片中的表格"""
        if not self.api_key:
            return {"success": False, "error": "未配置API Key"}

        import base64
        import requests

        with open(image_path, 'rb') as f:
            img_data = base64.b64encode(f.read()).decode()

        api_url = f"https://{self.endpoint}/api/v3/chat/completions"

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }

        prompt = """请识别这张银行年报截图中的财务表格，以JSON数组格式输出。

要求：
1. 只输出JSON数组，不要其他文字
2. 每行数据是一个数组，格式如：["项目名称", "2023年", "2022年", "同比增减"]
3. 保持原表格的行列结构
4. 数字去掉逗号，保留原数值
5. 表头也要包含在结果中

输出示例格式：
[["项目", "2023年末", "2022年末", "变动率"], ["流动资产", "100,000", "90,000", "11.11%"], ...]"""

        data = {
            "model": self.model,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_data}"}}
                    ]
                }
            ],
            "max_tokens": 4000,
            "temperature": 0.1
        }

        try:
            resp = requests.post(api_url, headers=headers, json=data, timeout=120)
            resp.raise_for_status()
            result = resp.json()

            if 'choices' in result and len(result['choices']) > 0:
                content = result['choices'][0]['message']['content']
                return {"success": True, "data": content}
            else:
                return {"success": False, "error": "API返回格式错误"}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def batch_recognize(self, image_paths, progress_callback=None):
        """批量识别图片"""
        results = []

        for i, img_path in enumerate(image_paths):
            if progress_callback:
                progress_callback(i + 1, len(image_paths))

            result = self.recognize_table(img_path)
            results.append({
                "page": i + 1,
                "image": img_path,
                "result": result
            })

            if i < len(image_paths) - 1:
                time.sleep(1)

        return results


# ============================================================
# 表格上下文LLM模块 - 为表格生成名称和摘要
# ============================================================
class TableContextLLM:
    """文本型LLM接口 - 使用 deepseek-v4 等模型为表格生成名称和摘要"""

    MAX_PREVIEW_ROWS = 5          # 表格预览最多行数
    MAX_PREVIEW_CELL_LEN = 30     # 单格最大字符数

    def __init__(self, api_key=None, endpoint=None, model=None):
        self.config = load_config()
        self.api_key = api_key or self.config.get("deepseek_api_key", "")
        self.endpoint = endpoint or self.config.get("deepseek_endpoint", "api.deepseek.com")
        self.model = model or self.config.get("deepseek_model", "deepseek-chat")

    def test_connection(self):
        """测试 deepseek API 连通性"""
        import requests
        api_url = f"https://{self.endpoint}/v1/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        data = {
            "model": self.model,
            "messages": [{"role": "user", "content": "Hi"}],
            "max_tokens": 10
        }
        try:
            resp = requests.post(api_url, headers=headers, json=data, timeout=15)
            resp.raise_for_status()
            return True, "DeepSeek API 连接成功！"
        except requests.exceptions.Timeout:
            return False, "连接超时，请检查网络或API地址"
        except requests.exceptions.HTTPError as e:
            return False, f"HTTP错误: {e.response.status_code} - {e.response.reason}"
        except requests.exceptions.RequestException as e:
            return False, f"连接失败: {str(e)}"

    @staticmethod
    def _build_table_preview(table_data, max_rows=None, max_cell_len=None):
        """将表格数据压缩为文本预览（用于 LLM prompt）"""
        if max_rows is None:
            max_rows = TableContextLLM.MAX_PREVIEW_ROWS
        if max_cell_len is None:
            max_cell_len = TableContextLLM.MAX_PREVIEW_CELL_LEN

        if not table_data:
            return "（表格数据为空）"

        lines = []
        for row_idx, row in enumerate(table_data):
            if row_idx >= max_rows:
                lines.append(f"... (共{len(table_data)}行，仅显示前{max_rows}行)")
                break
            cells = []
            for cell in row:
                s = str(cell).strip()
                if len(s) > max_cell_len:
                    s = s[:max_cell_len - 3] + "..."
                cells.append(s)
            lines.append(" | ".join(cells))

        return "\n".join(lines)

    def generate_table_name(self, context_text, table_data):
        """将上下文文本 + 表格摘要发给 LLM，返回 {"title": str, "summary": str}"""
        if not self.api_key:
            return {"title": "", "summary": "", "error": "未配置 DeepSeek API Key"}

        import json as _json
        import requests

        table_preview = self._build_table_preview(table_data)

        prompt = f"""你是一个财务文档分析专家。下面是一份银行年报中提取的表格数据。
请根据表格上方的描述文字和表格内容，为这个表格生成一个规范的名称和一句话摘要。

表格上方描述文字：
{context_text if context_text else "（无上方描述文字）"}

表格内容预览（前{self.MAX_PREVIEW_ROWS}行）：
{table_preview}

请以JSON格式返回，只返回JSON，不要其他文字：
{{"title": "表格的规范名称（如'合并资产负债表'、'利润表'）", "summary": "一句话描述该表格的主要内容（不超过50字）"}}"""

        api_url = f"https://{self.endpoint}/v1/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        data = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": "你是一个专业的财务表格分析助手，擅长为财务表格生成规范的标题和摘要。请只返回JSON格式的结果。"},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 500,
            "temperature": 0.1
        }

        try:
            resp = requests.post(api_url, headers=headers, json=data, timeout=30)
            resp.raise_for_status()
            result = resp.json()

            if 'choices' in result and len(result['choices']) > 0:
                content = result['choices'][0]['message']['content']
                # 尝试解析 JSON
                try:
                    parsed = _json.loads(content)
                    return {
                        "title": parsed.get("title", ""),
                        "summary": parsed.get("summary", "")
                    }
                except _json.JSONDecodeError:
                    # 尝试从文本中提取 JSON
                    import re
                    match = re.search(r'\{[^}]*\}', content, re.DOTALL)
                    if match:
                        try:
                            parsed = _json.loads(match.group())
                            return {
                                "title": parsed.get("title", ""),
                                "summary": parsed.get("summary", "")
                            }
                        except _json.JSONDecodeError:
                            pass
                    return {"title": content.strip()[:100], "summary": "", "error": "JSON解析失败，已用原始返回"}
            else:
                return {"title": "", "summary": "", "error": "API返回格式错误"}

        except requests.exceptions.Timeout:
            return {"title": "", "summary": "", "error": "请求超时"}
        except Exception as e:
            return {"title": "", "summary": "", "error": str(e)}

    def batch_generate(self, tables, progress_callback=None):
        """批量处理多个表格，返回每个表格的命名结果列表

        Args:
            tables: [{"context_text": str, "data": [[...], ...]}, ...]
            progress_callback: callable(current, total, message)

        Returns:
            [{"title": str, "summary": str, "error": str or None}, ...]
        """
        results = []
        total = len(tables)

        for i, table in enumerate(tables):
            context_text = table.get("context_text", "")
            table_data = table.get("data", [])

            if progress_callback:
                progress_callback(i + 1, total, f"正在为第{i+1}个表格生成名称...")

            result = self.generate_table_name(context_text, table_data)
            results.append(result)

            if i < total - 1:
                time.sleep(0.3)  # 限速，避免触发 API 频率限制

        return results


# ============================================================
# Excel导出模块
# ============================================================
class ExcelExporter:
    """Excel导出器"""

    @staticmethod
    def parse_json_table(json_str):
        """解析LLM返回的JSON字符串"""
        import json
        import re

        try:
            return json.loads(json_str)
        except:
            pass

        match = re.search(r'\[.*\]', json_str, re.DOTALL)
        if match:
            try:
                return json.loads(match.group())
            except:
                pass

        return None

    @staticmethod
    def export_tables(tables_data, output_path):
        """将表格数据导出为Excel——每个表格一个独立Sheet。

        Sheet 命名: P{页码}-T{序号}-{标题前若干字}
        表头区: 标题行(粗体) + 来源行(灰色) + 空白行
        数据区: 左右各留1列空白，上下各留1行空白
        """
        wb = Workbook()

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 过滤掉空数据表格，按页码+表格序号排序
        valid_tables = [t for t in tables_data if t.get("data")]
        if not valid_tables:
            # 创建一个空 sheet 避免报错
            ws = wb.active
            ws.title = "无数据"
            ws.cell(row=1, column=1, value="未提取到表格数据")
            wb.save(output_path)
            return True

        # 按页排序，同页按 extractor（docx 优先）
        def sort_key(t):
            page = t.get("page", 0)
            ext = t.get("extractor", "")
            is_docx = 0 if ext.startswith("docx") else 1
            return (page, is_docx)

        valid_tables.sort(key=sort_key)

        # 按页分组，分配每页内序号
        page_tables = {}
        for t in valid_tables:
            page = t.get("page", 0)
            if page not in page_tables:
                page_tables[page] = []
            page_tables[page].append(t)

        # 计算最大页码位数，用于零补位
        max_page = max(page_tables.keys()) if page_tables else 1
        page_digits = max(3, len(str(max_page)))  # 至少 3 位

        # 计算全局最大本页表数，用于零补位
        max_per_page = max(len(v) for v in page_tables.values()) if page_tables else 1
        seq_digits = max(2, len(str(max_per_page)))  # 至少 2 位

        global_idx = 0
        for page in sorted(page_tables.keys()):
            tables_on_page = page_tables[page]
            for seq, table_info in enumerate(tables_on_page, 1):
                table_data = table_info.get("data", [])
                if not table_data:
                    continue

                # ---- Sheet 命名（零补位确保字符串排序=数值排序）----
                # 优先级：llm_title > title > 自动提取
                title = table_info.get("llm_title") or table_info.get("title") or PDFProcessor._extract_table_title(table_data)
                sheet_name = f"P{page:0{page_digits}d}-T{seq:0{seq_digits}d}-{title}"
                # Sheet 名最长 31 字符
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:28] + "..."
                # 替换 Excel Sheet 名非法字符: \ / * ? : [ ]
                for ch in r'\/:*?[]':
                    sheet_name = sheet_name.replace(ch, "-")

                if global_idx == 0:
                    ws = wb.active
                else:
                    ws = wb.create_sheet()
                ws.title = sheet_name
                global_idx += 1

                # 计算列布局
                max_cols = 0
                for row in table_data:
                    max_cols = max(max_cols, len(row))
                data_start_col = 2  # 左侧留白
                data_end_col = data_start_col + max_cols  # 数据结束列（也是右侧留白）

                row_num = 1

                # ---- 表头区 ----
                # 标题行
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=data_end_col)
                title_cell = ws.cell(row=row_num, column=1, value=title)
                title_cell.font = Font(bold=True, size=12)
                title_cell.alignment = Alignment(horizontal="left", vertical="center")
                row_num += 1

                # 来源行
                ext_label = table_info.get("extractor", "unknown")
                ext_display = "docx精准通道" if ext_label.startswith("docx") else "V2快速通道"
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=data_end_col)
                source_cell = ws.cell(row=row_num, column=1,
                                      value=f"来源: PDF第{page}页 | 提取方式: {ext_display}")
                source_cell.font = Font(color="808080", size=9)
                source_cell.alignment = Alignment(horizontal="left", vertical="center")
                row_num += 1

                # 标题与数据之间的空行
                row_num += 1

                # ---- 数据区：顶部空白行 ----
                for col in range(1, data_end_col + 1):
                    cell = ws.cell(row=row_num, column=col, value="")
                    cell.border = thin_border
                row_num += 1

                # ---- 写入表格数据 ----
                for row in table_data:
                    # 左侧空白列
                    cell = ws.cell(row=row_num, column=1, value="")
                    cell.border = thin_border

                    for col_idx, value in enumerate(row):
                        col = data_start_col + col_idx
                        cell = ws.cell(row=row_num, column=col, value=str(value) if value is not None else "")
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    # 右侧空白列
                    cell = ws.cell(row=row_num, column=data_end_col, value="")
                    cell.border = thin_border

                    row_num += 1

                # ---- 底部空白行 ----
                for col in range(1, data_end_col + 1):
                    cell = ws.cell(row=row_num, column=col, value="")
                    cell.border = thin_border
                row_num += 1

                # ---- 自动列宽 ----
                for col_idx in range(1, data_end_col + 1):
                    max_length = 0
                    col_letter = get_column_letter(col_idx)
                    for r in range(1, row_num):
                        cv = ws.cell(row=r, column=col_idx).value
                        if cv:
                            max_length = max(max_length, len(str(cv)))
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(output_path)
        return True


# ============================================================
# 自动清洗 & 合并 工具函数（数据处理层，无 UI 依赖）
# ============================================================

def _remove_spaces_data(data):
    """智能删除空单元格：全局对齐裁剪左右两端空列。

    策略：
    - 扫描所有行，找到全局最左非空列索引和最右非空列索引
    - 所有行统一裁剪到 [left, right+1] 范围
    - 保留内部空列结构不变（中间的空白列会被保留）

    Returns:
        新数组（list of lists）
    """
    if not data:
        return []

    # 找到全局左右边界
    left = len(data[0]) if data else 0
    right = 0
    for row in data:
        for i, cell in enumerate(row):
            if cell is not None and str(cell).strip():
                left = min(left, i)
                right = max(right, i)

    # 如果全是空行，保留原样
    if left > right:
        return [list(row) for row in data]

    # 所有行统一裁剪到全局边界
    result = []
    for row in data:
        trimmed = row[left:right + 1]
        result.append(list(trimmed))

    return result


def _clean_data_cells(data):
    """清洗数值类单元格：含中文或英文的跳过，其余移除非数值字符。原地修改。

    纯横线单元格（--、---、—等）视为省略标记，保留不动。
    横线与数字混合时（如 ---100），横线被清洗掉。
    """
    import re
    for row in data:
        for c in range(len(row)):
            cell = row[c]
            if not cell:
                continue
            text = str(cell).strip()
            if not text:
                row[c] = ""
                continue
            # 省略标记（只含横线/破折号，无文本数字）→ 保留不动
            if re.fullmatch(r'[\-\u2014\u2015\u2500\uFF0D]+', text):
                row[c] = text
                continue
            # 含中文或英文字母 → 跳过
            if re.search(r'[\u4e00-\u9fff]|[a-zA-Z]', text):
                row[c] = text
                continue
            # 清洗：保留数字、逗号、小数点、负号、括号、百分号、空格
            # （空格必须保留：列边界缺失时可能合并为 "V1 V2"，后续由
            #  _split_merged_numeric_cells 根据空格拆分；仅删下划线等杂质）
            cleaned = re.sub(r'[^0-9,\.\-\(\)\% ]', '', text).strip()
            row[c] = cleaned
    return data


def _is_numeric_data(text):
    """判断文本是否为纯数值类（含数字格式字符），用于检测无表头数据行。"""
    if not text or not str(text).strip():
        return False
    import re
    t = str(text).strip()
    # 只要不含中文或英文字母，且至少有一个数字，就认为是数值
    if re.search(r'[\u4e00-\u9fff]|[a-zA-Z]', t):
        return False
    return bool(re.search(r'[0-9]', t))


def _remove_empty_rows(data):
    """删除表格中的全空行。只含空字符串或 None 的行会被移除。原地修改。"""
    if not data:
        return data
    data[:] = [
        row for row in data
        if any(cell is not None and str(cell).strip() for cell in row)
    ]
    return data


def _compact_empty_cells(data):
    """删除整列为空的列（所有行同步），压缩为紧凑列。打标记 cells_compacted。

    与 _remove_spaces_data 不同：后者只删两端空列，此函数删除所有整列为空的列。
    关键改进：检测整列为空的列，所有行同步删除，保证列方向对齐。

    Returns:
        (cleaned_data, compacted): 去空列后的数据 + 是否压缩了
    """
    if not data:
        return data, False

    max_cols = max((len(row) for row in data), default=0)
    if max_cols == 0:
        return data, False

    # 1. 先规范化：所有行补齐到最大列数
    normalized = []
    for row in data:
        nr = list(row)
        while len(nr) < max_cols:
            nr.append("")
        normalized.append(nr)

    # 2. 辅助：判断是否为跨列标题行（仅首列有内容）
    def _is_header_only(row):
        if not row or len(row) < 2:
            return False
        first = row[0]
        if first is None or not str(first).strip():
            return False
        return all(
            cell is None or not str(cell).strip()
            for cell in row[1:]
        )

    # 3. 区分标题行和数据行，仅基于数据行检测空列
    header_rows = {r for r in range(len(normalized)) if _is_header_only(normalized[r])}
    data_rows = [r for r in range(len(normalized)) if r not in header_rows]

    if not data_rows:
        return data, False

    # 4. 检测整列为空的列索引（仅基于数据行）
    cols_to_remove = set()
    for c in range(max_cols):
        is_empty_col = True
        for r in data_rows:
            cell = normalized[r][c]
            if cell is not None and str(cell).strip():
                is_empty_col = False
                break
        if is_empty_col:
            cols_to_remove.add(c)

    if not cols_to_remove:
        return data, False

    # 5. 所有行同步删除空列
    cols_to_keep = [c for c in range(max_cols) if c not in cols_to_remove]

    if not cols_to_keep:
        # 极端情况：所有列都空 → 保留原数据
        return data, False

    result = []
    for row in normalized:
        result.append([row[c] for c in cols_to_keep])

    return result, True


def _remove_sparse_columns(data, fill_threshold=0.15, min_nonempty=2):
    """删除数据行填充率极低的稀疏列（比 _compact_empty_cells 更宽松）。

    与 _compact_empty_cells 的区别：
    - _compact_empty_cells：只删 100% 为空的列
    - _remove_sparse_columns：删除填充率 < 有效阈值的列（含少量表头溢出）

    保护规则：
    - 自动识别表头行（仅首列有内容 OR 前3行默认视为表头区）
    - 不删除数据行中唯一非空列，避免丢失行标签
    - 删除前列内容自动合并到左侧相邻保留列，避免内容丢失

    自适应阈值：
    - 当列的非空数据行数 ≤ min_nonempty 时，使用 50% 作为有效阈值
      （防止少量异常行如脚注编号导致整列被误保留）
    - 否则使用 fill_threshold（默认 15%）

    Args:
        data: list[list[str]]
        fill_threshold: 数据行填充率阈值（默认 0.15 = 15%），仅非空数 > min_nonempty 时生效
        min_nonempty: 最少非空数据行数（默认 2），低于或等于此数时触发自适应阈值

    Returns:
        (cleaned_data, removed_count): 去稀疏列后的数据 + 删除的列数
    """
    if not data or len(data) < 2:
        return data, 0

    max_cols = max((len(row) for row in data), default=0)
    if max_cols == 0:
        return data, 0

    # 1. 规范化所有行到相同列数
    normalized = []
    for row in data:
        nr = list(row)
        while len(nr) < max_cols:
            nr.append("")
        normalized.append(nr)

    total_rows = len(normalized)

    # 2. 辅助：判断是否为跨列标题行（仅首列有内容）
    def _is_header_only(row):
        if not row or len(row) < 2:
            return False
        first = row[0]
        if first is None or not str(first).strip():
            return False
        return all(
            cell is None or not str(cell).strip()
            for cell in row[1:]
        )

    # 3. 识别表头行集（前3行默认视为表头区 + 仅首列有内容的行）
    header_row_set = set()
    header_row_set.update(range(min(3, total_rows)))
    for r in range(total_rows):
        if _is_header_only(normalized[r]):
            header_row_set.add(r)

    data_rows = [r for r in range(total_rows) if r not in header_row_set]
    if not data_rows:
        return data, 0

    # 4. 计算每列的数据行填充率
    cols_to_remove = set()
    for c in range(max_cols):
        nonempty_count = 0
        for r in data_rows:
            cell = normalized[r][c]
            if cell is not None and str(cell).strip():
                nonempty_count += 1

        fill_rate = nonempty_count / len(data_rows) if data_rows else 0

        # 当绝对非空数极少时（≤ min_nonempty），使用更慷慨的阈值（50%），
        # 防止少量异常行（如脚注编号、分隔文字）导致整列被误保留。
        # 例：5行数据中仅2行有内容 → 40% < 50% → 删除 ✓
        #     3行数据中2行有内容 → 67% > 50% → 保留 ✓（小表不误删）
        effective_threshold = 0.5 if nonempty_count <= min_nonempty else fill_threshold
        if fill_rate < effective_threshold:
            cols_to_remove.add(c)

    if not cols_to_remove:
        return data, 0

    # 5. 安全检查：不删除"行唯一非空列"（防止丢失标签）
    protected_cols = set()
    for r in data_rows:
        # 统计该行在待保留列中的非空单元格数
        nonempty_in_keeping = sum(
            1 for c in range(max_cols)
            if c not in cols_to_remove
            and normalized[r][c] is not None
            and str(normalized[r][c]).strip()
        )

        if nonempty_in_keeping > 0:
            continue

        # 如果该行在保留列中全空，则在待删列中保留第一个有内容的列
        for c in range(max_cols):
            if (normalized[r][c] is not None
                    and str(normalized[r][c]).strip()
                    and c in cols_to_remove):
                protected_cols.add(c)
                break

    cols_to_remove -= protected_cols

    if not cols_to_remove:
        return data, 0

    # 5.5 删除前合并：将待删列中的内容合并到左侧最近的非删除列，
    # 防止脚注编号、分隔文字等内容在列删除时丢失。
    # 例：Col 1 中 "(2)" 和 "7" → 合并到 Col 0
    kept_set = set(range(max_cols)) - cols_to_remove
    for c in sorted(cols_to_remove):
        for r in range(total_rows):
            cell_val = normalized[r][c]
            if cell_val is None or not str(cell_val).strip():
                continue
            # 找到左侧最近的保留列
            left_kept = None
            for lc in range(c - 1, -1, -1):
                if lc in kept_set:
                    left_kept = lc
                    break
            if left_kept is not None:
                existing = normalized[r][left_kept]
                existing_str = str(existing).strip() if existing is not None else ""
                merge_str = str(cell_val).strip()
                if existing_str:
                    normalized[r][left_kept] = existing_str + " " + merge_str
                else:
                    normalized[r][left_kept] = merge_str

    # 6. 所有行同步删除稀疏列
    cols_to_keep = [c for c in range(max_cols) if c not in cols_to_remove]
    if not cols_to_keep:
        return data, 0

    result = []
    for row in normalized:
        result.append([row[c] for c in cols_to_keep])

    return result, len(cols_to_remove)


def _deduplicate_columns(data):
    """删除因合并单元格展开导致的完全重复相邻列。

    pdf2docx 在展开跨列合并的标题格时，会把标题文本复制到每个子列，
    导致相邻多列内容完全相同。此函数检测并删除重复列。

    判定条件：如果两个相邻列在【所有行】中的值完全相同 → 列 c+1 被删除。
    从右向左扫描以避免索引偏移，迭代直到无重复列。

    Args:
        data: list[list[str]]

    Returns:
        (cleaned_data, removed_count): 去重后数据 + 删除的列数
    """
    if len(data) < 2:
        return data, 0

    max_cols = max((len(row) for row in data), default=0)
    if max_cols < 2:
        return data, 0

    removed_count = 0
    # 从右向左扫描，避免删除后索引偏移影响左侧比较
    for c in range(max_cols - 2, -1, -1):
        all_match = True
        for row in data:
            v1 = str(row[c]).strip() if c < len(row) and row[c] else ""
            v2 = str(row[c + 1]).strip() if c + 1 < len(row) and row[c + 1] else ""
            if v1 != v2:
                all_match = False
                break

        if all_match:
            for row in data:
                if c + 1 < len(row):
                    row.pop(c + 1)
            removed_count += 1

    return data, removed_count


def _remove_phantom_header_columns(data):
    """删除 pdf2docx 合并单元格展开导致的幽灵列（phantom columns）。

    pdf2docx 在展开跨列表头合并格时，会产生大量"表头文本相同但数据密度极低"
    的幽灵列。这些列夹在真实列之间，破坏列对齐，仅凭完全匹配的去重检测不到。

    算法：
    1. 规范化所有行到相同列数
    2. 按首行值将列分组（相同 row-0 值的列归为一组）
    3. 对每组（>1列），统计每列在数据行中的非空密度
    4. 密度低于组内最高密度的 25% → 视为幽灵列 → 删除
    5. 若组内所有列密度一致（差异<5%），视为真实子列，保留全部
    6. 所有行同步删除幽灵列

    安全边界：
    - 单列表 → 不处理（无法判定）
    - 不足 3 行 → 不处理（无足够数据行可统计）
    - 密度均匀组 → 保留全部（如多年度金额+占比子列均为真实列）
    - 密度阈值保守（最高密度的 25%），避免误删稀疏但合法的列

    Args:
        data: list[list[str]]，表格数据

    Returns:
        (cleaned_data, removed_count): 清洗后数据 + 删除的幽灵列数
    """
    if not data or len(data) < 3:
        return data, 0

    max_cols = max((len(row) for row in data), default=0)
    if max_cols < 2:
        return data, 0

    # ---- 1. 规范化所有行到相同列数 ----
    normalized = []
    for row in data:
        nr = list(row)
        while len(nr) < max_cols:
            nr.append("")
        normalized.append(nr[:max_cols])

    # ---- 2. 收集每列的 row-0 签名并分组 ----
    # 使用前 3 行的值组成复合签名（提升 multi-level header 的区分度）
    header_rows_count = min(3, len(normalized))
    col_signatures = []
    for c in range(max_cols):
        sig_parts = []
        for r in range(header_rows_count):
            v = str(normalized[r][c]).strip() if normalized[r][c] else ""
            sig_parts.append(v)
        col_signatures.append(tuple(sig_parts))

    # ---- 3. 按签名分组 ----
    from collections import defaultdict
    groups = defaultdict(list)
    for c, sig in enumerate(col_signatures):
        if any(s for s in sig):  # 跳过全空签名列（已由 _compact_empty_cells 处理）
            groups[sig].append(c)

    # ---- 4. 对每组（>1列）检测幽灵列 ----
    total_rows = len(normalized)
    data_start_row = min(3, total_rows - 1)  # 跳过表头区 0~2 行
    data_row_count = total_rows - data_start_row
    if data_row_count < 1:
        return data, 0

    phantom_cols = set()

    for sig, cols in groups.items():
        if len(cols) <= 1:
            continue

        # 计算该组每列的数据密度（表头行以下）
        densities = []
        for c in cols:
            non_empty = 0
            for r in range(data_start_row, total_rows):
                cell = normalized[r][c]
                if cell is not None and str(cell).strip():
                    non_empty += 1
            density = non_empty / data_row_count if data_row_count > 0 else 0
            densities.append(density)

        max_density = max(densities) if densities else 0

        # 若组内最高密度也很低（<10%），跳过——可能整个组都是表头装饰
        if max_density < 0.10:
            continue

        # 计算组内密度标准差，判断是否均匀散布
        if len(densities) >= 2:
            avg = sum(densities) / len(densities)
            variance = sum((d - avg) ** 2 for d in densities) / len(densities)
            std_dev = variance ** 0.5
            # 密度均匀、无显著离群值 → 保留全部（都是真实子列）
            if std_dev < 0.05 and max_density > 0.5:
                continue

        # 密度低于组内最高密度的 25% → 标记为幽灵列
        threshold = max_density * 0.25
        for c, density in zip(cols, densities):
            if density < threshold:
                phantom_cols.add(c)

    if not phantom_cols:
        return data, 0

    # ---- 5. 所有行同步删除幽灵列 ----
    cols_to_keep = [c for c in range(max_cols) if c not in phantom_cols]
    if not cols_to_keep:
        return data, 0

    result = []
    for row in normalized:
        result.append([row[c] for c in cols_to_keep])

    return result, len(phantom_cols)


def _deduplicate_subset_rows(data):
    """消除 pdf2docx 合并格展开导致的子集重复行（幽灵行）。

    检测模式：相邻两行标签相同，且其中一行的【非空值】是另一行的严格子集。
    较空的那行是合并格展开残留，删掉它，保留数据更完整的行。

    例如：
      Row A: [信用减值损失, (120,700), 空,    空,      空,    空      ]
      Row B: [信用减值损失, (120,700), (136,774), (11.75), 空, (154,535)]
      → A 的非空值 {信用减值损失, (120,700)} ⊂ B 的非空值
      → 删除 A，保留 B

    注意：只检测相邻行，避免跨大范围误删（如"资产总额"和"负债总额"可能共享值）。

    Args:
        data: list[list[str]]

    Returns:
        (cleaned_data, removed_count)
    """
    if not data or len(data) < 2:
        return data, 0

    dedup_count = 0
    rows_to_remove = set()
    max_cols = max(len(r) for r in data)

    for i in range(len(data) - 1):
        if i in rows_to_remove:
            continue

        row_a = data[i]
        row_b = data[i + 1]

        # 条件1：标签相同（第一列）
        label_a = str(row_a[0]).strip() if row_a and row_a[0] else ""
        label_b = str(row_b[0]).strip() if row_b and row_b[0] else ""
        if not label_a or label_a != label_b:
            continue

        # 条件2：逐列收集非空值
        a_nonempty = set()
        b_nonempty = set()

        for c in range(max_cols):
            va = str(row_a[c]).strip() if c < len(row_a) and row_a[c] else ""
            vb = str(row_b[c]).strip() if c < len(row_b) and row_b[c] else ""

            if va:
                a_nonempty.add((c, va))
            if vb:
                b_nonempty.add((c, vb))

        # 用归一化值比较（带列索引，防止不同列同值混淆）
        a_vals = {_normalize_for_lookup(v) for _, v in a_nonempty}
        b_vals = {_normalize_for_lookup(v) for _, v in b_nonempty}

        # A 的所有非空归一化值都在 B 中，且 B 有 A 没有的值 → A 是幽灵行
        if a_vals and a_vals.issubset(b_vals) and (b_vals - a_vals):
            rows_to_remove.add(i)
            dedup_count += 1
            continue

        # B 的所有非空归一化值都在 A 中，且 A 有 B 没有的值 → B 是幽灵行
        if b_vals and b_vals.issubset(a_vals) and (a_vals - b_vals):
            rows_to_remove.add(i + 1)
            dedup_count += 1
            continue

    if rows_to_remove:
        cleaned = [row for idx, row in enumerate(data) if idx not in rows_to_remove]
        return cleaned, dedup_count

    return data, 0


# 汇总行关键词：这些行可能合法地共享上级行的数值（如资产总额=负债总额）
_SUMMARY_KEYWORDS = [
    "总额", "总计", "合计", "净额", "净收入", "净利息",
    "净利", "净收益", "净值", "小计", "余额",
]


def _is_summary_row(first_col_text: str) -> bool:
    """判断首列文本是否属于汇总/合计行，这些行的值可能合法地与上级行一致。"""
    if not first_col_text:
        return False
    for kw in _SUMMARY_KEYWORDS:
        if kw in first_col_text:
            return True
    return False


def _clean_merged_cell_bleed(data):
    """激进清除 pdf2docx 合并格渗透值，后续由 liteparse 补回合法值。

    策略：同一列中，一个值出现在多个行，保留第一次出现（最可能是源头），
    删除后续所有行的该列值。不设保护——汇总行也会被删。
    
    这会产生大量误删，但会由 _restore_from_liteparse 根据 liteparse 原文补回。
    liteparse 是独立解析器，不受 pdf2docx 合并格影响，能准确判断原文中是否存在。
    """
    if len(data) < 2:
        return data, 0

    cleaned_count = 0
    max_cols = max(len(r) for r in data)

    for c in range(1, max_cols):
        first_seen = {}  # {normalized_value: row_index}
        for r in range(len(data)):
            if c >= len(data[r]):
                continue
            val = str(data[r][c]).strip() if data[r][c] else ""
            if not val:
                continue
            norm = _normalize_for_lookup(val)
            if norm not in first_seen:
                first_seen[norm] = r
            else:
                # 后续出现 → 渗透，直接删除
                data[r][c] = ""
                cleaned_count += 1

    return data, cleaned_count


def _normalize_for_lookup(val: str) -> str:
    """归一化数值，用于在 liteparse 文本中查找匹配。

    去掉千分位逗号、空格、括号转负号，统一格式。
    """
    s = val.strip().replace(',', '').replace(' ', '')
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    # 也保留原始千分位格式用于匹配
    return s


def _restore_from_liteparse(table_data, liteparse_text_items, original_data):
    """将激进清洗中误删的值根据 liteparse 原文补回。

    关键：只恢复在 liteparse 中同一行 Y 坐标范围内出现的值。
    避免把其他行（如总生息资产）的合法值误恢复到当前行（如非生息资产）。

    Args:
        table_data: 清洗后的 2D 表格（某些单元格已被清空）
        liteparse_text_items: [{text, x0, y0, x1, y1}, ...] 该页的文本片段
        original_data: 清洗前的原始数据（用于获取被删前的值）

    Returns:
        (restored_data, restore_count)
    """
    if not liteparse_text_items or not original_data:
        return table_data, 0

    # 构建带坐标的 text_items 列表
    items = []
    for ti in liteparse_text_items:
        if isinstance(ti, dict):
            t = ti.get("text", "").strip()
            y0 = ti.get("y0", 0)
            y1 = ti.get("y1", 0)
            if t and y1 > y0:
                items.append({"text": t, "y0": y0, "y1": y1, "y_mid": (y0 + y1) / 2})

    if not items:
        return table_data, 0

    restore_count = 0

    for r in range(min(len(table_data), len(original_data))):
        row = table_data[r]
        orig_row = original_data[r]

        # 1. 通过首列文本定位该行的 Y 坐标范围
        row_label = str(row[0]).strip() if row[0] else ""
        row_y_range = _find_row_y_range(row_label, items)
        if row_y_range is None:
            # 找不到行定位 → 退化为全文搜索
            row_items = items
        else:
            y_min, y_max = row_y_range
            row_items = [
                it for it in items
                if not (it["y1"] < y_min or it["y0"] > y_max)
            ]

        # 2. 对该行每个被清空的单元格，在行范围内的 items 中搜索
        for c in range(min(len(row), len(orig_row))):
            current = str(row[c]).strip() if row[c] else ""
            original = str(orig_row[c]).strip() if orig_row[c] else ""

            if current or not original:
                continue

            norm_orig = _normalize_for_lookup(original)
            if not norm_orig:
                continue

            found = False
            for it in row_items:
                it_norm = _normalize_for_lookup(it["text"])
                if norm_orig == it_norm or norm_orig in it_norm:
                    found = True
                    break

            if found:
                row[c] = original
                restore_count += 1

    return table_data, restore_count


def _find_row_y_range(row_label: str, items: list, threshold: float = 5.0):
    """通过行标签在 TextItems 中定位该行的 Y 坐标范围。

    用 y_mid 聚类：跟标签的 y_mid 差在 threshold 内的 items 视为同行。
    阈值设为 5pt —— 远小于典型行间距（~16-20pt），确保不会误收相邻行。
    """
    if not row_label or not items:
        return None

    label_items = [it for it in items if row_label in it["text"]]
    if not label_items:
        return None

    # 用标签的 y_mid 来定位
    label_y = label_items[0]["y_mid"]
    row_items = [it for it in items if abs(it["y_mid"] - label_y) <= threshold]

    if row_items:
        y_min = min(it["y0"] for it in row_items)
        y_max = max(it["y1"] for it in row_items)
        return (y_min, y_max)

    # 兜底：直接用标签 y_mid ± threshold
    return (label_y - threshold, label_y + threshold)


def _get_liteparse_page(liteparse_data: dict, page_num: int) -> dict | None:
    """从 liteparse ParseResult 字典中获取指定页的数据。"""
    pages = liteparse_data.get("pages", [])
    for p in pages:
        if p.get("page_number") == page_num:
            return p
    return None


def _infer_numeric_pattern(data, skip_col):
    """从同一表格的其他列推断常见数值格式，用于拆分拼接列。

    扫描非跳过列的值，寻找小数/百分比等常见金融数值格式，
    返回一个正则模式字符串给 re.findall 使用。

    Args:
        data: list[list[str]]
        skip_col: 要跳过（不参考）的列索引

    Returns:
        str: 正则模式，如 r'\d+\.\d{2}' 或默认回退模式
    """
    import re

    max_cols = max((len(row) for row in data), default=0)

    for c in range(max_cols):
        if c == skip_col:
            continue
        for row in data:
            val = str(row[c]).strip() if c < len(row) and row[c] else ""
            if not val:
                continue
            # 检测 "1.33" 格式（1-3位小数）
            if re.match(r'^-?\d+\.\d{1,3}$', val):
                return r'-?\d+\.\d{1,3}'
            # 检测 "1.33%" 格式
            if re.match(r'^-?\d+\.\d{1,3}\s*%$', val):
                return r'-?\d+\.\d{1,3}\s*%'
            # 检测 "(1.33)" 格式
            if re.match(r'^\(\d+\.\d{1,3}\)$', val):
                return r'\(\d+\.\d{1,3}\)'
            # 检测负数 "-1.33"
            if re.match(r'^ -\d+\.\d{1,3}$', val):
                return r' -\d+\.\d{1,3}'

    # 回退：宽数值模式（匹配含逗号分隔符、可选百分号、可选括号的数值）
    return r'\(?-?[\d,]+\.?\d*[%％]?\)?'


def _find_reference_row_count(data, skip_col, target_rows):
    """找同一表格中格式正常的"参考列"，用于确定拼接串应拆分成几个独立值。

    参考列条件：与目标列填充了相同数量的数据行，且每行值各不相同。

    Args:
        data: list[list[str]]
        skip_col: 不参考的列索引
        target_rows: 目标列有非空值的行索引列表

    Returns:
        int: 参考列建议的拆分数量；若无合适参考列则返回 len(target_rows)
    """
    max_cols = max((len(row) for row in data), default=0)
    target_set = set(target_rows)

    for c in range(max_cols):
        if c == skip_col:
            continue
        # 收集该列在有值的行索引及值集合
        filled_rows = set()
        values_set = set()
        for r, row in enumerate(data):
            val = str(row[c]).strip() if c < len(row) and row[c] else ""
            if val:
                filled_rows.add(r)
                values_set.add(val)

        # 条件1: 该列填充的行数 >= 目标行数的一半（排除稀疏列）
        # 条件2: 值各不相同（不是被复制的假数据列）
        # 条件3: 填充行与目标行有较高重叠
        overlap = len(filled_rows & target_set)
        if (len(filled_rows) >= len(target_rows) * 0.5
                and len(values_set) >= len(filled_rows) * 0.7
                and overlap >= len(target_rows) * 0.7):
            return len(filled_rows)

    return len(target_rows)


def _try_split_numeric_patterns(concat_str):
    """尝试多种正则模式拆分拼接串，返回拆分后的 token 列表。

    按精确度从高到低依次尝试，每种模式要求 tokens 拼回去等于原串
    （即 join(tokens) == concat_str），确保拆分无遗漏/无重叠。

    Args:
        concat_str: 拼接串

    Returns:
        list[str] or None: 拆分成功返回 token 列表，失败返回 None
    """
    # 模式列表：从精确到宽松
    candidates = [
        r'\d+\.\d{2}',           # 1.33  (两位小数)
        r'\d+\.\d{3}',           # 1.331 (三位小数)
        r'\d+\.\d{2}%',          # 1.33%
        r'\d+\.\d{1,3}\s*%',     # 1.33% / 5.6 %
        r'\(\d+\.\d{1,3}\)',     # (1.33)
        r'-\d+\.\d{1,3}',        # -1.33
        r'\d+\.\d{1,3}',         # 1.33 / 10.5 (通用小数)
        r'[\d,]+\.\d*[%％]?',    # 1,234.56 / 5% (宽数值)
    ]

    import re

    for pattern in candidates:
        tokens = re.findall(pattern, concat_str)
        if len(tokens) <= 1:
            continue
        # 关键验证：拼回去必须等于原串
        if ''.join(tokens) == concat_str:
            return tokens

    return None


def _split_concatenated_column(data):
    """拆分因 pdf2docx 垂直合并格展开导致的拼接值列。

    pdf2docx 在展开垂直跨行合并单元格时，有时会把合并区域内所有行的文本
    拼接成一个长串，然后复制到展开后的每一行对应列中。
    此函数通过多策略检测并拆分为每行一个独立值。

    检测条件：
      1. 某列在所有非空行中值完全一致
      2. 该值长度 > 10 字符（排除普通的共享值/短标签）

    拆分策略（按优先级）：
      策略1: 从同行其他列推断数值格式 → 正则拆分 → 验证 token 数与数据行数匹配
      策略2: 固定模式库逐一尝试 → 验证拆分回原串一致
      策略3: 用参考列确定 N → 尝试等分子串
      全部失败: 保留原值，不修改

    Args:
        data: list[list[str]]

    Returns:
        (cleaned_data, split_count): 拆分后数据 + 成功拆分的列数
    """
    import re

    if not data or len(data) < 2:
        return data, 0

    max_cols = max((len(row) for row in data), default=0)
    if max_cols < 2:
        return data, 0

    split_count = 0

    for c in range(max_cols):
        # ---- 阶段A: 检测拼接列 ----
        value_rows = []        # 该列有非空值的行索引
        first_val = None
        all_same = True

        for r, row in enumerate(data):
            val = str(row[c]).strip() if c < len(row) and row[c] else ""
            if val:
                if first_val is None:
                    first_val = val
                elif val != first_val:
                    all_same = False
                value_rows.append(r)

        if not value_rows or not all_same:
            continue
        if len(first_val) <= 10:
            continue

        n_rows = len(value_rows)
        concat = first_val

        # ---- 阶段B: 策略1 — 从同行列推断格式 ----
        numeric_pattern = _infer_numeric_pattern(data, c)
        tokens = re.findall(numeric_pattern, concat)

        # 验证：非空 token 数量与数据行数匹配
        valid_tokens = [t for t in tokens if t.strip()]
        if len(valid_tokens) == n_rows:
            for idx, r in enumerate(value_rows):
                data[r][c] = valid_tokens[idx]
            split_count += 1
            continue

        # ---- 阶段C: 策略2 — 固定模式库逐一尝试 ----
        tokens = _try_split_numeric_patterns(concat)
        if tokens and len(tokens) == n_rows:
            for idx, r in enumerate(value_rows):
                data[r][c] = tokens[idx]
            split_count += 1
            continue

        # ---- 阶段D: 策略3 — 参考列确定N + 等分子串 ----
        ref_n = _find_reference_row_count(data, c, value_rows)
        if ref_n > 1 and len(concat) % ref_n == 0:
            seg_len = len(concat) // ref_n
            segments = [concat[i:i + seg_len] for i in range(0, len(concat), seg_len)]
            # 每段至少含数字或字母，避免切出纯空白
            if all(any(ch.isdigit() or ch.isalpha() for ch in s) for s in segments):
                for idx, r in enumerate(value_rows[:ref_n]):
                    if idx < len(segments):
                        data[r][c] = segments[idx]
                split_count += 1
                continue

    return data, split_count


def _looks_like_independent_number(s):
    """检查一个字符串是否像一个完整的独立数值。

    合法数值特征：去除千分位逗号、百分号、货币符号后，可通过 float() 解析，
    且小数点不超过 1 个。不含字母、中文等非数值内容。

    用于检测"空格分隔的多数字合并单元格"——合法数据中不会在同一个单元格
    内用空格分隔两个独立数值。
    """
    s = s.strip().strip('%').strip('￥$€').strip()
    if not s:
        return False
    if s.startswith('-'):
        s = s[1:]
    elif s.startswith('+'):
        s = s[1:]
    cleaned = s.replace(',', '')
    if not any(c.isdigit() for c in cleaned):
        return False
    if cleaned.count('.') > 1:
        return False
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _split_merged_numeric_cells(data):
    """检测并拆分因列边界缺失导致的数值合并单元格。

    场景：PDF 表格列间距不明显，列边界检测漏掉中间列分隔线，
    导致相邻两列数值被分配到同一个网格单元格中，如 "453,098,733 407,794,724"。

    检测逻辑：
      - 对每个单元格按空格 split
      - 如果所有子串都像独立数值（≥2个），确认是合并错误
      - 仅当 ALL parts 都是数值时才拆分（保守策略，避免误伤正常文本）

    行列对齐策略（从右向左逐行处理，保证列对齐）：
      1. 找到某行中所有合并单元格的位置
      2. 从最右侧的合并格开始处理（保持索引稳定）
      3. 拆分时优先吸收左侧连续空列（最常见的场景：合并格左侧
         有未检测到列边界的空列，如 ['资产总额', '', 'V1 V2', ...]）
      4. 左侧空列不足以容纳所有子串时，右侧部分原地 shift 展开
      5. 处理完所有行后统一补齐到最大列数

    Returns:
        (cleaned_data, split_count, has_unresolved):
          - cleaned_data: 拆分并重新对齐后的数据
          - split_count: 成功拆分的额外数值个数
          - has_unresolved: True 表示存在无法完美对齐的行（保留原值未拆）
    """
    import copy

    if not data or len(data) < 2:
        return data, 0, False

    result = copy.deepcopy(data)
    split_count = 0
    has_unresolved = False

    # 先确定"基准列数"：取不含合并格的行中最大的非空列数作为对齐目标
    baseline_cols = 0
    for row in result:
        has_any_merge = False
        non_empty = 0
        for c in range(len(row)):
            cell = str(row[c]).strip() if row[c] else ""
            if not cell:
                continue
            non_empty += 1
            if ' ' in cell:
                parts = cell.split()
                if len(parts) >= 2 and all(_looks_like_independent_number(p) for p in parts):
                    has_any_merge = True
        if not has_any_merge and non_empty > baseline_cols:
            baseline_cols = non_empty

    for r in range(len(result)):
        row = result[r]

        # ---- 第1步：找出该行所有合并位置 ----
        merge_list = []  # [(col_index, split_parts)]
        for c in range(len(row)):
            cell_str = str(row[c]).strip() if row[c] else ""
            if not cell_str or ' ' not in cell_str:
                continue
            parts = cell_str.split()
            if len(parts) >= 2 and all(_looks_like_independent_number(p) for p in parts):
                merge_list.append((c, parts))

        if not merge_list:
            continue

        # ---- 第2步：从右向左处理合并格（保持左侧索引稳定） ----
        # 先转为可变列表
        row_work = list(row)

        for c, parts in reversed(merge_list):
            n = len(parts)

            # 统计左侧连续空列数
            left_empty = 0
            for j in range(c - 1, -1, -1):
                if not row_work[j] or str(row_work[j]).strip() == '':
                    left_empty += 1
                else:
                    break

            # 吸收左侧空列：拆分出的前 left_empty 个值向左填入空列
            use_left = min(left_empty, n - 1)

            if use_left > 0:
                # 从 (c - use_left) 开始到 c，依次填入 use_left + 1 个值
                start_pos = c - use_left
                for k in range(use_left + 1):
                    row_work[start_pos + k] = parts[k]

                # 剩余值（如果有）需要向右插入
                remaining = parts[use_left + 1:]
                if remaining:
                    before = row_work[:c + 1]
                    after = row_work[c + 1:]
                    row_work = before + remaining + after
                    # 更新 merge_list 中后续项的列索引（它们还没处理，但在 reversed 中已经处理过了）
                    # 由于是 reversed 迭代，当前 c 右侧的项已经处理过了，无需修正
            else:
                # 没有左侧空列可吸收，直接从当前位置向右展开
                before = row_work[:c]
                after = row_work[c + 1:]
                row_work = before + parts + after

            split_count += n - 1

        # ---- 第3步：对齐检查 ----
        # 如果该行拆分后列数与基准列数差距大，标记为未解决
        if baseline_cols > 0:
            expanded_cols = len(row_work)
            if expanded_cols > baseline_cols + 2:  # 允许 2 列浮动
                has_unresolved = True

        result[r] = row_work

    if split_count == 0:
        return result, 0, False

    # ---- 最终对齐：补齐到最大列数 ----
    max_cols = max((len(row) for row in result), default=0)
    if max_cols > 0:
        for r in range(len(result)):
            while len(result[r]) < max_cols:
                result[r].append("")

    return result, split_count, has_unresolved


def _auto_merge_split_tables(results, liteparse_data=None):
    """检测并合并被 pdf2docx 拆分断裂的相邻表格。
    无论同页还是跨页，只要相邻都参与合并判断。

    两层判定：
    1. liteparse 辅助（同页多表 → region 数量判断）
    2. 启发式规则（列数 + 表头 + 上下文阻断）

    仅合并 data（清洗后），original_data 不动。
    """
    import re

    if not results:
        return results

    is_docx_per_page = results[0].get("extractor") == "docx_per_page"
    results.sort(key=lambda x: (x.get("page", 0), _sort_y_for_page_order(x)))

    # ---- liteparse 辅助的相邻拆分检测 ----
    confident_merges = []
    suggested_merges = []

    if liteparse_data:
        try:
            from codes.table_validator.table_boundary import detect_adjacent_splits
            confident_merges, suggested_merges = detect_adjacent_splits(
                results, liteparse_data
            )
        except Exception as e:
            print(f"  [auto-merge] liteparse 检测失败（降级为纯规则）: {e}")

    # ---- 自动合并高置信对 ----
    to_remove = set()
    if confident_merges:
        # 去重 & 拓扑排序（避免连环合并中的索引错乱）
        merged_set = set()
        for i, j in confident_merges:
            if i in to_remove or j in to_remove:
                continue
            _do_merge(results, i, j)
            to_remove.add(j)
            merged_set.add((i, j))

        if to_remove:
            for j in sorted(to_remove, reverse=True):
                results.pop(j)
            merge_type = "同/跨页" if not is_docx_per_page else "同页"
            print(f"  [auto-merge] liteparse辅助{merge_type}合并: {len(to_remove)}对")

    # ---- 对 suggested merges 打标记 ----
    for i, j, reason in suggested_merges:
        if j >= len(results):
            continue
        results[j]["_suggest_merge_to"] = i
        results[j]["_merge_reason"] = reason
        pa, pb = results[i].get("page", "?"), results[j].get("page", "?")
        cross = "跨页" if pa != pb else "同页"
        print(f"  [建议] {cross} P{pa}→P{pb}: 表格#{j} 🔗建议合并到表格#{i} ({reason})")

    # ---- 启发式规则兜底（仅非 docx_per_page 的跨页合并） ----
    if not is_docx_per_page:
        valid_indices = [i for i, t in enumerate(results) if t.get("data")]

        merged_any = True
        merge_count = 0

        while merged_any:
            merged_any = False
            to_remove_heuristic = set()
            merges_this_round = []

            for idx_i in range(len(valid_indices) - 1):
                i = valid_indices[idx_i]
                j = valid_indices[idx_i + 1]
                if i in to_remove_heuristic or j in to_remove_heuristic:
                    continue

                table_a = results[i]
                table_b = results[j]

                cols_a = max((len(r) for r in table_a["data"]), default=0)
                cols_b = max((len(r) for r in table_b["data"]), default=0)
                if cols_a == 0 or cols_b == 0 or cols_a != cols_b:
                    continue

                # 辅助判断
                def _is_first_on_page(idx):
                    page = results[idx].get("page", 0)
                    for k in valid_indices:
                        if k == idx:
                            break
                        if results[k].get("page") == page and results[k].get("data"):
                            return False
                    return True

                ctx_b = table_b.get("context_text", "")
                meaningful_lines = []
                if ctx_b:
                    meaningful_lines = [l for l in ctx_b.split('\n') if l.strip() and not re.match(r'^[\d\-\s]+$', l.strip())]
                if len(meaningful_lines) >= 2:
                    continue
                if _is_first_on_page(j) and len(meaningful_lines) >= 1:
                    continue

                first_row_b = table_b["data"][0] if table_b["data"] else []
                if not first_row_b or not all(_is_numeric_data(cell) for cell in first_row_b):
                    numeric_count = sum(1 for cell in first_row_b if _is_numeric_data(cell))
                    if numeric_count < max(1, len(first_row_b) * 0.5):
                        continue

                _do_merge(results, i, j)
                to_remove_heuristic.add(j)
                merge_count += 1
                merges_this_round.append((i, j, table_a.get("page", 0), table_b.get("page", 0), len(table_b["data"])))

            if to_remove_heuristic:
                merged_any = True
                for j in sorted(to_remove_heuristic, reverse=True):
                    results.pop(j)
                valid_indices = [i for i, t in enumerate(results) if t.get("data")]

            for i, j, pa, pb, rows in merges_this_round:
                print(f"  [auto-merge] P{pa}+P{pb}: 表格#{i}+#{j} → +{rows}行")

        if merge_count > 0:
            print(f"  [auto-merge] 共合并 {merge_count} 对断裂表格")

    return results


def _do_merge(results, i, j):
    """将 results[j] 合并到 results[i]，保持列数对齐。"""
    table_a = results[i]
    table_b = results[j]
    cols_a = max((len(r) for r in table_a["data"]), default=0)

    for row in table_b["data"]:
        while len(row) < cols_a:
            row.append("")
    table_a["data"].extend([row[:cols_a] for row in table_b["data"]])
    table_a["rows"] = len(table_a["data"])

    if "original_data" in table_b:
        if "original_data" not in table_a:
            table_a["original_data"] = []
        for row in table_b["original_data"]:
            while len(row) < cols_a:
                row.append("")
        table_a["original_data"].extend([row[:cols_a] for row in table_b["original_data"]])


def _precompute_scoped_items(results, liteparse_data):
    """同页多表时，为每表预计算 liteparse region 限定后的 text_items。

    将 scoped items 存入 table["_liteparse_items"]，
    后续 _restore_from_liteparse 和 diff 对比优先使用。
    """
    if not liteparse_data:
        return

    try:
        from codes.table_validator.table_boundary import get_scoped_items_for_table

        # 按页分组
        page_tables: dict = {}
        for table in results:
            if not table.get("data"):
                continue
            page_tables.setdefault(table.get("page", 0), []).append(table)

        for page_num, tables_on_page in page_tables.items():
            if len(tables_on_page) <= 1:
                continue  # 单表不需要限定

            lp_page = _get_liteparse_page(liteparse_data, page_num)
            if not lp_page or not lp_page.get("table_regions"):
                continue

            for table in tables_on_page:
                scoped = get_scoped_items_for_table(
                    table, tables_on_page, lp_page
                )
                # 只有当 scoped 不是全量回退时才打标记
                all_items = lp_page.get("text_items", [])
                if scoped and (len(scoped) != len(all_items)):
                    table["_liteparse_items"] = scoped

            region_count = len(lp_page.get("table_regions", []))
            table_count = len(tables_on_page)
            if region_count != table_count:
                print(f"  [预检] P{page_num}: liteparse {region_count}个区域 vs pdf2docx {table_count}个表"
                      f"（已为各表限定文本范围）")

    except Exception as e:
        print(f"  [预检] scoped items 计算失败（降级为全页模式）: {e}")


def _normalize_table_columns(table_data):
    """规范化表格：所有行补齐到相同列数，剔除首尾全空行。

    独立函数版，供 _auto_clean_tables 等数据处理层调用（无类依赖）。
    """
    if not table_data or not isinstance(table_data, list):
        return table_data

    if len(table_data) == 0:
        return table_data

    max_cols = max((len(row) for row in table_data if row), default=0)

    if max_cols == 0:
        return table_data

    def _is_empty_row(row):
        if not row:
            return True
        return all(cell is None or str(cell).strip() == "" for cell in row)

    normalized = []
    for row in table_data:
        if not row:
            row = []
        while len(row) < max_cols:
            row.append(None)
        row = row[:max_cols]
        normalized.append(row)

    start_idx = 0
    while start_idx < len(normalized) and _is_empty_row(normalized[start_idx]):
        start_idx += 1

    end_idx = len(normalized)
    while end_idx > start_idx and _is_empty_row(normalized[end_idx - 1]):
        end_idx -= 1

    return normalized[start_idx:end_idx]


def _auto_clean_tables(results, progress_callback=None, liteparse_data=None):
    """对解析结果自动清洗：保存原始 → 清洗数值 → 删空格 → 删空行 → 合并断裂。
    
    Args:
        results: 表格列表
        progress_callback: 进度回调
        liteparse_data: liteparse ParseResult.to_dict()，用于激进清洗后恢复合法值
    """
    import copy

    # results 中的 data 字段是纯 Python list[list[str]]，
    # deepcopy 不涉及 PyMuPDF C 对象，无需 gc.collect()

    if progress_callback:
        progress_callback(93, "正在自动清洗数据...")

    total_cleaned_cells = 0
    total_removed_spaces = 0

    # ---- 预处理：同页多表时，为每表预计算 liteparse scoped text_items ----
    _precompute_scoped_items(results, liteparse_data)

    for i, table in enumerate(results):
        if not table.get("data"):
            continue

        # 1. 深拷贝原始数据
        table["original_data"] = copy.deepcopy(table["data"])
        table["is_cleaned"] = True

        # 2. 清洗数值（含中英文保护）
        prev_data = copy.deepcopy(table["data"])
        _clean_data_cells(table["data"])
        for r in range(len(prev_data)):
            for c in range(min(len(prev_data[r]), len(table["data"][r]))):
                if prev_data[r][c] != table["data"][r][c]:
                    total_cleaned_cells += 1

        # 3. 智能删除空格（仅两端空单元格）
        prev_data = copy.deepcopy(table["data"])
        cleaned = _remove_spaces_data(table["data"])
        table["data"] = cleaned
        for r in range(len(prev_data)):
            for c in range(min(len(prev_data[r]), len(cleaned[r]))):
                if prev_data[r][c] != cleaned[r][c]:
                    total_removed_spaces += 1

        # 4. 删除全空行（不影响含省略标记 -- 的行）
        row_before = len(table["data"])
        _remove_empty_rows(table["data"])
        row_after = len(table["data"])
        if row_before != row_after:
            print(f"  [清洗] 表格{i+1} 删除空行: {row_before}→{row_after}行")

        # 4.5 删除因合并单元格展开导致的完全重复相邻列
        table["data"], dup_cols = _deduplicate_columns(table["data"])
        if dup_cols > 0:
            print(f"  [清洗] 表格{i+1} 列去重: 删除 {dup_cols} 个重复列")

        # 4.6 删除因合并单元格展开导致的幽灵列（表头重复 + 数据密度极低）
        table["data"], phantom_cols = _remove_phantom_header_columns(table["data"])
        if phantom_cols > 0:
            print(f"  [清洗] 表格{i+1} 幽灵列删除: {phantom_cols}个列")

        # 5. 删除行内空单元格（紧凑化，打标记 cells_compacted）
        data_before = sum(len(row) for row in table["data"])
        table["data"], compacted = _compact_empty_cells(table["data"])
        if compacted:
            data_after = sum(len(row) for row in table["data"])
            table["cells_compacted"] = True
            print(f"  [清洗] 表格{i+1} 紧凑化: {data_before}→{data_after}个单元格")

        # 5.1 删除稀疏列（数据行填充率 < 15% 的列，用于清理 pdf2docx 检测到的冗余间隔列）
        table["data"], sparse_count = _remove_sparse_columns(table["data"])
        if sparse_count > 0:
            print(f"  [清洗] 表格{i+1} 稀疏列删除: {sparse_count}个列")

        # 5.3 消除幽灵行（子集重复行，合并格展开残留）
        row_before_dedup = len(table["data"])
        table["data"], phantom_count = _deduplicate_subset_rows(table["data"])
        if phantom_count > 0:
            row_after_dedup = len(table["data"])
            print(f"  [清洗] 表格{i+1} 幽灵行去重: {row_before_dedup}→{row_after_dedup}行, 删除{phantom_count}行")

        # 5.5 激进清除 pdf2docx 合并格渗透值（同一列重复出现即删）
        # 后续由 liteparse 补回合法值
        original_before_bleed = copy.deepcopy(table["data"])
        table["data"], bleed_count = _clean_merged_cell_bleed(table["data"])
        if bleed_count > 0:
            print(f"  [清洗] 表格{i+1} 激进清除合并格渗透: {bleed_count}个单元格")

        # 5.6 根据 liteparse 原文补回误删的合法值
        if liteparse_data and bleed_count > 0:
            # 优先使用 scoped text_items（同页多表时限定区域），降级使用全页 items
            text_items = table.get("_liteparse_items")
            if not text_items:
                page_num = table.get("page", 0)
                lp_page = _get_liteparse_page(liteparse_data, page_num)
                if lp_page:
                    text_items = lp_page.get("text_items", [])
            if text_items:
                table["data"], restore_count = _restore_from_liteparse(
                    table["data"], text_items, original_before_bleed
                )
                if restore_count > 0:
                    scope_tag = " (scoped)" if table.get("_liteparse_items") else ""
                    print(f"  [清洗] 表格{i+1} liteparse补回合法值{scope_tag}: {restore_count}个单元格")

        # 5.7 拆分垂直合并格导致的拼接值列
        table["data"], split_count = _split_concatenated_column(table["data"])
        if split_count > 0:
            print(f"  [清洗] 表格{i+1} 拆分拼接列: {split_count}列")

        # 5.75 拆分因列边界缺失导致的数值合并单元格（如 "453,098,733 407,794,724"）
        table["data"], numeric_merge_count, has_unresolved = _split_merged_numeric_cells(table["data"])
        if numeric_merge_count > 0:
            table["_has_merged_numeric_cells"] = True
            if has_unresolved:
                table["_has_unresolved_merged_cells"] = True
            print(f"  [清洗] 表格{i+1} 拆分数值合并单元格: {numeric_merge_count}个数值")

        # 5.8 列规范化：确保清洗后所有行列数一致（修复紧凑化导致的锯齿数组）
        table["data"] = _normalize_table_columns(table["data"])

    # 6. 自动合并断裂表格
    results = _auto_merge_split_tables(results, liteparse_data=liteparse_data)

    # 6.5 合并后再次规范化（合并操作可能产生不同列数的行）
    for table in results:
        if table.get("data"):
            table["data"] = _normalize_table_columns(table["data"])

    if progress_callback:
        progress_callback(94, f"清洗完成({len(results)}个表格, {total_cleaned_cells}个单元格已清洗)")

    return results


# ============================================================
# 工作线程：PDF处理
# ============================================================
class ProcessingWorker(QThread):
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)
    warning = pyqtSignal(str)  # 添加警告信号

    def __init__(self, pdf_path, mode="auto", max_pages=100):
        super().__init__()
        self.pdf_path = pdf_path
        self.mode = mode
        self.max_pages = max_pages
        self.pdf_processor = PDFProcessor()
        self.llm = VisionLLM()

    def run(self):
        context = None
        try:
            from codes.pdf_extractor._log import write_log
            write_log(f"[Worker] 开始解析: {self.pdf_path}, mode={self.mode}")
            self.progress.emit(5, "正在检测PDF类型...")

            # ---- 创建 PDF 共享上下文（一次打开，全流程复用） ----
            print(f"  [Worker] 创建 PDFContext: {self.pdf_path}")
            context = PDFContext(self.pdf_path)
            total_pages = context.page_count
            if self.max_pages:
                total_pages = min(self.max_pages, total_pages)

            is_image = self.pdf_processor.is_image_pdf(context=context)

            results = []
            failed_pages = set()  # 用 set 防止 auto 降级时重复记录同一页
            image_cache_dir = ""
            image_paths = []

            if self.mode == "text_only" or self.mode == "auto":
                # 进度回调 lambda
                cb = lambda v, m: self.progress.emit(v, m)

                if self.mode == "auto" and not is_image:
                    # ---- auto 模式：逐页 pdf2word (页码100%准确) ----
                    self.progress.emit(20, "pdf2word 逐页提取表格...")
                    docx_tables = self.pdf_processor._extract_tables_via_docx_per_page(
                        pdf_path=self.pdf_path,
                        context=context,
                        progress_callback=cb
                    )
                    v2_tables = []  # auto 模式不跑完整 V2，docx 已覆盖
                else:
                    # ---- text_only 模式：纯 V2 ----
                    self.progress.emit(20, "V2提取表格...")
                    v2_tables = self.pdf_processor.extract_text_tables(
                        pdf_path=self.pdf_path,
                        max_pages=self.max_pages,
                        context=context,
                        progress_callback=cb
                    )
                    docx_tables = []

                # ---- 去重合并：docx 优先，V2 补漏 ----
                self.progress.emit(50, "正在去重合并表格...")
                merged_tables = self.pdf_processor._deduplicate_v2_docx(
                    v2_tables, docx_tables
                )
                # 质量过滤：去掉单行/无数字的无效表格
                merged_tables = self.pdf_processor._filter_table_quality(merged_tables)

                for t in merged_tables:
                    ext = t.get("extractor", "v2")
                    data = t.get("data", [])
                    item_type = t.get("type", "text")

                    # 段落条目：跳过表格纠错，用独立格式
                    if item_type == "paragraph":
                        entry = {
                            "page": t["page"],
                            "type": "paragraph",
                            "data": data,  # data 是纯文本字符串
                            "extractor": ext,
                            "parse_status": "success" if data else "empty",
                            "parse_message": "V2段切分离",
                            "columns": 1 if data else 0,
                        }
                        if t.get("bbox"):
                            entry["bbox"] = t["bbox"]
                        results.append(entry)
                        continue

                    # 表格条目：自动纠错 + 标准格式
                    if isinstance(data, list) and len(data) >= 2:
                        corrected = self.pdf_processor.TableAutoCorrector.correct(data)
                        if corrected and len(corrected) != len(data):
                            print(f"  [自动纠错] P{t.get('page')}: {len(data)}行→{len(corrected)}行")
                            data = corrected
                            t["data"] = data
                            t["rows"] = len(corrected)

                    is_docx = ext.startswith("docx")
                    entry = {
                        "page": t["page"],
                        "type": "text",
                        "data": data,
                        "extractor": ext,
                        "parse_status": "success" if data else "failed",
                        "parse_message": "docx通道提取" if (is_docx and data) else ("V2提取" if data else "未检测到表格")
                    }
                    if t.get("context_text"):
                        entry["context_text"] = t["context_text"]
                    results.append(entry)

                extracted_pages = {t["page"] for t in merged_tables}
                for page_num in range(1, total_pages + 1):
                    if page_num not in extracted_pages:
                        failed_pages.add(page_num)

                if self.mode == "auto" and is_image and not v2_tables:
                    print(f"  [auto模式] 文本提取未找到表格，降级到图片处理...")

            if self.mode == "ai_only" or (self.mode == "auto" and is_image and not results):
                # 图片型PDF处理（ai_only 模式，或 auto 模式文本提取失败后的降级）
                self.progress.emit(20, "正在转换为图片...")

                # 使用 context 生成 LLM 图片
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                llm_image_dir = Path(TEMP_DIR) / f"pdf_images_{timestamp}"
                image_paths = context.generate_all_llm_images(llm_image_dir)
                image_cache_dir = str(llm_image_dir)

                if not self.llm.api_key:
                    # 没有API Key时，给每页生成空数据
                    self.warning.emit("未配置API Key，已将PDF转为图片缓存，每页生成空数据")
                    for img_path in image_paths:
                        page_num = int(os.path.basename(img_path).split('_')[1].split('.')[0])
                        results.append({
                            "page": page_num,
                            "type": "ai",
                            "data": [],
                            "parse_status": "empty",
                            "parse_message": "未配置API Key（空数据）"
                        })
                    results.sort(key=lambda x: (x["page"], _sort_y_for_page_order(x)))
                    self.progress.emit(95, "正在整理数据...")

                    # 生成预览图
                    from codes.pdf_extractor.utils import get_pdf_preview_dir
                    preview_dir = get_pdf_preview_dir(self.pdf_path)
                    context.generate_all_previews(preview_dir)

                    context.close()
                    self.finished.emit({
                        "success": True,
                        "is_image_pdf": True,
                        "image_cache_dir": image_cache_dir,
                        "tables": results,
                        "total_tables": len(results),
                        "success_count": 0,
                        "empty_count": len(results),
                        "failed_count": len(results),
                        "total_pages": total_pages
                    })
                    return

                self.progress.emit(40, "正在调用AI识别表格...")
                llm_results = self.llm.batch_recognize(
                    image_paths,
                    progress_callback=lambda x, y: self.progress.emit(
                        40 + int(x / y * 50),
                        f"正在识别第 {x}/{y} 页..."
                    )
                )

                successful_pages = set()
                for res in llm_results:
                    if res["result"]["success"]:
                        table_data = ExcelExporter.parse_json_table(res["result"]["data"])
                        if table_data:
                            results.append({
                                "page": res["page"],
                                "type": "ai",
                                "data": table_data,
                                "parse_status": "success",
                                "parse_message": "AI识别成功"
                            })
                            successful_pages.add(res["page"])
                        else:
                            results.append({
                                "page": res["page"],
                                "type": "ai",
                                "data": [],
                                "parse_status": "failed",
                                "parse_message": "AI返回数据解析失败"
                            })
                            successful_pages.add(res["page"])
                    else:
                        results.append({
                            "page": res["page"],
                            "type": "ai",
                            "data": [],
                            "parse_status": "failed",
                            "parse_message": res["result"].get("error", "AI识别失败")
                        })
                        successful_pages.add(res["page"])

                for page_num in range(1, total_pages + 1):
                    if page_num not in successful_pages:
                        failed_pages.add(page_num)

            for page_num in sorted(failed_pages):
                results.append({
                    "page": page_num,
                    "type": "failed",
                    "data": [],
                    "parse_status": "failed",
                    "parse_message": "未提取到表格数据"
                })

            # 按页码+页内Y坐标排序（保证表格和文本原始数据顺序）
            results.sort(key=lambda x: (x["page"], _sort_y_for_page_order(x)))

            # ---- [旁路] liteparse 通道：全页空间布局文本解析 ----
            self._run_liteparse_side_channel(results, total_pages)

            self.progress.emit(95, "正在整理数据...")

            # 获取图片缓存目录（如果之前没有设置）
            if not image_cache_dir and image_paths:
                image_cache_dir = str(Path(image_paths[0]).parent) if image_paths else ""

            # ---- 统一生成预览图到磁盘（UI 层复用） ----
            from codes.pdf_extractor.utils import get_pdf_preview_dir
            preview_dir = get_pdf_preview_dir(self.pdf_path)
            need_preview = True
            if os.path.isdir(preview_dir):
                cached = [f for f in os.listdir(preview_dir) if f.startswith("preview_")]
                if cached:
                    need_preview = False
            if need_preview:
                self.progress.emit(96, "正在生成预览图...")
                context.generate_all_previews(preview_dir)

            # ---- 自动清洗数据 + 合并断裂表格 ----
            # 尝试加载 liteparse 数据，用于激进清洗后补回合法值
            liteparse_data = None
            try:
                from codes.liteparse_extractor.cache_manager import load_parse_result
                lp_result = load_parse_result(self.pdf_path)
                if lp_result is not None:
                    liteparse_data = lp_result.to_dict()
            except Exception:
                pass

            # 保存真·原始数据（清洗前的提取原样），供追溯
            import copy
            for table in results:
                if table.get("data") and "raw_data" not in table:
                    table["raw_data"] = copy.deepcopy(table["data"])

            results = _auto_clean_tables(
                results,
                progress_callback=lambda v, m: self.progress.emit(v, m),
                liteparse_data=liteparse_data,
            )
            # 最终确保按页码+页内Y坐标排序
            results.sort(key=lambda x: (x["page"], _sort_y_for_page_order(x)))

            # ---- 混合表格分割（liteparse 边界 + pdf2docx 单元格融合） ----
            tables_before_segmentation = copy.deepcopy(results)
            seg_report = None
            liteparse_seg_tables = None
            try:
                if liteparse_data and results:
                    from codes.table_validator.hybrid_segmenter import (
                        hybrid_segment_tables,
                    )
                    from codes.table_validator.liteparse_table_segmenter import (
                        extract_paragraphs_from_liteparse,
                    )
                    self.progress.emit(96, "正在混合分割表格...")
                    # 只取表格类型条目（过滤段落/注解等非表格条目）
                    docx_only = [
                        t for t in results
                        if isinstance(t.get("data"), list) and t.get("type") != "paragraph"
                    ]
                    seg_tables, seg_report = hybrid_segment_tables(
                        liteparse_data,
                        docx_tables=copy.deepcopy(docx_only),
                        enable_cross_page=False,
                    )
                    if seg_tables:
                        liteparse_seg_tables = seg_tables

                        # 为原始表打上质量标记（基于分类器）
                        from codes.table_validator.table_classifier import classify_page
                        for t in tables_before_segmentation:
                            t_data = t.get("data", [])
                            if t_data:
                                cr = classify_page(t_data, t.get("page", 0))
                                t["is_real_table"] = cr.is_real_table
                                t["is_complete"] = cr.is_real_table
                                t["table_category"] = "财务数据表" if cr.is_real_table else "非表格"
                                t["has_header"] = cr.checks.get("has_numeric_col", False)
                                t["has_numeric_data"] = cr.checks.get("has_numeric_col", False)
                            t["segment_source"] = "original"

                        results = seg_tables
                        # 确保分割后的表格按页内 Y 顺序排列
                        results.sort(key=lambda r: (
                            r.get("page", 0), _sort_y_for_page_order(r)
                        ))

                        # 从未被表格覆盖的 liteparse text_items 提取段落
                        try:
                            # 构建兼容格式供 extract_paragraphs_from_liteparse 使用
                            # 混合表格的 data 列数用于检测覆盖范围
                            lp_paragraphs = _extract_paragraphs_for_hybrid(
                                liteparse_data, results
                            )
                            if lp_paragraphs:
                                from codes.table_validator.cell_differ import (
                                    _normalize_for_search,
                                )
                                import re as _re

                                def _entry_text_blob(r: dict) -> str:
                                    return str(
                                        r.get("context_text")
                                        or r.get("data")
                                        or r.get("text")
                                        or ""
                                    ).strip()

                                existing_keys = set()
                                for r in results:
                                    if r.get("type") not in (
                                        "text", "paragraph", "annotation",
                                    ):
                                        continue
                                    blob = _entry_text_blob(r)
                                    if blob:
                                        existing_keys.add((
                                            r.get("page", 0),
                                            _normalize_for_search(blob[:300]),
                                        ))

                                merged_paras = []
                                for para in lp_paragraphs:
                                    blob = _entry_text_blob(para)
                                    if not blob:
                                        continue
                                    key = (
                                        para.get("page", 0),
                                        _normalize_for_search(blob[:300]),
                                    )
                                    if key in existing_keys:
                                        continue
                                    # 与已有文本条目高度重合则跳过
                                    dup = False
                                    norm_new = _normalize_for_search(blob)
                                    parts = _re.findall(
                                        r"[\u4e00-\u9fff]{2,}", norm_new,
                                    )
                                    for r in results:
                                        if r.get("type") not in (
                                            "text", "paragraph", "annotation",
                                        ):
                                            continue
                                        if r.get("page") != para.get("page"):
                                            continue
                                        norm_old = _normalize_for_search(
                                            _entry_text_blob(r),
                                        )
                                        if not norm_old:
                                            continue
                                        if norm_new in norm_old or norm_old in norm_new:
                                            dup = True
                                            break
                                        if parts:
                                            hit = sum(
                                                1 for p in parts if p in norm_old
                                            )
                                            if hit >= 3 and hit / len(parts) >= 0.6:
                                                dup = True
                                                break
                                    if dup:
                                        continue
                                    existing_keys.add(key)
                                    merged_paras.append(para)

                                if merged_paras:
                                    results.extend(merged_paras)
                                    results.sort(key=lambda r: (
                                        r.get("page", 0),
                                        _sort_y_for_page_order(r),
                                    ))
                                    print(
                                        f"  [段切] liteparse 提取到 "
                                        f"{len(merged_paras)} 个文本段落"
                                    )
                        except Exception:
                            pass

                        # 混合融合不需要 extracted_text_entries（pdf2docx cell 不混入注解文本）
                        # 但仍需处理 seg_report 中的段落提取结果
                        extracted_entries = seg_report.get("extracted_text_entries", [])
                        if extracted_entries:
                            new_entries = []
                            existing_texts = {
                                (r.get("page"), r.get("data", "").strip())
                                for r in results
                                if r.get("type") in ("paragraph", "annotation")
                            }
                            seen = set()
                            for entry in extracted_entries:
                                entry_data = entry.get("data", "").strip()
                                if not entry_data:
                                    continue
                                key = (entry.get("page"), entry_data)
                                if key in existing_texts or key in seen:
                                    continue
                                seen.add(key)
                                new_entries.append(entry)
                            if new_entries:
                                results.extend(new_entries)
                                results.sort(key=lambda r: (
                                    r.get("page", 0), _sort_y_for_page_order(r)
                                ))

                        accepted_tables = [t for t in results if t.get("quality_decision") == "accepted"]
                        review_tables = [t for t in results if t.get("quality_decision") == "review"]
                        rejected_tables = [t for t in results if t.get("quality_decision") == "rejected"]
                        real_count = len(accepted_tables)
                        print(f"  [混合分割] 原始 {len(tables_before_segmentation)} 张表 → "
                              f"分割后 {len(results)} 张表, "
                              f"可信 {real_count} 张, 待复核 {len(review_tables)} 张, "
                              f"已拒绝 {len(rejected_tables)} 张")

                        # 跨表去重 V3: 使用统一去重引擎
                        if len(results) > 1:
                            self.progress.emit(97, "正在跨表去重...")
                            dedup_engine = DeduplicationEngine()
                            results = dedup_engine.dedup_adjacent(results)
                            if dedup_engine._debug_log:
                                print(f"  [混合分割去重] {len(dedup_engine._debug_log)} 项")
                                for entry in dedup_engine._debug_log:
                                    print(f"    {entry}")
                    else:
                        print("  [混合分割] 未产生结果，保留原始表格")
            except ImportError:
                print("  [混合分割] hybrid segmenter 模块不可用，跳过")
            except Exception as e:
                import traceback
                print(f"  [混合分割] 分割异常（不影响主流程）: {e}")
                traceback.print_exc()

            # ---- 兜底表格分类（liteparse 不可用或分割失败时，确保每个表都有 category 标记）----
            if not seg_report:
                from codes.table_validator.table_classifier import classify_page
                print("  [兜底分类] liteparse 未产出分割结果，使用基础分类器标记表格类别...")

                def _apply_classify(tables_list: list):
                    for t in tables_list:
                        t_data = t.get("data", [])
                        if t_data:
                            cr = classify_page(t_data, t.get("page", 0))
                            t["is_real_table"] = cr.is_real_table
                            t["is_complete"] = cr.is_real_table
                            t["table_category"] = "财务数据表" if cr.is_real_table else "非表格"
                            t["has_header"] = cr.checks.get("has_numeric_col", False)
                            t["has_numeric_data"] = cr.checks.get("has_numeric_col", False)
                            t["quality_decision"] = "accepted" if cr.is_real_table else "rejected"
                            t["quality_decision_reason"] = cr.reason
                            t["quality_decision_score"] = cr.confidence
                            t["quality_flags"] = cr.checks
                        else:
                            t["table_category"] = "非表格"
                            t["quality_decision"] = "rejected"
                        t["segment_source"] = "original"

                _apply_classify(results)
                _apply_classify(tables_before_segmentation)
                classified_real = sum(1 for t in results if t.get("is_real_table"))
                print(f"  [兜底分类] 完成: {len(results)} 张表, "
                      f"财务数据表 {classified_real} 张, "
                      f"非表格 {len(results) - classified_real} 张")

            self.progress.emit(98, "处理完成")

            # ---- 页面类型标记：纯表格 / 半表格（文本+表格）----
            _mark_page_types(results)
            if tables_before_segmentation:
                _mark_page_types(tables_before_segmentation)

            # ---- 表格→文本去重 V3: 使用统一去重引擎 ----

            # ---- 数据顺序一致性验证 ----
            _verify_results_ordering(results)

            self.finished.emit({
                "success": True,
                "is_image_pdf": is_image,
                "image_cache_dir": image_cache_dir,
                "tables": results,
                "tables_before_segmentation": tables_before_segmentation,
                "liteparse_seg_tables": liteparse_seg_tables,  # liteparse 原始格式（rows[dict]，供导出 CSV）
                "segmentation_report": seg_report if seg_report else {},
                "total_tables": len(results),
                "success_count": len([r for r in results if r.get("quality_decision") == "accepted"]),
                "review_count": len([r for r in results if r.get("quality_decision") == "review"]),
                "rejected_count": len([r for r in results if r.get("quality_decision") == "rejected"]),
                "failed_count": len([r for r in results if r.get("parse_status") == "failed"]),
                "total_pages": total_pages
            })
            from codes.pdf_extractor._log import write_log
            write_log(f"[Worker] 解析完成: {len(results)} 表, {total_pages} 页", "OK")

        except BaseException as e:
            # 用 BaseException 而非 Exception：liteparse 底层 Rust/PyO3
            # 的 PanicException 继承自 BaseException，否则抓不住会闪退
            if isinstance(e, (KeyboardInterrupt, SystemExit)):
                raise
            from codes.pdf_extractor._log import log_exception
            log_exception(f"[Worker] run() 异常: {e}")
            traceback.print_exc()
            self.error.emit(str(e))
        finally:
            if context:
                context.close()

    def _run_liteparse_side_channel(self, results, total_pages):
        """[旁路] liteparse 全页空间布局解析。

        对 PDF 全部页（1..total_pages，受 max_pages 约束）做 liteparse 解析，
        不仅限于 pdf2docx 已识别出表格的页，避免纯文字页丢失坐标文本。
        结果缓存到 data/mid_cache/<pdf>/liteparse/，供混合分割与差分对比使用。

        此步骤失败不阻塞主流程。
        """
        try:
            from codes.liteparse_extractor import LiteParseParser

            if total_pages < 1:
                print("  [liteparse] 无有效页，跳过")
                return

            all_page_numbers = list(range(1, total_pages + 1))

            self.progress.emit(
                93, f"liteparse 正在解析全部 {len(all_page_numbers)} 页...",
            )
            print(f"  [liteparse] 开始全页解析: 1-{total_pages} "
                  f"({len(all_page_numbers)} 页)")

            parser = LiteParseParser()
            result = parser.parse(
                self.pdf_path, target_pages=all_page_numbers,
            )

            print(f"  [liteparse] 完成: "
                  f"{len(result.pages)} 页已解析, "
                  f"{result.page_count_with_table} 页含表格区域, "
                  f"耗时 {result.parse_time_sec:.1f}s")

        except ImportError:
            print("  [liteparse] liteparse 未安装，跳过旁路解析")
            print("  [liteparse] 安装方法: pip install liteparse")
        except BaseException as e:
            # 用 BaseException 而非 Exception，因为 liteparse 底层是 Rust/PyO3，
            # PanicException 继承自 BaseException，不用 Exception 抓不住会闪退
            print(f"  [liteparse] 旁路解析异常（不影响主流程）: {e}")
            from codes.pdf_extractor._log import log_exception
            log_exception(f"[liteparse] 旁路异常: {e}")

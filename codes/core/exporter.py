# -*- coding: utf-8 -*-
"""
PDF表格提取器 - Excel导出模块
"""

import json
from pathlib import Path


class ExcelExporter:
    """Excel导出器"""
    
    @staticmethod
    def export(tables_data, output_path):
        """导出表格到Excel
        
        Args:
            tables_data: 表格数据列表
            output_path: 输出文件路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "财务表格"
        
        # 设置样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_num = 1
        
        for i, table_info in enumerate(tables_data):
            # 添加表格分隔
            if i > 0:
                row_num += 2  # 空行分隔
            
            # 添加标题
            page = table_info.get("page", "?")
            ws.cell(row=row_num, column=1, value=f"第 {page} 页")
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            row_num += 1
            
            # 添加表格数据
            table_data = table_info.get("data", [])
            if not table_data:
                continue
            
            for row in table_data:
                for col_idx, value in enumerate(row):
                    cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
                    cell.border = thin_border
                row_num += 1
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
    
    @staticmethod
    def export_to_json(tables_data, output_path):
        """导出为JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(tables_data, f, ensure_ascii=False, indent=2)
    
    @staticmethod
    def parse_json_table(json_str):
        """解析JSON格式的表格"""
        try:
            return json.loads(json_str)
        except:
            return None
